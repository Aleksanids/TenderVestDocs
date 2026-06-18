"""Safe Excel/CSV registry import for TenderVestDocs user projects."""

from __future__ import annotations

import csv
from dataclasses import dataclass, field
from datetime import datetime, timezone
import json
from pathlib import Path
import re
import shutil
from typing import Any, Iterable, Mapping, Sequence

from openpyxl import load_workbook

from tendervestdocs.domain.procurement import parse_source_raw

from .project_service import CANONICAL_UNKNOWN_PRIORITY_FOLDER, ProjectService
from .source_resolver import SourceRouteResolver


SCHEMA_VERSION = "1.0"
SUPPORTED_EXTENSIONS = {".xlsx", ".csv"}
BAD_PRIORITY_LABELS = {"Без приоритета", "Приоритет не определён"}

DATASET_PROCUREMENT = "procurement"
DATASET_CONTRACT = "contract"
DATASET_UNKNOWN = "unknown"

PROFILE_PROCUREMENT_REGISTRY = "procurement_registry"
PROFILE_CONTRACT_REGISTRY = "contract_registry"
PROFILE_COMBINED_WORKBOOK = "combined_workbook"
PROFILE_UNKNOWN_REQUIRES_MAPPING = "unknown_requires_mapping"

NORMALIZED_REGISTRY_JSON_NAME = "normalized_registry_v1.json"
NORMALIZED_REGISTRY_CSV_NAME = "normalized_registry_v1.csv"
NORMALIZATION_REPORT_JSON_NAME = "import_normalization_report.json"
NORMALIZATION_REPORT_MD_NAME = "import_normalization_report.md"
COLUMN_MAPPING_FOLDER_NAME = "column_mappings"
COLUMN_MAPPING_INDEX_NAME = "index_v1.json"
COLUMN_MAPPING_SCHEMA_VERSION = "tendervestdocs_column_mapping_v1"
MAPPING_REQUIRED_STATUS = "mapping_required"
MAPPING_REQUIRED_ACTION = "open_column_mapping"

NORMALIZED_COLUMNS: tuple[str, ...] = (
    "row_id",
    "row_sequence_number",
    "source_dataset_type",
    "source_file",
    "source_sheet",
    "source_row_number",
    "purchase_number",
    "procurement_number",
    "notice_number",
    "procedure_number",
    "contract_number",
    "contract_registry_number",
    "registry_identifier",
    "law_type",
    "law",
    "source_raw",
    "source_url",
    "contract_url",
    "source_text",
    "source_type",
    "route",
    "route_status",
    "route_confidence",
    "route_reason",
    "normalized_source_url",
    "normalized_domain",
    "normalized_eis_number",
    "normalized_etp_number",
    "normalized_procedure_number",
    "source_warnings",
    "next_action",
    "rostender_number",
    "eis_number",
    "platform_number",
    "source_label",
    "platform_code",
    "platform_name",
    "generated_eis_url",
    "generated_platform_url",
    "priority",
    "customer_name",
    "customer_inn",
    "region",
    "nmck",
    "name",
    "subject",
    "supplier_name",
    "supplier_inn",
    "contract_price",
    "contract_date",
    "execution_status",
    "okpd2",
    "ktru",
    "import_status",
    "import_warnings",
    "raw_values",
    "warnings",
    "errors",
)

FIELD_ALIASES: Mapping[str, tuple[str, ...]] = {
    "purchase_number": (
        "Номер закупки",
        "Номер закупки/извещения",
        "Номер закупки / извещения",
        "Номер извещения",
        "Извещение",
        "Номер ЕИС",
        "Реестровый номер",
        "Закупка",
    ),
    "procurement_number": (
        "Номер",
        "Номер закупки",
        "Номер закупки/извещения",
        "Номер закупки / извещения",
        "Номер извещения",
        "Номер процедуры",
        "procurement_number",
    ),
    "notice_number": (
        "Номер извещения ЕИС",
        "Извещение ЕИС",
        "notice_number",
    ),
    "procedure_number": (
        "Номер процедуры",
        "Номер процедуры на площадке",
        "procedure_number",
    ),
    "contract_number": (
        "Номер контракта",
        "Номер договора",
        "contract_number",
    ),
    "registry_identifier": (
        "Реестровый идентификатор",
        "Идентификатор реестра",
        "registry_identifier",
    ),
    "law_type": (
        "Закон",
        "Закон / тип",
        "Тип закупки",
        "ФЗ",
        "44-ФЗ",
        "223-ФЗ",
        "Коммерция",
        "law_type",
    ),
    "law": (
        "Закон",
        "Закон / тип",
        "Тип закупки",
        "ФЗ",
        "law",
    ),
    "source_raw": (
        "Источник / ссылка на закупку",
        "Источник / ссылка",
        "Ссылка / идентификатор источника",
        "Источник",
        "Адрес закупки",
        "source_raw",
    ),
    "source_url": (
        "Ссылка",
        "URL",
        "Ссылка на закупку",
        "Адрес URL",
        "source_url",
    ),
    "contract_url": (
        "Ссылка на контракт",
        "Ссылка на договор",
        "Ссылка / идентификатор источника",
        "Источник / ссылка на контракт",
        "Источник / ссылка на договор",
        "URL контракта",
        "contract_url",
    ),
    "source_text": (
        "Площадка",
        "ЭТП",
        "Источник текст",
        "source_text",
    ),
    "platform_name": (
        "Площадка",
        "Название площадки",
        "ЭТП",
        "platform_name",
    ),
    "domain": (
        "Домен",
        "Домен источника",
        "domain",
    ),
    "purchase_method": (
        "Способ закупки",
        "Способ определения поставщика",
        "purchase_method",
    ),
    "publication_date": (
        "Дата публикации",
        "Дата размещения",
        "publication_date",
    ),
    "deadline_date": (
        "Срок подачи заявок",
        "Дата окончания подачи заявок",
        "deadline_date",
    ),
    "status": (
        "Статус закупки",
        "Статус",
        "status",
    ),
    "lot_number": (
        "Номер лота",
        "Лот",
        "lot_number",
    ),
    "position_name": (
        "Позиция / товар / работа / услуга",
        "Позиция",
        "Товар",
        "position_name",
    ),
    "quantity": (
        "Количество",
        "quantity",
    ),
    "unit": (
        "Единица измерения",
        "Ед. изм.",
        "unit",
    ),
    "priority": (
        "Приоритет",
        "priority",
    ),
    "customer_name": (
        "Заказчик",
        "Наименование заказчика",
        "customer_name",
    ),
    "customer_inn": (
        "ИНН заказчика",
        "customer_inn",
    ),
    "region": (
        "Регион",
        "Регион поставки / исполнения",
        "Регион поставки",
        "Регион исполнения",
        "region",
    ),
    "nmck": (
        "Начальная цена закупки, руб.",
        "НМЦ закупки, руб.",
        "НМЦ закупки, руб",
        "НМЦК",
        "НМЦ",
        "Начальная цена",
        "Начальная цена закупки",
        "Начальная максимальная цена",
        "Цена закупки",
        "Цена",
        "Сумма",
        "nmck",
    ),
    "name": (
        "Наименование",
        "Название",
        "Наименование закупки / предмет закупки",
        "Предмет закупки",
        "name",
    ),
    "subject": (
        "Наименование закупки / предмет контракта",
        "Предмет контракта",
        "Предмет договора",
        "Поставляемый товар",
        "Товар / модель / аналог",
        "subject",
    ),
    "contract_registry_number": (
        "Реестровый номер контракта",
        "Реестровый номер договора",
        "Номер контракта / договора",
        "Номер контракта",
        "Номер договора",
        "contract_registry_number",
    ),
    "supplier_name": (
        "Поставщик / подрядчик / исполнитель",
        "Поставщик",
        "Подрядчик",
        "Исполнитель",
        "supplier_name",
    ),
    "supplier_inn": (
        "ИНН поставщика / подрядчика / исполнителя",
        "ИНН поставщика",
        "ИНН подрядчика",
        "ИНН исполнителя",
        "supplier_inn",
    ),
    "contract_price": (
        "Цена контракта / договора, руб.",
        "Цена контракта",
        "Цена договора",
        "contract_price",
    ),
    "contract_date": (
        "Дата контракта / договора",
        "Дата контракта",
        "Дата договора",
        "contract_date",
    ),
    "execution_status": (
        "Статус исполнения",
        "Статус исполнения контракта",
        "Статус исполнения договора",
        "execution_status",
    ),
    "contract_status": (
        "Статус контракта",
        "Статус договора",
        "contract_status",
    ),
    "contract_end_date": (
        "Дата окончания",
        "Дата окончания контракта",
        "contract_end_date",
    ),
    "delivery_region": (
        "Регион поставки / исполнения",
        "Регион поставки",
        "delivery_region",
    ),
    "source_comment": (
        "Комментарий по источнику",
        "source_comment",
    ),
    "comment": (
        "Комментарий",
        "Примечание",
        "comment",
    ),
    "okpd2": (
        "ОКПД2",
        "ОКПД 2",
        "Код ОКПД2",
        "okpd2",
    ),
    "ktru": (
        "КТРУ",
        "Код КТРУ",
        "ktru",
    ),
}

CONTRACT_ALIAS_FIELDS = {
    "contract_registry_number",
    "contract_url",
    "supplier_name",
    "supplier_inn",
    "contract_price",
    "contract_date",
    "execution_status",
    "subject",
}


@dataclass(frozen=True)
class ColumnMappingTarget:
    """User-facing target field for Excel/CSV column mapping."""

    field: str
    label: str
    required: bool = False
    requirement_label: str = "опционально"
    dataset_type: str = DATASET_PROCUREMENT
    group: str = "Поля реестра закупок"

    def to_json_dict(self) -> dict[str, str | bool]:
        return {
            "field": self.field,
            "label": self.label,
            "required": self.required,
            "requirement_label": self.requirement_label,
            "dataset_type": self.dataset_type,
            "group": self.group,
        }


COMMON_MAPPING_TARGETS: tuple[ColumnMappingTarget, ...] = (
    ColumnMappingTarget("ignore_column", "Не импортировать эту колонку", False, "всегда доступно", DATASET_PROCUREMENT, "Общие поля"),
    ColumnMappingTarget("source_row_number", "Исходный номер строки", False, "служебное", DATASET_PROCUREMENT, "Общие поля"),
    ColumnMappingTarget("source_sheet", "Исходный лист", False, "служебное", DATASET_PROCUREMENT, "Общие поля"),
    ColumnMappingTarget("source_file", "Исходный файл", False, "служебное", DATASET_PROCUREMENT, "Общие поля"),
    ColumnMappingTarget("comment", "Комментарий / примечание", False, "опционально", DATASET_PROCUREMENT, "Общие поля"),
)

PROCUREMENT_MAPPING_TARGETS: tuple[ColumnMappingTarget, ...] = (
    ColumnMappingTarget("procurement_number", "Номер закупки / извещения / процедуры", True, "обязательно", DATASET_PROCUREMENT),
    ColumnMappingTarget("law", "Закон / тип", True, "обязательно", DATASET_PROCUREMENT),
    ColumnMappingTarget("source_raw", "Исходное значение источника", False, "обязательно: source_raw/source_url", DATASET_PROCUREMENT),
    ColumnMappingTarget("source_url", "Ссылка / URL", False, "обязательно: source_raw/source_url", DATASET_PROCUREMENT),
    ColumnMappingTarget("customer_name", "Заказчик", True, "обязательно", DATASET_PROCUREMENT),
    ColumnMappingTarget("subject", "Предмет закупки", True, "обязательно", DATASET_PROCUREMENT),
    ColumnMappingTarget("nmck", "НМЦК / НМЦД", False, "желательно", DATASET_PROCUREMENT),
    ColumnMappingTarget("priority", "Приоритет", False, "желательно", DATASET_PROCUREMENT),
    ColumnMappingTarget("okpd2", "ОКПД2", False, "желательно", DATASET_PROCUREMENT),
    ColumnMappingTarget("ktru", "КТРУ", False, "желательно", DATASET_PROCUREMENT),
    ColumnMappingTarget("region", "Регион", False, "желательно", DATASET_PROCUREMENT),
    ColumnMappingTarget("platform_name", "Площадка", False, "желательно", DATASET_PROCUREMENT),
    ColumnMappingTarget("domain", "Домен источника", False, "желательно", DATASET_PROCUREMENT),
    ColumnMappingTarget("purchase_method", "Способ закупки", False, "желательно", DATASET_PROCUREMENT),
    ColumnMappingTarget("publication_date", "Дата публикации", False, "желательно", DATASET_PROCUREMENT),
    ColumnMappingTarget("deadline_date", "Срок подачи заявок", False, "желательно", DATASET_PROCUREMENT),
    ColumnMappingTarget("status", "Статус закупки", False, "желательно", DATASET_PROCUREMENT),
    ColumnMappingTarget("lot_number", "Номер лота", False, "желательно", DATASET_PROCUREMENT),
    ColumnMappingTarget("position_name", "Позиция / товар / работа / услуга", False, "желательно", DATASET_PROCUREMENT),
    ColumnMappingTarget("quantity", "Количество", False, "желательно", DATASET_PROCUREMENT),
    ColumnMappingTarget("unit", "Единица измерения", False, "желательно", DATASET_PROCUREMENT),
)

CONTRACT_MAPPING_TARGETS: tuple[ColumnMappingTarget, ...] = (
    ColumnMappingTarget("contract_registry_number", "Реестровый номер контракта / договора", True, "обязательно", DATASET_CONTRACT, "Поля реестра контрактов"),
    ColumnMappingTarget("procurement_number", "Номер закупки / извещения", False, "ключевое", DATASET_CONTRACT, "Поля реестра контрактов"),
    ColumnMappingTarget("contract_url", "Ссылка на контракт / договор / источник", False, "ключевое", DATASET_CONTRACT, "Поля реестра контрактов"),
    ColumnMappingTarget("customer_name", "Заказчик", False, "желательно", DATASET_CONTRACT, "Поля реестра контрактов"),
    ColumnMappingTarget("supplier_name", "Поставщик", False, "желательно", DATASET_CONTRACT, "Поля реестра контрактов"),
    ColumnMappingTarget("supplier_inn", "ИНН поставщика", False, "желательно", DATASET_CONTRACT, "Поля реестра контрактов"),
    ColumnMappingTarget("contract_price", "Цена контракта / договора", False, "желательно", DATASET_CONTRACT, "Поля реестра контрактов"),
    ColumnMappingTarget("subject", "Предмет контракта / договора", False, "желательно", DATASET_CONTRACT, "Поля реестра контрактов"),
    ColumnMappingTarget("contract_date", "Дата контракта", False, "желательно", DATASET_CONTRACT, "Поля реестра контрактов"),
    ColumnMappingTarget("execution_status", "Статус исполнения", False, "желательно", DATASET_CONTRACT, "Поля реестра контрактов"),
    ColumnMappingTarget("contract_status", "Статус контракта", False, "желательно", DATASET_CONTRACT, "Поля реестра контрактов"),
    ColumnMappingTarget("contract_end_date", "Дата окончания", False, "желательно", DATASET_CONTRACT, "Поля реестра контрактов"),
    ColumnMappingTarget("delivery_region", "Регион поставки / исполнения", False, "желательно", DATASET_CONTRACT, "Поля реестра контрактов"),
    ColumnMappingTarget("okpd2", "ОКПД2", False, "желательно", DATASET_CONTRACT, "Поля реестра контрактов"),
    ColumnMappingTarget("ktru", "КТРУ", False, "желательно", DATASET_CONTRACT, "Поля реестра контрактов"),
    ColumnMappingTarget("source_raw", "Исходное значение источника", False, "желательно", DATASET_CONTRACT, "Поля реестра контрактов"),
    ColumnMappingTarget("source_url", "Ссылка / URL", False, "желательно", DATASET_CONTRACT, "Поля реестра контрактов"),
)

SOURCE_MAPPING_TARGETS: tuple[ColumnMappingTarget, ...] = (
    ColumnMappingTarget("source_raw", "Исходное поле источника как в Excel", False, "желательно", DATASET_PROCUREMENT, "Поля источника / площадки"),
    ColumnMappingTarget("source_url", "Прямая ссылка", False, "желательно", DATASET_PROCUREMENT, "Поля источника / площадки"),
    ColumnMappingTarget("platform_name", "Название площадки", False, "желательно", DATASET_PROCUREMENT, "Поля источника / площадки"),
    ColumnMappingTarget("domain", "Домен", False, "желательно", DATASET_PROCUREMENT, "Поля источника / площадки"),
    ColumnMappingTarget("procedure_number", "Номер процедуры на ЭТП", False, "желательно", DATASET_PROCUREMENT, "Поля источника / площадки"),
    ColumnMappingTarget("eis_number", "Номер ЕИС", False, "желательно", DATASET_PROCUREMENT, "Поля источника / площадки"),
    ColumnMappingTarget("etp_number", "Номер ЭТП", False, "желательно", DATASET_PROCUREMENT, "Поля источника / площадки"),
    ColumnMappingTarget("source_comment", "Комментарий по источнику", False, "опционально", DATASET_PROCUREMENT, "Поля источника / площадки"),
)

MAPPING_ONLY_FIELDS = {
    "ignore_column",
    "source_row_number",
    "source_sheet",
    "source_file",
}


class ImportServiceError(Exception):
    """Base error for ImportService operations."""

    code = "import_service_error"

    def __init__(self, message: str, *, code: str | None = None) -> None:
        super().__init__(message)
        if code is not None:
            self.code = code


class UnsupportedInputFileError(ImportServiceError):
    """Raised when an input file extension is not supported."""

    code = "unsupported_file_type"


class ImportValidationError(ImportServiceError):
    """Raised when the import request cannot be validated safely."""

    code = "import_validation_error"


@dataclass(frozen=True)
class ImportRequest:
    """Input model for a registry import operation."""

    project_root: Path
    source_file: Path
    sheet_name: str | None = None


@dataclass(frozen=True)
class ImportWarning:
    """Structured import warning."""

    code: str
    message: str
    source_row_number: int | None = None


@dataclass(frozen=True)
class ImportErrorInfo:
    """Structured import error."""

    code: str
    message: str
    source_row_number: int | None = None


@dataclass(frozen=True)
class ImportRow:
    """Normalized registry row v1."""

    row_id: str
    row_sequence_number: int
    source_row_number: int
    source_dataset_type: str = DATASET_PROCUREMENT
    source_file: str = ""
    source_sheet: str = ""
    purchase_number: str = ""
    procurement_number: str = ""
    notice_number: str = ""
    procedure_number: str = ""
    contract_number: str = ""
    contract_registry_number: str = ""
    registry_identifier: str = ""
    law_type: str = ""
    law: str = ""
    source_url: str = ""
    contract_url: str = ""
    source_raw: str = ""
    source_text: str = ""
    source_type: str = ""
    route: str = ""
    route_status: str = ""
    route_confidence: str = ""
    route_reason: str = ""
    normalized_source_url: str = ""
    normalized_domain: str = ""
    normalized_eis_number: str = ""
    normalized_etp_number: str = ""
    normalized_procedure_number: str = ""
    source_warnings: tuple[str, ...] = ()
    next_action: str = ""
    rostender_number: str = ""
    eis_number: str = ""
    platform_number: str = ""
    source_label: str = ""
    platform_code: str = ""
    platform_name: str = ""
    generated_eis_url: str = ""
    generated_platform_url: str = ""
    priority: str = CANONICAL_UNKNOWN_PRIORITY_FOLDER
    customer_name: str = ""
    customer_inn: str = ""
    region: str = ""
    nmck: str = ""
    name: str = ""
    subject: str = ""
    supplier_name: str = ""
    supplier_inn: str = ""
    contract_price: str = ""
    contract_date: str = ""
    execution_status: str = ""
    okpd2: str = ""
    ktru: str = ""
    import_status: str = "ok"
    import_warnings: tuple[str, ...] = ()
    raw_values: Mapping[str, str] = field(default_factory=dict)
    warnings: tuple[str, ...] = ()
    errors: tuple[str, ...] = ()

    def to_json_dict(self) -> dict[str, Any]:
        return {
            "row_id": self.row_id,
            "row_sequence_number": self.row_sequence_number,
            "source_dataset_type": self.source_dataset_type,
            "source_file": self.source_file,
            "source_sheet": self.source_sheet,
            "source_row_number": self.source_row_number,
            "purchase_number": self.purchase_number,
            "procurement_number": self.procurement_number or self.purchase_number,
            "notice_number": self.notice_number,
            "procedure_number": self.procedure_number,
            "contract_number": self.contract_number,
            "contract_registry_number": self.contract_registry_number,
            "registry_identifier": self.registry_identifier,
            "law_type": self.law_type,
            "law": self.law or self.law_type,
            "source_raw": self.source_raw,
            "source_url": self.source_url,
            "contract_url": self.contract_url,
            "source_text": self.source_text,
            "source_type": self.source_type,
            "route": self.route,
            "route_status": self.route_status,
            "route_confidence": self.route_confidence,
            "route_reason": self.route_reason,
            "normalized_source_url": self.normalized_source_url,
            "normalized_domain": self.normalized_domain,
            "normalized_eis_number": self.normalized_eis_number,
            "normalized_etp_number": self.normalized_etp_number,
            "normalized_procedure_number": self.normalized_procedure_number,
            "source_warnings": list(self.source_warnings),
            "next_action": self.next_action,
            "rostender_number": self.rostender_number,
            "eis_number": self.eis_number,
            "platform_number": self.platform_number,
            "source_label": self.source_label,
            "platform_code": self.platform_code,
            "platform_name": self.platform_name,
            "generated_eis_url": self.generated_eis_url,
            "generated_platform_url": self.generated_platform_url,
            "priority": self.priority,
            "customer_name": self.customer_name,
            "customer_inn": self.customer_inn,
            "region": self.region,
            "nmck": self.nmck,
            "name": self.name,
            "subject": self.subject or self.name,
            "supplier_name": self.supplier_name,
            "supplier_inn": self.supplier_inn,
            "contract_price": self.contract_price,
            "contract_date": self.contract_date,
            "execution_status": self.execution_status,
            "okpd2": self.okpd2,
            "ktru": self.ktru,
            "import_status": self.import_status,
            "import_warnings": list(self.import_warnings or self.warnings),
            "raw_values": dict(self.raw_values),
            "warnings": list(self.warnings),
            "errors": list(self.errors),
        }

    def to_csv_dict(self) -> dict[str, str | int]:
        row = self.to_json_dict()
        return {
            column: (
                json.dumps(row[column], ensure_ascii=False)
                if column in {"raw_values", "import_warnings", "source_warnings", "warnings", "errors"}
                else row[column]
            )
            for column in NORMALIZED_COLUMNS
        }


@dataclass(frozen=True)
class ImportResult:
    """Result model returned by ImportService."""

    status: str
    project_root: Path
    source_file: Path
    copied_source_file: Path
    normalized_registry_json_path: Path
    normalized_registry_csv_path: Path
    normalization_report_json_path: Path
    normalization_report_md_path: Path
    sheet_name: str | None
    file_profile: str
    file_profile_label: str
    sheet_names: tuple[str, ...]
    dataset_counts: Mapping[str, int]
    sheet_profiles: tuple[Mapping[str, Any], ...]
    header_row: int | None
    rows_count: int
    physical_data_rows_count: int
    skipped_empty_rows_count: int
    columns_detected: Mapping[str, str]
    missing_required_columns: tuple[str, ...]
    warnings: tuple[ImportWarning, ...]
    errors: tuple[ImportErrorInfo, ...]
    rows: tuple[ImportRow, ...]
    created_files: tuple[Path, ...]
    source_files: tuple[Path, ...] = field(default_factory=tuple)
    copied_source_files: tuple[Path, ...] = field(default_factory=tuple)


@dataclass(frozen=True)
class ImportPreviewResult:
    """Read-only preview used by the UI field-mapping step."""

    source_file: Path
    sheet_name: str | None
    header_row: int | None
    columns: tuple[str, ...]
    columns_detected: Mapping[str, str]
    suggested_field_mapping: Mapping[str, str]
    preview_rows: tuple[Mapping[str, str], ...]
    sample_values: Mapping[str, tuple[str, ...]] = field(default_factory=dict)
    physical_rows_count: int = 0
    data_rows_count: int = 0
    skipped_rows_count: int = 0
    file_profile: str = "Нестандартный файл"
    profile_code: str = PROFILE_UNKNOWN_REQUIRES_MAPPING
    source_dataset_type: str = DATASET_UNKNOWN
    sheet_profiles: tuple[Mapping[str, Any], ...] = ()
    recognition_status: str = "Нужна настройка сопоставления"
    manual_mapping_required: bool = True
    missing_required_columns: tuple[str, ...] = ()
    warnings: tuple[ImportWarning, ...] = ()
    errors: tuple[ImportErrorInfo, ...] = ()
    source_files: tuple[Path, ...] = field(default_factory=tuple)

    def to_json_dict(self) -> dict[str, Any]:
        return {
            "source_file": str(self.source_file),
            "sheet_name": self.sheet_name,
            "header_row": self.header_row,
            "columns": list(self.columns),
            "columns_detected": dict(self.columns_detected),
            "suggested_field_mapping": dict(self.suggested_field_mapping),
            "preview_rows": [dict(row) for row in self.preview_rows],
            "sample_values": {column: list(values) for column, values in self.sample_values.items()},
            "physical_rows_count": self.physical_rows_count,
            "data_rows_count": self.data_rows_count,
            "skipped_rows_count": self.skipped_rows_count,
            "file_profile": self.file_profile,
            "profile_code": self.profile_code,
            "source_dataset_type": self.source_dataset_type,
            "sheet_profiles": [dict(profile) for profile in self.sheet_profiles],
            "recognition_status": self.recognition_status,
            "manual_mapping_required": self.manual_mapping_required,
            "missing_required_columns": list(self.missing_required_columns),
            "warnings": [warning.__dict__ for warning in self.warnings],
            "errors": [error.__dict__ for error in self.errors],
            "source_files": [str(path) for path in (self.source_files or (self.source_file,))],
        }


@dataclass(frozen=True)
class _TabularData:
    rows: tuple[tuple[str, ...], ...]
    sheet_name: str | None
    source_type: str
    warnings: tuple[ImportWarning, ...] = ()


@dataclass(frozen=True)
class _TableClassification:
    tabular: _TabularData
    profile_code: str
    profile_label: str
    source_dataset_type: str
    header_index: int | None
    columns_by_index: Mapping[int, str]
    columns_detected: Mapping[str, str]
    missing_required_columns: tuple[str, ...]
    score: int = 0


class ImportService:
    """Import .xlsx and .csv registries into an existing user project."""

    def __init__(self, *, project_service: ProjectService | None = None) -> None:
        self.project_service = project_service or ProjectService()
        self.source_resolver = SourceRouteResolver()

    def import_registry(
        self,
        project_root: Path,
        source_file: Path,
        *,
        sheet_name: str | None = None,
        field_mapping: Mapping[str, str] | None = None,
        source_dataset_type: str | None = None,
    ) -> ImportResult:
        """Copy a source file into the project and build normalized registry v1."""

        project = self.project_service.open_project(project_root)
        safe_source_file = self._validate_source_file(source_file)
        copied_source_file = self._copy_source_file(
            safe_source_file,
            project.folders["source_exports"],
        )

        selected_tables, file_profile, file_profile_label = self._select_import_tables(
            self._read_tabular_tables(copied_source_file, sheet_name=sheet_name)
        )
        apply_mapping = field_mapping if len(selected_tables) == 1 else None
        mapped_dataset_type = _dataset_type_for_mapping(
            apply_mapping,
            requested=source_dataset_type,
        )
        warnings: list[ImportWarning] = []
        errors: list[ImportErrorInfo] = []
        normalized_rows: list[ImportRow] = []
        skipped_empty_rows_count = 0
        physical_data_rows_count = 0
        missing_required_columns: list[str] = []
        columns_detected = _merge_columns_detected(selected_tables)

        for selected in selected_tables:
            tabular = selected.tabular
            header_index = selected.header_index
            columns_by_index = dict(selected.columns_by_index)
            selected_columns_detected = dict(selected.columns_detected)
            effective_dataset_type = selected.source_dataset_type
            if selected.profile_code == PROFILE_UNKNOWN_REQUIRES_MAPPING and apply_mapping:
                effective_dataset_type = mapped_dataset_type
            if apply_mapping:
                mapping_header_index = _header_index_for_mapping(
                    tabular.rows,
                    apply_mapping,
                    default=header_index,
                )
                if mapping_header_index != header_index:
                    columns_by_index = {}
                    selected_columns_detected = {}
                header_index = mapping_header_index

            mapping_warnings: tuple[ImportWarning, ...] = ()
            if apply_mapping:
                columns_by_index, selected_columns_detected, mapping_warnings = self._apply_field_mapping(
                    rows=tabular.rows,
                    header_index=header_index,
                    columns_by_index=columns_by_index,
                    columns_detected=selected_columns_detected,
                    field_mapping=apply_mapping,
                )

            selected_missing = _critical_missing_required_fields(
                effective_dataset_type,
                selected_columns_detected,
            )
            for missing_field in selected_missing:
                marker = (
                    missing_field
                    if len(selected_tables) == 1
                    else f"{selected.source_dataset_type}.{missing_field}"
                )
                if marker not in missing_required_columns:
                    missing_required_columns.append(marker)

            warnings.extend(tabular.warnings)
            warnings.extend(mapping_warnings)
            if selected.profile_code == PROFILE_UNKNOWN_REQUIRES_MAPPING and not apply_mapping:
                warnings.append(
                    ImportWarning(
                        code=PROFILE_UNKNOWN_REQUIRES_MAPPING,
                        message="Профиль файла не определен уверенно; требуется будущая настройка сопоставления.",
                    )
                )
            for missing_field in selected_missing:
                warnings.append(
                    ImportWarning(
                        code="missing_required_column",
                        message=f"Не найдена обязательная колонка: {missing_field}.",
                    )
                )

            table_rows, table_skipped_empty_rows_count = self._build_rows(
                rows=tabular.rows,
                header_index=header_index,
                columns_by_index=columns_by_index,
                source_type=tabular.source_type,
                source_dataset_type=effective_dataset_type,
                source_name=copied_source_file.name,
                source_sheet=tabular.sheet_name or "",
                row_id_offset=len(normalized_rows),
            )
            normalized_rows.extend(table_rows)
            skipped_empty_rows_count += table_skipped_empty_rows_count
            physical_data_rows_count += _count_physical_data_rows(tabular.rows, header_index)

        for row in normalized_rows:
            for warning_code in row.warnings:
                warnings.append(
                    ImportWarning(
                        code=warning_code,
                        message=self._warning_message(warning_code),
                        source_row_number=row.source_row_number,
                    )
                )

        if field_mapping and len(selected_tables) > 1:
            warnings.append(
                ImportWarning(
                    code="field_mapping_skipped_for_combined_workbook",
                    message="Ручное сопоставление не применено к книге с несколькими распознанными листами.",
                )
            )

        if not normalized_rows:
            warnings.append(
                ImportWarning(
                    code="empty_registry",
                    message="После чтения источника не найдено строк реестра.",
                )
            )

        sheet_names = tuple(table.tabular.sheet_name or "CSV" for table in selected_tables)
        sheet_name_result = (
            sheet_names[0]
            if len(sheet_names) == 1
            else "; ".join(sheet_names)
        ) if sheet_names else None
        first_selected = selected_tables[0] if selected_tables else None
        header_row = (
            None
            if first_selected is None or first_selected.header_index is None
            else first_selected.header_index + 1
        )
        dataset_counts = _count_rows_by_dataset(normalized_rows)
        sheet_profiles = _build_sheet_profiles(selected_tables, normalized_rows)

        registry_json_path = self._next_available_path(
            project.folders["clean_registers"] / NORMALIZED_REGISTRY_JSON_NAME,
        )
        registry_csv_path = self._next_available_path(
            project.folders["clean_registers"] / NORMALIZED_REGISTRY_CSV_NAME,
        )
        report_json_path = self._next_available_path(
            project.folders["app_working_docs"] / NORMALIZATION_REPORT_JSON_NAME,
        )
        report_md_path = self._next_available_path(
            project.folders["app_working_docs"] / NORMALIZATION_REPORT_MD_NAME,
        )

        created_files = (
            copied_source_file,
            registry_json_path,
            registry_csv_path,
            report_json_path,
            report_md_path,
        )
        result = ImportResult(
            status="ok_with_warnings" if warnings or errors else "ok",
            project_root=project.project_root,
            source_file=safe_source_file,
            copied_source_file=copied_source_file,
            normalized_registry_json_path=registry_json_path,
            normalized_registry_csv_path=registry_csv_path,
            normalization_report_json_path=report_json_path,
            normalization_report_md_path=report_md_path,
            sheet_name=sheet_name_result,
            file_profile=file_profile,
            file_profile_label=file_profile_label,
            sheet_names=sheet_names,
            dataset_counts=dataset_counts,
            sheet_profiles=sheet_profiles,
            header_row=header_row,
            rows_count=len(normalized_rows),
            physical_data_rows_count=physical_data_rows_count,
            skipped_empty_rows_count=skipped_empty_rows_count,
            columns_detected=columns_detected,
            missing_required_columns=tuple(missing_required_columns),
            warnings=tuple(warnings),
            errors=tuple(errors),
            rows=tuple(normalized_rows),
            created_files=created_files,
            source_files=(safe_source_file,),
            copied_source_files=(copied_source_file,),
        )

        self._write_registry_json(result)
        self._write_registry_csv(result)
        self._write_normalization_report(result)
        return result

    def import_registries(
        self,
        project_root: Path,
        source_files: Sequence[Path],
    ) -> ImportResult:
        """Import separate registry workbooks as one normalized registry."""

        source_files_tuple = tuple(Path(item) for item in source_files)
        if not source_files_tuple:
            raise ImportValidationError(
                "Файлы для импорта не переданы.",
                code="source_files_missing",
            )
        if len(source_files_tuple) == 1:
            return self.import_registry(project_root, source_files_tuple[0])

        project = self.project_service.open_project(project_root)
        safe_source_files = tuple(self._validate_source_file(source_file) for source_file in source_files_tuple)
        copied_source_files = tuple(
            self._copy_source_file(safe_source_file, project.folders["source_exports"])
            for safe_source_file in safe_source_files
        )

        selected_groups: list[tuple[Path, tuple[_TableClassification, ...]]] = []
        group_profiles: list[tuple[str, str]] = []
        for copied_source_file in copied_source_files:
            selected_tables, profile_code, profile_label = self._select_import_tables(
                self._read_tabular_tables(copied_source_file, sheet_name=None)
            )
            selected_groups.append((copied_source_file, selected_tables))
            group_profiles.append((profile_code, profile_label))

        selected_tables = tuple(
            selected
            for _, group_tables in selected_groups
            for selected in group_tables
        )
        dataset_types = {
            selected.source_dataset_type
            for selected in selected_tables
            if selected.source_dataset_type != DATASET_UNKNOWN
        }
        if DATASET_PROCUREMENT in dataset_types and DATASET_CONTRACT in dataset_types:
            file_profile = PROFILE_COMBINED_WORKBOOK
            file_profile_label = "Несколько Excel: закупки + контракты"
        elif any(profile_code == PROFILE_UNKNOWN_REQUIRES_MAPPING for profile_code, _ in group_profiles):
            file_profile = PROFILE_UNKNOWN_REQUIRES_MAPPING
            file_profile_label = "Нестандартные файлы"
        elif len(group_profiles) == 1:
            file_profile, file_profile_label = group_profiles[0]
        else:
            file_profile = PROFILE_COMBINED_WORKBOOK
            file_profile_label = "Несколько Excel-таблиц"

        warnings: list[ImportWarning] = []
        errors: list[ImportErrorInfo] = []
        normalized_rows: list[ImportRow] = []
        skipped_empty_rows_count = 0
        physical_data_rows_count = 0
        missing_required_columns: list[str] = []
        columns_detected = _merge_columns_detected(selected_tables)

        for copied_source_file, group_tables in selected_groups:
            for selected in group_tables:
                tabular = selected.tabular
                header_index = selected.header_index
                columns_by_index = dict(selected.columns_by_index)
                selected_columns_detected = dict(selected.columns_detected)
                effective_dataset_type = selected.source_dataset_type

                selected_missing = _critical_missing_required_fields(
                    effective_dataset_type,
                    selected_columns_detected,
                )
                for missing_field in selected_missing:
                    marker = f"{copied_source_file.name}.{selected.source_dataset_type}.{missing_field}"
                    if marker not in missing_required_columns:
                        missing_required_columns.append(marker)

                warnings.extend(tabular.warnings)
                if selected.profile_code == PROFILE_UNKNOWN_REQUIRES_MAPPING:
                    warnings.append(
                        ImportWarning(
                            code=PROFILE_UNKNOWN_REQUIRES_MAPPING,
                            message="Профиль одного из файлов не определен уверенно; требуется настройка сопоставления.",
                        )
                    )
                for missing_field in selected_missing:
                    warnings.append(
                        ImportWarning(
                            code="missing_required_column",
                            message=f"Не найдена обязательная колонка: {missing_field}.",
                        )
                    )

                table_rows, table_skipped_empty_rows_count = self._build_rows(
                    rows=tabular.rows,
                    header_index=header_index,
                    columns_by_index=columns_by_index,
                    source_type=tabular.source_type,
                    source_dataset_type=effective_dataset_type,
                    source_name=copied_source_file.name,
                    source_sheet=tabular.sheet_name or "",
                    row_id_offset=len(normalized_rows),
                )
                normalized_rows.extend(table_rows)
                skipped_empty_rows_count += table_skipped_empty_rows_count
                physical_data_rows_count += _count_physical_data_rows(tabular.rows, header_index)

        for row in normalized_rows:
            for warning_code in row.warnings:
                warnings.append(
                    ImportWarning(
                        code=warning_code,
                        message=self._warning_message(warning_code),
                        source_row_number=row.source_row_number,
                    )
                )

        if not normalized_rows:
            warnings.append(
                ImportWarning(
                    code="empty_registry",
                    message="После чтения источников не найдено строк реестра.",
                )
            )

        sheet_names = tuple(table.tabular.sheet_name or "CSV" for table in selected_tables)
        sheet_name_result = "; ".join(sheet_names) if sheet_names else None
        first_selected = selected_tables[0] if selected_tables else None
        header_row = (
            None
            if first_selected is None or first_selected.header_index is None
            else first_selected.header_index + 1
        )
        dataset_counts = _count_rows_by_dataset(normalized_rows)
        sheet_profiles = _build_sheet_profiles(selected_tables, normalized_rows)

        registry_json_path = self._next_available_path(
            project.folders["clean_registers"] / NORMALIZED_REGISTRY_JSON_NAME,
        )
        registry_csv_path = self._next_available_path(
            project.folders["clean_registers"] / NORMALIZED_REGISTRY_CSV_NAME,
        )
        report_json_path = self._next_available_path(
            project.folders["app_working_docs"] / NORMALIZATION_REPORT_JSON_NAME,
        )
        report_md_path = self._next_available_path(
            project.folders["app_working_docs"] / NORMALIZATION_REPORT_MD_NAME,
        )

        created_files = (
            *copied_source_files,
            registry_json_path,
            registry_csv_path,
            report_json_path,
            report_md_path,
        )
        result = ImportResult(
            status="ok_with_warnings" if warnings or errors else "ok",
            project_root=project.project_root,
            source_file=safe_source_files[0],
            copied_source_file=copied_source_files[0],
            normalized_registry_json_path=registry_json_path,
            normalized_registry_csv_path=registry_csv_path,
            normalization_report_json_path=report_json_path,
            normalization_report_md_path=report_md_path,
            sheet_name=sheet_name_result,
            file_profile=file_profile,
            file_profile_label=file_profile_label,
            sheet_names=sheet_names,
            dataset_counts=dataset_counts,
            sheet_profiles=sheet_profiles,
            header_row=header_row,
            rows_count=len(normalized_rows),
            physical_data_rows_count=physical_data_rows_count,
            skipped_empty_rows_count=skipped_empty_rows_count,
            columns_detected=columns_detected,
            missing_required_columns=tuple(missing_required_columns),
            warnings=tuple(warnings),
            errors=tuple(errors),
            rows=tuple(normalized_rows),
            created_files=created_files,
            source_files=safe_source_files,
            copied_source_files=copied_source_files,
        )

        self._write_registry_json(result)
        self._write_registry_csv(result)
        self._write_normalization_report(result)
        return result

    def preview_registry(
        self,
        source_file: Path,
        *,
        sheet_name: str | None = None,
    ) -> ImportPreviewResult:
        """Read a source file without copying or writing project artifacts."""

        safe_source_file = self._validate_source_file(source_file)
        selected_tables, profile_code, profile_label = self._select_import_tables(
            self._read_tabular_tables(safe_source_file, sheet_name=sheet_name)
        )
        selected = selected_tables[0] if selected_tables else None
        tabular = selected.tabular if selected is not None else _TabularData(
            rows=(),
            sheet_name=None,
            source_type="",
        )
        header_index = selected.header_index if selected is not None else None
        columns_by_index = selected.columns_by_index if selected is not None else {}
        columns_detected = selected.columns_detected if selected is not None else {}
        if header_index is None:
            header_index = _first_non_empty_row_index(tabular.rows)
        header_row = tabular.rows[header_index] if header_index is not None else ()
        data_rows_count = 0
        skipped_rows_count = 0
        for table in selected_tables:
            table_data_count, table_skipped_count = _count_keyed_data_rows(
                rows=table.tabular.rows,
                header_index=table.header_index,
                columns_by_index=table.columns_by_index,
                source_dataset_type=table.source_dataset_type,
            )
            data_rows_count += table_data_count
            skipped_rows_count += table_skipped_count
        missing_required_columns = _preview_missing_required_fields(
            selected_tables,
            selected=selected,
            profile_code=profile_code,
            columns_detected=columns_detected,
        )
        manual_mapping_required = bool(missing_required_columns) or profile_code == PROFILE_UNKNOWN_REQUIRES_MAPPING
        recognition_status = "Файл распознан" if not manual_mapping_required else "Нужна настройка сопоставления"

        preview_rows: list[Mapping[str, str]] = []
        if header_index is not None:
            for row in tabular.rows[header_index + 1 : header_index + 6]:
                if not _has_any_value(row):
                    continue
                preview_rows.append(
                    {
                        _to_text(header_row[index]) or f"column_{index + 1}": _sanitize_output_value(value)
                        for index, value in enumerate(row)
                    }
                )

        return ImportPreviewResult(
            source_file=safe_source_file,
            sheet_name=tabular.sheet_name,
            header_row=None if header_index is None else header_index + 1,
            columns=tuple(_to_text(value) for value in header_row if _to_text(value)),
            columns_detected=columns_detected,
            suggested_field_mapping={field: header for field, header in columns_detected.items()},
            preview_rows=tuple(preview_rows),
            sample_values=_sample_values_for_columns(
                tuple(_to_text(value) for value in header_row if _to_text(value)),
                tuple(preview_rows),
            ),
            physical_rows_count=_count_physical_data_rows(tabular.rows, header_index),
            data_rows_count=data_rows_count,
            skipped_rows_count=skipped_rows_count,
            file_profile=profile_label,
            profile_code=profile_code,
            source_dataset_type=selected.source_dataset_type if selected is not None else DATASET_UNKNOWN,
            sheet_profiles=tuple(_classification_to_profile(table) for table in selected_tables),
            recognition_status=recognition_status,
            manual_mapping_required=manual_mapping_required,
            missing_required_columns=missing_required_columns,
            warnings=tuple(warning for table in selected_tables for warning in table.tabular.warnings),
            errors=(),
            source_files=(safe_source_file,),
        )

    def preview_registries(
        self,
        source_files: Sequence[Path],
    ) -> ImportPreviewResult:
        """Preview separate registry workbooks as one import source."""

        source_files_tuple = tuple(Path(item) for item in source_files)
        if not source_files_tuple:
            raise ImportValidationError(
                "Файлы для предварительной проверки не переданы.",
                code="source_files_missing",
            )
        if len(source_files_tuple) == 1:
            return self.preview_registry(source_files_tuple[0])

        safe_source_files = tuple(self._validate_source_file(source_file) for source_file in source_files_tuple)
        selected_groups: list[tuple[Path, tuple[_TableClassification, ...], str, str]] = []
        for safe_source_file in safe_source_files:
            selected_tables, profile_code, profile_label = self._select_import_tables(
                self._read_tabular_tables(safe_source_file, sheet_name=None)
            )
            selected_groups.append((safe_source_file, selected_tables, profile_code, profile_label))

        selected_tables = tuple(
            selected
            for _, group_tables, _, _ in selected_groups
            for selected in group_tables
        )
        selected = selected_tables[0] if selected_tables else None
        tabular = selected.tabular if selected is not None else _TabularData(
            rows=(),
            sheet_name=None,
            source_type="",
        )
        header_index = selected.header_index if selected is not None else None
        if header_index is None:
            header_index = _first_non_empty_row_index(tabular.rows)
        header_row = tabular.rows[header_index] if header_index is not None else ()

        data_rows_count = 0
        skipped_rows_count = 0
        for table in selected_tables:
            table_data_count, table_skipped_count = _count_keyed_data_rows(
                rows=table.tabular.rows,
                header_index=table.header_index,
                columns_by_index=table.columns_by_index,
                source_dataset_type=table.source_dataset_type,
            )
            data_rows_count += table_data_count
            skipped_rows_count += table_skipped_count

        dataset_types = {
            table.source_dataset_type
            for table in selected_tables
            if table.source_dataset_type != DATASET_UNKNOWN
        }
        if DATASET_PROCUREMENT in dataset_types and DATASET_CONTRACT in dataset_types:
            profile_code = PROFILE_COMBINED_WORKBOOK
            profile_label = "Несколько Excel: закупки + контракты"
        elif any(group_profile == PROFILE_UNKNOWN_REQUIRES_MAPPING for _, _, group_profile, _ in selected_groups):
            profile_code = PROFILE_UNKNOWN_REQUIRES_MAPPING
            profile_label = "Нестандартные файлы"
        else:
            profile_code = selected_groups[0][2] if selected_groups else PROFILE_UNKNOWN_REQUIRES_MAPPING
            profile_label = selected_groups[0][3] if selected_groups else "Нестандартный файл"

        missing_required_columns = _preview_missing_required_fields(
            selected_tables,
            selected=selected,
            profile_code=profile_code,
            columns_detected=selected.columns_detected if selected is not None else {},
        )
        manual_mapping_required = bool(missing_required_columns) or profile_code == PROFILE_UNKNOWN_REQUIRES_MAPPING
        recognition_status = "Файлы распознаны" if not manual_mapping_required else "Нужна настройка сопоставления"

        preview_rows: list[Mapping[str, str]] = []
        if header_index is not None:
            for row in tabular.rows[header_index + 1 : header_index + 6]:
                if not _has_any_value(row):
                    continue
                preview_rows.append(
                    {
                        _to_text(header_row[index]) or f"column_{index + 1}": _sanitize_output_value(value)
                        for index, value in enumerate(row)
                    }
                )

        sheet_names = tuple(table.tabular.sheet_name or "CSV" for table in selected_tables)
        return ImportPreviewResult(
            source_file=safe_source_files[0],
            sheet_name="; ".join(sheet_names) if sheet_names else None,
            header_row=None if header_index is None else header_index + 1,
            columns=tuple(_to_text(value) for value in header_row if _to_text(value)),
            columns_detected=_merge_columns_detected(selected_tables),
            suggested_field_mapping={
                field: header
                for field, header in (selected.columns_detected if selected is not None else {}).items()
            },
            preview_rows=tuple(preview_rows),
            sample_values=_sample_values_for_columns(
                tuple(_to_text(value) for value in header_row if _to_text(value)),
                tuple(preview_rows),
            ),
            physical_rows_count=sum(
                _count_physical_data_rows(table.tabular.rows, table.header_index)
                for table in selected_tables
            ),
            data_rows_count=data_rows_count,
            skipped_rows_count=skipped_rows_count,
            file_profile=profile_label,
            profile_code=profile_code,
            source_dataset_type=DATASET_UNKNOWN if len(dataset_types) != 1 else next(iter(dataset_types)),
            sheet_profiles=tuple(_classification_to_profile(table) for table in selected_tables),
            recognition_status=recognition_status,
            manual_mapping_required=manual_mapping_required,
            missing_required_columns=missing_required_columns,
            warnings=tuple(warning for table in selected_tables for warning in table.tabular.warnings),
            errors=(),
            source_files=safe_source_files,
        )

    def column_mapping_settings_path(self, project_root: Path) -> Path:
        """Return the project-scoped mapping settings path."""

        project = self.project_service.open_project(project_root)
        return project.folders["app_working_docs"] / COLUMN_MAPPING_FOLDER_NAME / COLUMN_MAPPING_INDEX_NAME

    def load_column_mappings(self, project_root: Path) -> dict[str, Any]:
        """Load project-scoped column mappings without failing on a missing file."""

        settings_path = self.column_mapping_settings_path(project_root)
        if not settings_path.exists():
            return {
                "schema_version": COLUMN_MAPPING_SCHEMA_VERSION,
                "mappings": [],
            }
        try:
            data = json.loads(settings_path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            return {
                "schema_version": COLUMN_MAPPING_SCHEMA_VERSION,
                "mappings": [],
            }
        mappings = data.get("mappings") if isinstance(data, Mapping) else []
        return {
            "schema_version": str(data.get("schema_version") or COLUMN_MAPPING_SCHEMA_VERSION) if isinstance(data, Mapping) else COLUMN_MAPPING_SCHEMA_VERSION,
            "mappings": [dict(item) for item in mappings if isinstance(item, Mapping)],
        }

    def save_column_mapping(
        self,
        project_root: Path,
        *,
        dataset_type: str,
        field_mapping: Mapping[str, str],
        source_columns: Iterable[str],
        sample_values: Mapping[str, Iterable[str]] | None = None,
        source_file: Path | str = "",
        sheet_name: str | None = None,
        header_row: int | None = None,
    ) -> dict[str, Any]:
        """Persist a reusable mapping inside the active user project."""

        clean_mapping = _clean_field_mapping(field_mapping)
        column_mapping = _column_to_field_mapping(_clean_field_mapping_for_storage(field_mapping))
        ignored_columns = [
            source_column
            for field_name, source_column in field_mapping.items()
            if _to_text(field_name) == "ignore_column" and _to_text(source_column)
        ]
        validation_errors = validate_column_mapping(dataset_type, clean_mapping)
        if validation_errors:
            raise ImportValidationError(
                "; ".join(validation_errors),
                code="column_mapping_validation_error",
            )

        settings_path = self.column_mapping_settings_path(project_root)
        settings_path.parent.mkdir(parents=True, exist_ok=True)
        current = self.load_column_mappings(project_root)
        now = datetime.now(timezone.utc).isoformat()
        source_columns_tuple = tuple(_to_text(column) for column in source_columns if _to_text(column))
        mapping_id = _mapping_id(dataset_type, source_columns_tuple)
        mapping_file_name = _mapping_file_name(source_file, sheet_name)
        mapping_file_path = settings_path.parent / mapping_file_name
        entry = {
            "mapping_id": mapping_id,
            "dataset_type": _normalize_dataset_type(dataset_type),
            "mapping_file": str(mapping_file_path),
            "source_file": str(source_file),
            "source_file_name": Path(str(source_file)).name if source_file else "",
            "sheet_name": sheet_name,
            "detected_profile": PROFILE_UNKNOWN_REQUIRES_MAPPING,
            "header_row": header_row,
            "source_columns": list(source_columns_tuple),
            "sample_values": {
                _to_text(column): [str(value) for value in values if _to_text(value)][:3]
                for column, values in (sample_values or {}).items()
            },
            "mapping": column_mapping,
            "field_mapping": clean_mapping,
            "ignored_columns": ignored_columns,
            "required_fields_status": _required_fields_status(dataset_type, clean_mapping),
            "target_fields": [
                target.to_json_dict()
                for target in mapping_targets_for_dataset(dataset_type)
            ],
            "created_at": now,
            "updated_at": now,
        }

        mappings = [
            item
            for item in current.get("mappings", [])
            if item.get("mapping_id") != mapping_id
        ]
        for old in current.get("mappings", []):
            if old.get("mapping_id") == mapping_id and old.get("created_at"):
                entry["created_at"] = old["created_at"]
                break
        mappings.append(entry)
        settings = {
            "schema_version": COLUMN_MAPPING_SCHEMA_VERSION,
            "mappings": mappings,
        }
        settings_path.write_text(
            json.dumps(settings, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )
        mapping_file_path.write_text(
            json.dumps(
                {
                    "schema_version": COLUMN_MAPPING_SCHEMA_VERSION,
                    "source_file": entry["source_file"],
                    "source_sheet": sheet_name,
                    "detected_profile": entry["detected_profile"],
                    "header_row": header_row,
                    "mapping": column_mapping,
                    "field_mapping": clean_mapping,
                    "ignored_columns": ignored_columns,
                    "required_fields_status": entry["required_fields_status"],
                    "created_at": entry["created_at"],
                    "updated_at": entry["updated_at"],
                },
                ensure_ascii=False,
                indent=2,
            )
            + "\n",
            encoding="utf-8",
        )
        return entry

    def find_reusable_column_mapping(
        self,
        project_root: Path,
        *,
        dataset_type: str,
        source_columns: Iterable[str],
    ) -> dict[str, str] | None:
        """Find the latest saved mapping compatible with the current source columns."""

        settings = self.load_column_mappings(project_root)
        normalized_columns = {_normalize_header(column) for column in source_columns if _to_text(column)}
        candidates = [
            item
            for item in settings.get("mappings", [])
            if _normalize_dataset_type(str(item.get("dataset_type") or "")) == _normalize_dataset_type(dataset_type)
        ]
        for item in reversed(candidates):
            field_mapping = _clean_field_mapping(item.get("field_mapping") or {})
            if not field_mapping:
                continue
            mapped_columns = {_normalize_header(column) for column in field_mapping.values()}
            if mapped_columns <= normalized_columns:
                return field_mapping
        return None

    def clear_column_mappings(self, project_root: Path) -> Path:
        """Reset saved mappings for the active project by writing an empty settings file."""

        settings_path = self.column_mapping_settings_path(project_root)
        settings_path.parent.mkdir(parents=True, exist_ok=True)
        settings_path.write_text(
            json.dumps({"schema_version": COLUMN_MAPPING_SCHEMA_VERSION, "mappings": []}, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )
        return settings_path

    def _select_import_tables(
        self,
        tables: tuple[_TabularData, ...],
    ) -> tuple[tuple[_TableClassification, ...], str, str]:
        classifications = tuple(self._classify_table(table) for table in tables)
        procurement_tables = tuple(
            item for item in classifications if item.source_dataset_type == DATASET_PROCUREMENT
        )
        contract_tables = tuple(
            item for item in classifications if item.source_dataset_type == DATASET_CONTRACT
        )
        if procurement_tables and contract_tables:
            selected = tuple(
                item
                for item in classifications
                if item.source_dataset_type in {DATASET_PROCUREMENT, DATASET_CONTRACT}
            )
            return selected, PROFILE_COMBINED_WORKBOOK, "Книга Excel: закупки + контракты"

        recognized = tuple(
            item
            for item in classifications
            if item.profile_code != PROFILE_UNKNOWN_REQUIRES_MAPPING
        )
        if recognized:
            selected = (max(recognized, key=lambda item: item.score),)
            return selected, selected[0].profile_code, selected[0].profile_label

        if classifications:
            return (classifications[0],), PROFILE_UNKNOWN_REQUIRES_MAPPING, "Нестандартный файл"
        return (), PROFILE_UNKNOWN_REQUIRES_MAPPING, "Нестандартный файл"

    def _classify_table(self, tabular: _TabularData) -> _TableClassification:
        procurement_header, procurement_columns_by_index, procurement_columns = self._detect_columns(
            tabular.rows,
            source_dataset_type=DATASET_PROCUREMENT,
        )
        contract_header, contract_columns_by_index, contract_columns = self._detect_columns(
            tabular.rows,
            source_dataset_type=DATASET_CONTRACT,
        )
        sheet_hint = _normalize_header(tabular.sheet_name or "")
        procurement_score = _profile_score(
            DATASET_PROCUREMENT,
            sheet_hint=sheet_hint,
            columns_detected=procurement_columns,
        )
        contract_score = _profile_score(
            DATASET_CONTRACT,
            sheet_hint=sheet_hint,
            columns_detected=contract_columns,
        )

        if contract_score >= 5 and contract_score > procurement_score:
            return _TableClassification(
                tabular=tabular,
                profile_code=PROFILE_CONTRACT_REGISTRY,
                profile_label="Стандартный лист 5",
                source_dataset_type=DATASET_CONTRACT,
                header_index=contract_header,
                columns_by_index=contract_columns_by_index,
                columns_detected=contract_columns,
                missing_required_columns=_missing_required_fields(DATASET_CONTRACT, contract_columns),
                score=contract_score,
            )
        if procurement_score >= 5:
            return _TableClassification(
                tabular=tabular,
                profile_code=PROFILE_PROCUREMENT_REGISTRY,
                profile_label="Стандартный лист 4",
                source_dataset_type=DATASET_PROCUREMENT,
                header_index=procurement_header,
                columns_by_index=procurement_columns_by_index,
                columns_detected=procurement_columns,
                missing_required_columns=_missing_required_fields(DATASET_PROCUREMENT, procurement_columns),
                score=procurement_score,
            )
        weak_header_detected = max(procurement_score, contract_score) >= 2
        header_index = (
            procurement_header
            if weak_header_detected and procurement_header is not None
            else contract_header if weak_header_detected else _first_non_empty_row_index(tabular.rows)
        )
        columns_by_index = (procurement_columns_by_index or contract_columns_by_index) if weak_header_detected else {}
        columns_detected = (procurement_columns or contract_columns) if weak_header_detected else {}
        return _TableClassification(
            tabular=tabular,
            profile_code=PROFILE_UNKNOWN_REQUIRES_MAPPING,
            profile_label="Нестандартный файл",
            source_dataset_type=DATASET_UNKNOWN,
            header_index=header_index,
            columns_by_index=columns_by_index,
            columns_detected=columns_detected,
            missing_required_columns=(),
            score=max(procurement_score, contract_score),
        )

    def _validate_source_file(self, source_file: Path) -> Path:
        try:
            safe_source_file = Path(source_file).resolve()
        except OSError as exc:
            raise ImportValidationError(
                f"Не удалось нормализовать путь исходного файла: {source_file}",
                code="source_file_path_error",
            ) from exc

        suffix = safe_source_file.suffix.lower()
        if suffix not in SUPPORTED_EXTENSIONS:
            raise UnsupportedInputFileError(
                f"Неподдерживаемый тип исходного файла: {safe_source_file.suffix}",
                code="unsupported_file_type",
            )
        if not safe_source_file.exists():
            raise ImportValidationError(
                f"Исходный файл не найден: {safe_source_file}",
                code="source_file_not_found",
            )
        if not safe_source_file.is_file():
            raise ImportValidationError(
                f"Исходный путь не является файлом: {safe_source_file}",
                code="source_file_is_not_file",
            )
        return safe_source_file

    def _copy_source_file(self, source_file: Path, target_folder: Path) -> Path:
        target_folder.mkdir(parents=True, exist_ok=True)
        target_file = self._next_available_path(target_folder / source_file.name)
        shutil.copy2(source_file, target_file)
        return target_file

    def _read_tabular_file(self, source_file: Path, *, sheet_name: str | None) -> _TabularData:
        return self._read_tabular_tables(source_file, sheet_name=sheet_name)[0]

    def _read_tabular_tables(self, source_file: Path, *, sheet_name: str | None) -> tuple[_TabularData, ...]:
        suffix = source_file.suffix.lower()
        if suffix == ".xlsx":
            return self._read_xlsx_tables(source_file, sheet_name=sheet_name)
        if suffix == ".csv":
            return (self._read_csv(source_file),)
        raise UnsupportedInputFileError(
            f"Неподдерживаемый тип исходного файла: {source_file.suffix}",
            code="unsupported_file_type",
        )

    def _read_xlsx(self, source_file: Path, *, sheet_name: str | None) -> _TabularData:
        return self._read_xlsx_tables(source_file, sheet_name=sheet_name)[0]

    def _read_xlsx_tables(self, source_file: Path, *, sheet_name: str | None) -> tuple[_TabularData, ...]:
        try:
            workbook = load_workbook(source_file, data_only=False, read_only=False)
        except Exception as exc:
            raise ImportValidationError(
                f"Не удалось прочитать .xlsx файл: {source_file}",
                code="xlsx_read_error",
            ) from exc

        if sheet_name is not None:
            if sheet_name not in workbook.sheetnames:
                raise ImportValidationError(
                    f"Лист Excel не найден: {sheet_name}",
                    code="xlsx_sheet_not_found",
                )
            sheets = (workbook[sheet_name],)
        else:
            sheets = tuple(workbook.worksheets)

        tables: list[_TabularData] = []
        for sheet in sheets:
            rows: list[tuple[str, ...]] = []
            for row in sheet.iter_rows():
                values = [_excel_cell_text(cell) for cell in row]
                rows.append(tuple(values))
            tables.append(
                _TabularData(
                    rows=tuple(rows),
                    sheet_name=sheet.title,
                    source_type="Excel",
                )
            )
        return tuple(tables)

    def _read_csv(self, source_file: Path) -> _TabularData:
        try:
            text = source_file.read_text(encoding="utf-8-sig")
        except UnicodeDecodeError as exc:
            raise ImportValidationError(
                f"CSV не читается как UTF-8 / UTF-8-SIG: {source_file}",
                code="csv_encoding_warning",
            ) from exc

        sample = text[:4096]
        try:
            dialect = csv.Sniffer().sniff(sample)
        except csv.Error:
            dialect = csv.excel

        rows = tuple(tuple(_to_text(value) for value in row) for row in csv.reader(text.splitlines(), dialect))
        return _TabularData(rows=rows, sheet_name=None, source_type="CSV")

    def _detect_columns(
        self,
        rows: tuple[tuple[str, ...], ...],
        *,
        source_dataset_type: str = DATASET_PROCUREMENT,
    ) -> tuple[int | None, dict[int, str], dict[str, str]]:
        alias_map = _build_alias_map(source_dataset_type=source_dataset_type)
        best_header_index: int | None = None
        best_score = 0
        best_columns_by_index: dict[int, str] = {}
        best_columns_detected: dict[str, str] = {}

        for row_index, row in enumerate(rows[:20]):
            columns_by_index: dict[int, str] = {}
            columns_detected: dict[str, str] = {}
            for column_index, cell_value in enumerate(row):
                field_name = alias_map.get(_normalize_header(cell_value))
                if field_name is None or field_name in columns_detected:
                    continue
                columns_by_index[column_index] = field_name
                columns_detected[field_name] = cell_value
            score = len(columns_detected)
            if score > best_score:
                best_header_index = row_index
                best_score = score
                best_columns_by_index = columns_by_index
                best_columns_detected = columns_detected

        if best_header_index is None:
            return None, {}, {}
        return best_header_index, best_columns_by_index, best_columns_detected

    def _build_rows(
        self,
        *,
        rows: tuple[tuple[str, ...], ...],
        header_index: int | None,
        columns_by_index: Mapping[int, str],
        source_type: str,
        source_dataset_type: str,
        source_name: str = "",
        source_sheet: str = "",
        row_id_offset: int = 0,
    ) -> tuple[list[ImportRow], int]:
        if header_index is None:
            return [], 0

        header_row = rows[header_index]
        normalized_rows: list[ImportRow] = []
        skipped_empty_rows_count = 0
        for source_row_index, source_row in enumerate(rows[header_index + 1 :], start=header_index + 2):
            if not _has_any_value(source_row):
                continue
            row_values = self._extract_row_values(source_row, columns_by_index)
            purchase_number = row_values.get("purchase_number", "")
            procurement_number = row_values.get("procurement_number", "") or purchase_number
            contract_registry_number = row_values.get("contract_registry_number", "")
            contract_number = row_values.get("contract_number", "") or contract_registry_number
            law_type = row_values.get("law_type", "") or row_values.get("law", "") or _infer_law_type(
                source_name=source_name,
                source_url=row_values.get("source_url", ""),
                source_text=row_values.get("source_text", ""),
            )
            contract_url = row_values.get("contract_url", "")
            source_raw = (
                row_values.get("source_raw", "")
                or row_values.get("source_url", "")
                or contract_url
                or row_values.get("source_text", "")
            )
            source_details = parse_source_raw(source_raw, law_type=law_type, purchase_number=procurement_number)
            source_url = (
                row_values.get("source_url", "")
                or contract_url
                or source_details["source_url"]
                or source_details["generated_eis_url"]
                or source_details["generated_platform_url"]
            )
            if source_dataset_type == DATASET_CONTRACT:
                contract_url = contract_url or source_url
            source_text = (
                row_values.get("source_text", "")
                or row_values.get("platform_name", "")
                or row_values.get("domain", "")
            )
            if not source_text and source_raw and source_raw != source_url:
                source_text = source_raw
            name = row_values.get("name", "") or row_values.get("subject", "")
            subject = row_values.get("subject", "") or name
            if not _has_key_registry_value(
                source_dataset_type=source_dataset_type,
                purchase_number=procurement_number,
                contract_registry_number=contract_registry_number,
                name=name,
                subject=subject,
                customer_name=row_values.get("customer_name", ""),
                supplier_name=row_values.get("supplier_name", ""),
                source_raw=source_raw,
                source_url=source_url,
                source_text=source_text,
            ):
                skipped_empty_rows_count += 1
                continue

            row_sequence_number = len(normalized_rows) + 1
            warnings: list[str] = []
            source_anchor = source_url or source_raw or contract_url or source_text
            if source_dataset_type == DATASET_CONTRACT and not contract_registry_number:
                warnings.append("missing_contract_registry_number")
            if (
                source_dataset_type == DATASET_CONTRACT
                and not contract_registry_number
                and not (procurement_number and source_anchor)
            ):
                warnings.append("missing_required_contract_identifier")
            if source_dataset_type == DATASET_PROCUREMENT and not procurement_number:
                warnings.append("missing_purchase_number")
            if not source_anchor:
                warnings.append("source_missing_manual_review")
            import_status = "ok_with_warnings" if warnings else "ok"

            raw_values = {
                _to_text(header_row[index]) or f"column_{index + 1}": _sanitize_output_value(value)
                for index, value in enumerate(source_row)
            }
            source_resolution = self.source_resolver.resolve(
                {
                    "source_dataset_type": source_dataset_type,
                    "law": law_type,
                    "law_type": law_type,
                    "procurement_number": procurement_number,
                    "purchase_number": procurement_number,
                    "notice_number": row_values.get("notice_number", ""),
                    "procedure_number": row_values.get("procedure_number", ""),
                    "contract_number": contract_number,
                    "contract_registry_number": contract_registry_number,
                    "contract_url": contract_url,
                    "source_url": source_url,
                    "source_raw": source_raw,
                    "source_text": source_text,
                    "platform_name": row_values.get("platform_name", "") or source_details["platform_name"],
                    "domain": row_values.get("domain", ""),
                    "eis_number": source_details["eis_number"],
                    "etp_number": source_details["platform_number"],
                    "platform_number": source_details["platform_number"],
                    "customer_name": row_values.get("customer_name", ""),
                    "subject": subject,
                    "raw_values": raw_values,
                }
            )
            normalized_rows.append(
                ImportRow(
                    row_id=f"row-{row_id_offset + row_sequence_number}",
                    row_sequence_number=row_id_offset + row_sequence_number,
                    source_dataset_type=source_dataset_type,
                    source_file=source_name,
                    source_sheet=source_sheet,
                    source_row_number=source_row_index,
                    purchase_number=procurement_number,
                    procurement_number=procurement_number,
                    notice_number=row_values.get("notice_number", ""),
                    procedure_number=row_values.get("procedure_number", ""),
                    contract_number=contract_number,
                    contract_registry_number=contract_registry_number,
                    registry_identifier=row_values.get("registry_identifier", "") or contract_registry_number,
                    law_type=law_type,
                    law=law_type,
                    source_raw=source_raw,
                    source_url=source_url,
                    contract_url=contract_url,
                    source_text=source_text,
                    source_type=source_resolution.source_type,
                    route=source_resolution.route,
                    route_status=source_resolution.route_status,
                    route_confidence=source_resolution.route_confidence,
                    route_reason=source_resolution.route_reason,
                    normalized_source_url=source_resolution.normalized_source_url,
                    normalized_domain=source_resolution.normalized_domain,
                    normalized_eis_number=source_resolution.normalized_eis_number,
                    normalized_etp_number=source_resolution.normalized_etp_number,
                    normalized_procedure_number=source_resolution.normalized_procedure_number,
                    source_warnings=source_resolution.source_warnings,
                    next_action=source_resolution.next_action,
                    rostender_number=source_details["rostender_number"],
                    eis_number=source_details["eis_number"],
                    platform_number=source_details["platform_number"],
                    source_label=source_details["source_label"],
                    platform_code=source_details["platform_code"],
                    platform_name=row_values.get("platform_name", "") or source_details["platform_name"],
                    generated_eis_url=source_details["generated_eis_url"],
                    generated_platform_url=source_details["generated_platform_url"],
                    priority=_normalize_priority(row_values.get("priority", "")),
                    customer_name=row_values.get("customer_name", ""),
                    customer_inn=row_values.get("customer_inn", ""),
                    region=row_values.get("region", ""),
                    nmck=row_values.get("nmck", ""),
                    name=name,
                    subject=subject,
                    supplier_name=row_values.get("supplier_name", ""),
                    supplier_inn=row_values.get("supplier_inn", ""),
                    contract_price=row_values.get("contract_price", ""),
                    contract_date=row_values.get("contract_date", ""),
                    execution_status=row_values.get("execution_status", ""),
                    okpd2=row_values.get("okpd2", ""),
                    ktru=row_values.get("ktru", ""),
                    import_status=import_status,
                    import_warnings=tuple(warnings),
                    raw_values=raw_values,
                    warnings=tuple(warnings),
                )
            )
        return normalized_rows, skipped_empty_rows_count

    def _apply_field_mapping(
        self,
        *,
        rows: tuple[tuple[str, ...], ...],
        header_index: int | None,
        columns_by_index: Mapping[int, str],
        columns_detected: Mapping[str, str],
        field_mapping: Mapping[str, str],
    ) -> tuple[dict[int, str], dict[str, str], tuple[ImportWarning, ...]]:
        if header_index is None:
            return dict(columns_by_index), dict(columns_detected), (
                ImportWarning(
                    code="field_mapping_without_header",
                    message="Сопоставление колонок не применено: строка заголовков не найдена.",
                ),
            )

        header_row = rows[header_index]
        header_by_normalized = {
            _normalize_header(header): (index, _to_text(header))
            for index, header in enumerate(header_row)
            if _to_text(header)
        }
        normalized_columns_by_index = dict(columns_by_index)
        normalized_columns_detected = dict(columns_detected)
        warnings: list[ImportWarning] = []

        valid_fields = set(FIELD_ALIASES)
        for field_name, source_header in field_mapping.items():
            field = _to_text(field_name)
            header = _to_text(source_header)
            if not field or not header:
                continue
            if field not in valid_fields:
                warnings.append(
                    ImportWarning(
                        code="field_mapping_unknown_field",
                        message=f"Неизвестное поле сопоставления пропущено: {field}.",
                    )
                )
                continue
            matched = header_by_normalized.get(_normalize_header(header))
            if matched is None:
                warnings.append(
                    ImportWarning(
                        code="field_mapping_source_column_not_found",
                        message=f"Колонка для поля {field} не найдена: {header}.",
                    )
                )
                continue
            column_index, original_header = matched
            for existing_index, existing_field in list(normalized_columns_by_index.items()):
                if existing_field == field and existing_index != column_index:
                    del normalized_columns_by_index[existing_index]
            normalized_columns_by_index[column_index] = field
            normalized_columns_detected[field] = original_header

        return normalized_columns_by_index, normalized_columns_detected, tuple(warnings)

    @staticmethod
    def _extract_row_values(
        row: tuple[str, ...],
        columns_by_index: Mapping[int, str],
    ) -> dict[str, str]:
        values = {field: "" for field in FIELD_ALIASES}
        for column_index, field_name in columns_by_index.items():
            if column_index >= len(row):
                continue
            values[field_name] = _sanitize_output_value(row[column_index])
        return values

    @staticmethod
    def _next_available_path(path: Path) -> Path:
        if not path.exists():
            return path

        suffix = 2
        while True:
            candidate = path.with_name(f"{path.stem}_{suffix}{path.suffix}")
            if not candidate.exists():
                return candidate
            suffix += 1

    def _write_registry_json(self, result: ImportResult) -> None:
        data = {
            "schema_version": SCHEMA_VERSION,
            "file_profile": result.file_profile,
            "file_profile_label": result.file_profile_label,
            "source_file": str(result.source_file),
            "copied_source_file": str(result.copied_source_file),
            "source_files": [str(path) for path in (result.source_files or (result.source_file,))],
            "copied_source_files": [str(path) for path in (result.copied_source_files or (result.copied_source_file,))],
            "sheet_name": result.sheet_name,
            "sheet_names": list(result.sheet_names),
            "dataset_counts": dict(result.dataset_counts),
            "sheet_profiles": [dict(profile) for profile in result.sheet_profiles],
            "header_row": result.header_row,
            "rows_count": result.rows_count,
            "physical_data_rows_count": result.physical_data_rows_count,
            "skipped_empty_rows_count": result.skipped_empty_rows_count,
            "columns_detected": dict(result.columns_detected),
            "rows": [row.to_json_dict() for row in result.rows],
        }
        result.normalized_registry_json_path.write_text(
            json.dumps(data, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )

    def _write_registry_csv(self, result: ImportResult) -> None:
        with result.normalized_registry_csv_path.open(
            "w",
            encoding="utf-8-sig",
            newline="",
        ) as csv_file:
            writer = csv.DictWriter(csv_file, fieldnames=NORMALIZED_COLUMNS)
            writer.writeheader()
            for row in result.rows:
                writer.writerow(row.to_csv_dict())

    def _write_normalization_report(self, result: ImportResult) -> None:
        report = {
            "file_profile": result.file_profile,
            "file_profile_label": result.file_profile_label,
            "source_file": str(result.source_file),
            "copied_source_file": str(result.copied_source_file),
            "source_files": [str(path) for path in (result.source_files or (result.source_file,))],
            "copied_source_files": [str(path) for path in (result.copied_source_files or (result.copied_source_file,))],
            "sheet_name": result.sheet_name,
            "sheet_names": list(result.sheet_names),
            "dataset_counts": dict(result.dataset_counts),
            "sheet_profiles": [dict(profile) for profile in result.sheet_profiles],
            "header_row": result.header_row,
            "rows_count": result.rows_count,
            "columns_detected": dict(result.columns_detected),
            "missing_required_columns": list(result.missing_required_columns),
            "warnings": [warning.__dict__ for warning in result.warnings],
            "errors": [error.__dict__ for error in result.errors],
            "created_files": [str(path) for path in result.created_files],
        }
        result.normalization_report_json_path.write_text(
            json.dumps(report, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )

        md_lines = [
            "# Отчет нормализации импорта",
            "",
            f"- file_profile: `{result.file_profile}`",
            f"- file_profile_label: `{result.file_profile_label}`",
            f"- source_file: `{result.source_file}`",
            f"- copied_source_file: `{result.copied_source_file}`",
            f"- source_files: `{', '.join(str(path) for path in (result.source_files or (result.source_file,)))}`",
            f"- copied_source_files: `{', '.join(str(path) for path in (result.copied_source_files or (result.copied_source_file,)))}`",
            f"- sheet_name: `{result.sheet_name or ''}`",
            f"- sheet_names: `{', '.join(result.sheet_names)}`",
            f"- dataset_counts: `{json.dumps(dict(result.dataset_counts), ensure_ascii=False)}`",
            f"- header_row: `{result.header_row or ''}`",
            f"- rows_count: `{result.rows_count}`",
            f"- physical_data_rows_count: `{result.physical_data_rows_count}`",
            f"- skipped_empty_rows_count: `{result.skipped_empty_rows_count}`",
            f"- missing_required_columns: `{', '.join(result.missing_required_columns)}`",
            "",
            "## Листы / профили",
            "",
        ]
        if result.sheet_profiles:
            md_lines.extend(
                "- `{sheet_name}`: `{profile_code}` / `{source_dataset_type}`, rows `{rows_count}`".format(
                    sheet_name=profile.get("sheet_name", ""),
                    profile_code=profile.get("profile_code", ""),
                    source_dataset_type=profile.get("source_dataset_type", ""),
                    rows_count=profile.get("rows_count", 0),
                )
                for profile in result.sheet_profiles
            )
        else:
            md_lines.append("- нет")

        md_lines.extend(
            [
                "",
                "## Найденные колонки",
                "",
            ]
        )
        if result.columns_detected:
            md_lines.extend(
                f"- `{field}`: `{source}`"
                for field, source in sorted(result.columns_detected.items())
            )
        else:
            md_lines.append("- не найдены")

        md_lines.extend(["", "## Warnings / предупреждения", ""])
        if result.warnings:
            md_lines.extend(
                f"- `{warning.code}`: {warning.message}"
                for warning in result.warnings
            )
        else:
            md_lines.append("- нет")

        md_lines.extend(["", "## Errors / ошибки", ""])
        if result.errors:
            md_lines.extend(
                f"- `{error.code}`: {error.message}"
                for error in result.errors
            )
        else:
            md_lines.append("- нет")

        md_lines.extend(["", "## Созданные файлы", ""])
        md_lines.extend(f"- `{path}`" for path in result.created_files)
        result.normalization_report_md_path.write_text(
            "\n".join(md_lines) + "\n",
            encoding="utf-8",
        )

    @staticmethod
    def _warning_message(code: str) -> str:
        messages = {
            "missing_purchase_number": "В строке не найден номер закупки / извещения.",
            "missing_contract_registry_number": "В строке не найден реестровый номер контракта / договора.",
            "missing_required_contract_identifier": "В строке не найден contract_registry_number или fallback-связка procurement_number + source.",
            "source_missing_manual_review": "В строке не найден первоисточник; требуется ручная проверка.",
            "empty_registry": "После чтения источника не найдено строк реестра.",
            "missing_required_column": "Не найдена обязательная колонка.",
        }
        return messages.get(code, code)


def mapping_targets_for_dataset(dataset_type: str = DATASET_PROCUREMENT) -> tuple[ColumnMappingTarget, ...]:
    normalized = _normalize_dataset_type(dataset_type)
    if normalized == DATASET_CONTRACT:
        return _unique_mapping_targets(
            COMMON_MAPPING_TARGETS + CONTRACT_MAPPING_TARGETS + PROCUREMENT_MAPPING_TARGETS + SOURCE_MAPPING_TARGETS
        )
    return _unique_mapping_targets(
        COMMON_MAPPING_TARGETS + PROCUREMENT_MAPPING_TARGETS + CONTRACT_MAPPING_TARGETS + SOURCE_MAPPING_TARGETS
    )


def mapping_targets_json(dataset_type: str = DATASET_PROCUREMENT) -> list[dict[str, str | bool]]:
    return [target.to_json_dict() for target in mapping_targets_for_dataset(dataset_type)]


def validate_column_mapping(dataset_type: str, field_mapping: Mapping[str, str]) -> tuple[str, ...]:
    clean_mapping = _clean_field_mapping(field_mapping)
    fields = set(clean_mapping)
    normalized_dataset = _normalize_dataset_type(dataset_type)
    if normalized_dataset == DATASET_CONTRACT:
        errors: list[str] = []
        if not (
            "contract_registry_number" in fields
            or ({"procurement_number", "purchase_number"} & fields and {"contract_url", "source_url", "source_raw"} & fields)
        ):
            errors.append("Не сопоставлен contract_registry_number или связка procurement_number + contract_url/source_url/source_raw.")
        return tuple(errors)

    errors = []
    if not ({"procurement_number", "purchase_number"} & fields):
        errors.append("Не сопоставлено обязательное поле: procurement_number.")
    if not ({"law", "law_type"} & fields):
        errors.append("Не сопоставлено обязательное поле: law.")
    if not ({"source_raw", "source_url"} & fields):
        errors.append("Не сопоставлено обязательное поле: source_url или source_raw.")
    if "customer_name" not in fields:
        errors.append("Не сопоставлено обязательное поле: customer_name.")
    if not ({"subject", "name"} & fields):
        errors.append("Не сопоставлено обязательное поле: subject.")
    return tuple(errors)


def _clean_field_mapping(field_mapping: Mapping[str, Any]) -> dict[str, str]:
    valid_fields = set(FIELD_ALIASES)
    result: dict[str, str] = {}
    for field_name, source_header in field_mapping.items():
        field = _to_text(field_name)
        header = _to_text(source_header)
        if not field or not header or field not in valid_fields:
            continue
        if field == "ignore_column":
            continue
        if field == "law":
            field = "law_type"
        if field == "etp_number":
            field = "platform_number"
        result[field] = header
    return result


def _clean_field_mapping_for_storage(field_mapping: Mapping[str, Any]) -> dict[str, str]:
    valid_fields = set(FIELD_ALIASES) | MAPPING_ONLY_FIELDS
    result: dict[str, str] = {}
    for field_name, source_header in field_mapping.items():
        field = _to_text(field_name)
        header = _to_text(source_header)
        if not field or not header or field not in valid_fields or field == "ignore_column":
            continue
        result[field] = header
    return result


def _column_to_field_mapping(field_mapping: Mapping[str, str]) -> dict[str, str]:
    return {
        source_column: field_name
        for field_name, source_column in field_mapping.items()
        if _to_text(source_column) and _to_text(field_name)
    }


def _unique_mapping_targets(targets: Iterable[ColumnMappingTarget]) -> tuple[ColumnMappingTarget, ...]:
    seen: set[tuple[str, str]] = set()
    unique: list[ColumnMappingTarget] = []
    for target in targets:
        key = (target.group, target.field)
        if key in seen:
            continue
        seen.add(key)
        unique.append(target)
    return tuple(unique)


def _required_fields_status(dataset_type: str, field_mapping: Mapping[str, str]) -> dict[str, bool]:
    fields = set(field_mapping)
    if _normalize_dataset_type(dataset_type) == DATASET_CONTRACT:
        return {
            "contract_registry_number": "contract_registry_number" in fields,
            "procurement_number": bool({"procurement_number", "purchase_number"} & fields),
            "contract_url": bool({"contract_url", "source_url", "source_raw"} & fields),
        }
    return {
        "procurement_number": bool({"procurement_number", "purchase_number"} & fields),
        "law": bool({"law", "law_type"} & fields),
        "source_raw_or_source_url": bool({"source_raw", "source_url"} & fields),
        "customer_name": "customer_name" in fields,
        "subject": bool({"subject", "name"} & fields),
    }


def _mapping_file_name(source_file: Path | str, sheet_name: str | None) -> str:
    source_name = Path(str(source_file)).stem if source_file else "source"
    sheet = sheet_name or "CSV"
    return f"{_safe_slug(source_name)}__{_safe_slug(sheet)}__mapping.json"


def _safe_slug(value: str) -> str:
    slug = re.sub(r"[^0-9A-Za-zА-Яа-я_-]+", "_", _to_text(value)).strip("_")
    return slug[:80] or "empty"


def _dataset_type_for_mapping(
    field_mapping: Mapping[str, str] | None,
    *,
    requested: str | None = None,
) -> str:
    normalized_requested = _normalize_dataset_type(requested or "")
    if normalized_requested in {DATASET_PROCUREMENT, DATASET_CONTRACT}:
        return normalized_requested
    fields = set(_clean_field_mapping(field_mapping or {}))
    contract_markers = {
        "contract_registry_number",
        "contract_url",
        "supplier_name",
        "supplier_inn",
        "contract_price",
        "contract_date",
        "execution_status",
    }
    if fields & contract_markers:
        return DATASET_CONTRACT
    return DATASET_PROCUREMENT


def _normalize_dataset_type(value: str) -> str:
    text = _to_text(value).lower()
    if text == DATASET_CONTRACT:
        return DATASET_CONTRACT
    if text == DATASET_PROCUREMENT:
        return DATASET_PROCUREMENT
    return DATASET_PROCUREMENT


def _mapping_id(dataset_type: str, source_columns: Iterable[str]) -> str:
    normalized_columns = "_".join(
        sorted(_normalize_header(column) for column in source_columns if _to_text(column))
    )
    compact = normalized_columns[:80] or "empty_columns"
    return f"{_normalize_dataset_type(dataset_type)}_{compact}"


def _sample_values_for_columns(
    columns: Iterable[str],
    preview_rows: Iterable[Mapping[str, str]],
) -> dict[str, tuple[str, ...]]:
    samples: dict[str, list[str]] = {_to_text(column): [] for column in columns if _to_text(column)}
    for row in preview_rows:
        for column in list(samples):
            value = _sanitize_output_value(row.get(column, ""))
            if value and value not in samples[column]:
                samples[column].append(value)
            if len(samples[column]) > 3:
                samples[column] = samples[column][:3]
    return {column: tuple(values[:3]) for column, values in samples.items()}


def _build_alias_map(*, source_dataset_type: str = DATASET_PROCUREMENT) -> dict[str, str]:
    aliases: dict[str, str] = {}
    for field_name, field_aliases in FIELD_ALIASES.items():
        for alias in field_aliases:
            aliases[_normalize_header(alias)] = field_name
    if source_dataset_type == DATASET_CONTRACT:
        for field_name in CONTRACT_ALIAS_FIELDS:
            for alias in FIELD_ALIASES.get(field_name, ()):
                aliases[_normalize_header(alias)] = field_name
        aliases[_normalize_header("Реестровый номер")] = "contract_registry_number"
    else:
        aliases[_normalize_header("Реестровый номер")] = "purchase_number"
    return aliases


def _excel_cell_text(cell: Any) -> str:
    value = cell.value
    if cell.hyperlink is not None and cell.hyperlink.target:
        value = cell.hyperlink.target
    formula_url = _parse_hyperlink_formula(value)
    if formula_url:
        value = formula_url
    return _to_text(value)


def _parse_hyperlink_formula(value: object) -> str:
    text = _to_text(value)
    if not text.startswith("="):
        return ""
    match = re.match(r"^=\s*(?:HYPERLINK|ГИПЕРССЫЛКА)\s*\((.*)\)\s*$", text, flags=re.IGNORECASE)
    if not match:
        return ""
    body = match.group(1)
    separator = ";" if ";" in body else ","
    first = body.split(separator, 1)[0].strip()
    if len(first) >= 2 and first[0] == first[-1] == '"':
        return first[1:-1]
    return first


def _normalize_header(value: object) -> str:
    text = _to_text(value).lower().replace("ё", "е")
    return re.sub(r"[^0-9a-zа-я]+", "", text)


def _to_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _has_any_value(row: Iterable[str]) -> bool:
    return any(_to_text(value) for value in row)


def _first_non_empty_row_index(rows: tuple[tuple[str, ...], ...]) -> int | None:
    for index, row in enumerate(rows[:20]):
        if _has_any_value(row):
            return index
    return None


def _header_index_for_mapping(
    rows: tuple[tuple[str, ...], ...],
    field_mapping: Mapping[str, str],
    *,
    default: int | None,
) -> int | None:
    expected_headers = {
        _normalize_header(header)
        for header in field_mapping.values()
        if _to_text(header)
    }
    best_index = default
    best_score = 0
    for index, row in enumerate(rows[:20]):
        normalized_headers = {_normalize_header(value) for value in row if _to_text(value)}
        score = len(expected_headers & normalized_headers)
        if score > best_score:
            best_score = score
            best_index = index
    if best_index is not None:
        return best_index
    return _first_non_empty_row_index(rows)


STANDARD_LIST4_REQUIRED_FIELDS = {
    "purchase_number",
    "law_type",
    "name",
    "customer_name",
}


def _standard_list4_missing_required_fields(columns_detected: Mapping[str, str]) -> tuple[str, ...]:
    source_present = "source_raw" in columns_detected or "source_url" in columns_detected
    detected = set(columns_detected)
    if "procurement_number" in detected:
        detected.add("purchase_number")
    if "law" in detected:
        detected.add("law_type")
    if "subject" in detected:
        detected.add("name")
    if source_present:
        detected.add("source_raw")
    return tuple(sorted(STANDARD_LIST4_REQUIRED_FIELDS - detected))


def _contract_missing_required_fields(columns_detected: Mapping[str, str]) -> tuple[str, ...]:
    return (
        ("missing_required_contract_identifier",)
        if _contract_identifier_missing(columns_detected)
        else ()
    )


def _missing_required_fields(
    source_dataset_type: str,
    columns_detected: Mapping[str, str],
) -> tuple[str, ...]:
    if source_dataset_type == DATASET_CONTRACT:
        return _contract_missing_required_fields(columns_detected)
    if source_dataset_type == DATASET_PROCUREMENT:
        return _standard_list4_missing_required_fields(columns_detected)
    return ()


def _critical_missing_required_fields(
    source_dataset_type: str,
    columns_detected: Mapping[str, str],
) -> tuple[str, ...]:
    if source_dataset_type == DATASET_CONTRACT:
        return _contract_missing_required_fields(columns_detected)
    if source_dataset_type == DATASET_PROCUREMENT:
        return (
            ("procurement_number",)
            if "purchase_number" not in columns_detected and "procurement_number" not in columns_detected
            else ()
        )
    return ()


def _contract_identifier_missing(columns_detected: Mapping[str, str]) -> bool:
    fields = set(columns_detected)
    if "contract_registry_number" in fields:
        return False
    has_procurement = bool({"procurement_number", "purchase_number"} & fields)
    has_source = bool({"contract_url", "source_url", "source_raw"} & fields)
    return not (has_procurement and has_source)


def _preview_missing_required_fields(
    selected_tables: tuple[_TableClassification, ...],
    *,
    selected: _TableClassification | None,
    profile_code: str,
    columns_detected: Mapping[str, str],
) -> tuple[str, ...]:
    if profile_code == PROFILE_COMBINED_WORKBOOK:
        missing: list[str] = []
        for table in selected_tables:
            for field_name in table.missing_required_columns:
                marker = f"{table.source_dataset_type}.{field_name}"
                if marker not in missing:
                    missing.append(marker)
        return tuple(missing)
    return _missing_required_fields(
        selected.source_dataset_type if selected is not None else DATASET_UNKNOWN,
        columns_detected=columns_detected,
    )


def _profile_score(
    source_dataset_type: str,
    *,
    sheet_hint: str,
    columns_detected: Mapping[str, str],
) -> int:
    fields = set(columns_detected)
    score = len(fields)
    if source_dataset_type == DATASET_CONTRACT:
        score += 4 if "контракт" in sheet_hint or "договор" in sheet_hint else 0
        score += 3 if "contract_registry_number" in fields else 0
        score += 2 if "supplier_name" in fields else 0
        score += 2 if "contract_price" in fields else 0
        score += 1 if "contract_date" in fields else 0
        return score
    score += 4 if "закуп" in sheet_hint and "контракт" not in sheet_hint else 0
    score += 3 if "purchase_number" in fields else 0
    score += 2 if "source_raw" in fields or "source_url" in fields else 0
    score += 1 if "nmck" in fields else 0
    return score


def _merge_columns_detected(
    selected_tables: tuple[_TableClassification, ...],
) -> dict[str, str]:
    if len(selected_tables) == 1:
        return dict(selected_tables[0].columns_detected)
    merged: dict[str, str] = {}
    for selected in selected_tables:
        sheet_name = selected.tabular.sheet_name or selected.source_dataset_type
        for field_name, source_header in selected.columns_detected.items():
            merged[f"{sheet_name}.{field_name}"] = source_header
    return merged


def _count_rows_by_dataset(rows: Iterable[ImportRow]) -> dict[str, int]:
    counts = {
        DATASET_PROCUREMENT: 0,
        DATASET_CONTRACT: 0,
    }
    for row in rows:
        counts[row.source_dataset_type] = counts.get(row.source_dataset_type, 0) + 1
    return counts


def _build_sheet_profiles(
    selected_tables: tuple[_TableClassification, ...],
    rows: Iterable[ImportRow],
) -> tuple[Mapping[str, Any], ...]:
    rows_by_sheet: dict[tuple[str, str], int] = {}
    for row in rows:
        key = (row.source_sheet, row.source_dataset_type)
        rows_by_sheet[key] = rows_by_sheet.get(key, 0) + 1
    profiles: list[Mapping[str, Any]] = []
    for selected in selected_tables:
        sheet_name = selected.tabular.sheet_name or ""
        effective_dataset_type = selected.source_dataset_type
        if selected.source_dataset_type == DATASET_UNKNOWN:
            dataset_counts = {
                dataset_type: count
                for (row_sheet, dataset_type), count in rows_by_sheet.items()
                if row_sheet == sheet_name and dataset_type != DATASET_UNKNOWN
            }
            if len(dataset_counts) == 1:
                effective_dataset_type = next(iter(dataset_counts))
        profile = dict(_classification_to_profile(selected))
        profile["source_dataset_type"] = effective_dataset_type
        profile["rows_count"] = rows_by_sheet.get((sheet_name, effective_dataset_type), 0)
        profiles.append(profile)
    return tuple(profiles)


def _classification_to_profile(selected: _TableClassification) -> Mapping[str, Any]:
    return {
        "sheet_name": selected.tabular.sheet_name,
        "profile_code": selected.profile_code,
        "profile_label": selected.profile_label,
        "source_dataset_type": selected.source_dataset_type,
        "header_row": None if selected.header_index is None else selected.header_index + 1,
        "columns_detected": dict(selected.columns_detected),
        "missing_required_columns": list(selected.missing_required_columns),
    }


def _detect_file_profile(*, sheet_name: str | None, columns_detected: Mapping[str, str]) -> str:
    missing = _standard_list4_missing_required_fields(columns_detected)
    if not missing and _normalize_header(sheet_name or "") == _normalize_header("Реестр закупок товаров"):
        return "Стандартный лист 4"
    if not missing:
        return "Стандартный лист 4"
    detected_values = {_normalize_header(value) for value in columns_detected.values()}
    if any("контракт" in value for value in detected_values):
        return "Стандартный лист 5"
    return "Нестандартный файл"


def _count_physical_data_rows(rows: tuple[tuple[str, ...], ...], header_index: int | None) -> int:
    if header_index is None:
        return 0
    return sum(1 for row in rows[header_index + 1 :] if _has_any_value(row))


def _count_keyed_data_rows(
    *,
    rows: tuple[tuple[str, ...], ...],
    header_index: int | None,
    columns_by_index: Mapping[int, str],
    source_dataset_type: str = DATASET_PROCUREMENT,
) -> tuple[int, int]:
    if header_index is None:
        return 0, 0
    data_count = 0
    skipped_count = 0
    for row in rows[header_index + 1 :]:
        if not _has_any_value(row):
            continue
        values = ImportService._extract_row_values(row, columns_by_index)
        source_raw = (
            values.get("source_raw", "")
            or values.get("source_url", "")
            or values.get("contract_url", "")
            or values.get("source_text", "")
        )
        if _has_key_registry_value(
            source_dataset_type=source_dataset_type,
            purchase_number=values.get("purchase_number", ""),
            contract_registry_number=values.get("contract_registry_number", ""),
            name=values.get("name", ""),
            subject=values.get("subject", ""),
            customer_name=values.get("customer_name", ""),
            supplier_name=values.get("supplier_name", ""),
            source_raw=source_raw,
            source_url=values.get("source_url", ""),
            source_text=values.get("source_text", ""),
        ):
            data_count += 1
        else:
            skipped_count += 1
    return data_count, skipped_count


def _has_key_registry_value(
    *,
    source_dataset_type: str,
    purchase_number: str,
    contract_registry_number: str,
    name: str,
    subject: str,
    customer_name: str,
    supplier_name: str,
    source_raw: str,
    source_url: str,
    source_text: str,
) -> bool:
    if source_dataset_type == DATASET_CONTRACT:
        return any(
            _to_text(value)
            for value in (
                contract_registry_number,
                purchase_number,
                subject,
                name,
                supplier_name,
                customer_name,
                source_raw,
                source_url,
                source_text,
            )
        )
    return any(
        _to_text(value)
        for value in (
            purchase_number,
            name,
            subject,
            customer_name,
            source_raw,
            source_url,
            source_text,
        )
    )


def _normalize_priority(value: str) -> str:
    text = _to_text(value)
    if not text:
        return CANONICAL_UNKNOWN_PRIORITY_FOLDER
    if text in BAD_PRIORITY_LABELS:
        return CANONICAL_UNKNOWN_PRIORITY_FOLDER
    return text


def _infer_law_type(*, source_name: str, source_url: str, source_text: str) -> str:
    joined = " ".join([source_name, source_url, source_text]).lower().replace("ё", "е")
    compact_name = re.sub(r"[^0-9a-zа-я]+", "", source_name.lower().replace("ё", "е"))
    if "коммер" in joined or "commercial" in joined:
        return "Коммерческая"
    if "notice223" in joined or "ea223" in joined or compact_name in {"223", "223xlsx", "223csv"}:
        return "223-ФЗ"
    if "fz223" in joined or "223фз" in joined:
        return "223-ФЗ"
    if "epz/order/notice" in joined or compact_name in {"44", "44xlsx", "44csv"}:
        return "44-ФЗ"
    if "44фз" in joined or "fz44" in joined:
        return "44-ФЗ"
    return ""


def _sanitize_output_value(value: object) -> str:
    text = _to_text(value)
    for bad_label in BAD_PRIORITY_LABELS:
        text = text.replace(bad_label, CANONICAL_UNKNOWN_PRIORITY_FOLDER)
    return text
