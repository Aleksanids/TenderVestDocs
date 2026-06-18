"""ETAP 32.1 export-data service.

The public workflow is named "Экспорт данных".  The historical module and
class names are kept as an internal compatibility detail for existing imports.
"""

from __future__ import annotations

from collections import Counter
import csv
from dataclasses import dataclass
from datetime import datetime, timezone
import json
from pathlib import Path
from typing import Any, Callable, Mapping, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter, quote_sheetname
from openpyxl.worksheet.datavalidation import DataValidation

from .import_service import NORMALIZED_REGISTRY_JSON_NAME
from .normative_ui_integration import (
    NORMATIVE_EXPORT_FOLDER_NAME,
    NORMATIVE_SHEET6_READINESS_JSON_NAME,
    NORMATIVE_UI_EXPORT_JSON_NAME,
)
from .procurement_contract_link import (
    COMPLETED_CONTRACT_REGISTRY_EXPORT_JSON_NAME,
    COMPLETED_PROCUREMENT_REGISTRY_EXPORT_JSON_NAME,
    CONTRACT_LINK_COVERAGE_SUMMARY_JSON_NAME,
    CONTRACT_LINK_EXPORT_FOLDER_NAME,
    CONTRACT_LINK_UI_EXPORT_JSON_NAME,
)
from .project_service import ProjectInfo, ProjectService
from .protocol_ui_integration import (
    PROTOCOL_CONTRACT_EVIDENCE_EXPORT_JSON_NAME,
    PROTOCOL_EXPORT_FOLDER_NAME,
    PROTOCOL_SUPPLIER_RESULTS_EXPORT_JSON_NAME,
    PROTOCOL_UI_EXPORT_JSON_NAME,
)
from .registry_enrichment import (
    REGISTRY_ENRICHMENT_FOLDER_NAME,
    REGISTRY_ENRICHMENT_MANIFEST_NAME,
)


EXPORT_DATA_SCHEMA_VERSION = "export_data_v1"
EXPORT_DATA_STAGE = "EXPORT_DATA_CARD_MODEL_32_2"
EXPORT_DATA_FOLDER_NAME = "Экспорт_данных"
EXPORT_DATA_MAIN_SHEET_NAME = "Экспорт данных"
EXPORT_DATA_XLSX_NAME_TEMPLATE = "Экспорт_данных_проверка_листов_4_5_YYYY-MM-DD.xlsx"
EXPORT_DATA_CSV_NAME = "EXPORT_DATA.csv"
EXPORT_DATA_JSON_NAME = "EXPORT_DATA.json"
EXPORT_DATA_TECHNICAL_JSON_NAME = "EXPORT_DATA_TECHNICAL_MATRIX.json"
EXPORT_DATA_HANDOFF_MD_NAME = "EXPORT_DATA_HANDOFF.md"
EXPORT_DATA_REPORT_MD_NAME = "EXPORT_DATA_REPORT.md"
EXPORT_DATA_NORMATIVE_READINESS_EVIDENCE = "NORMATIVE_EXPORT_READINESS"

# Backward-compatible internal names.  These names are not shown in the UI.
SHEET6_CHECK_DATA_EXPORT_SCHEMA_VERSION = EXPORT_DATA_SCHEMA_VERSION
SHEET6_EXPORT_STAGE = EXPORT_DATA_STAGE
SHEET6_EXPORT_FOLDER_NAME = EXPORT_DATA_FOLDER_NAME
SHEET6_CHECK_DATA_EXPORT_XLSX_NAME = EXPORT_DATA_XLSX_NAME_TEMPLATE
SHEET6_CHECK_DATA_EXPORT_CSV_NAME = EXPORT_DATA_CSV_NAME
SHEET6_CHECK_DATA_EXPORT_JSON_NAME = EXPORT_DATA_JSON_NAME
SHEET6_TECHNICAL_READINESS_MATRIX_JSON_NAME = EXPORT_DATA_TECHNICAL_JSON_NAME
SHEET6_GPT_HANDOFF_MD_NAME = EXPORT_DATA_HANDOFF_MD_NAME
SHEET6_EXPORT_REPORT_MD_NAME = EXPORT_DATA_REPORT_MD_NAME
SHEET6_CLIENT_SHEET_NAME = EXPORT_DATA_MAIN_SHEET_NAME

EXPORT_DATA_COLUMNS: tuple[str, ...] = (
    "№",
    "Метка",
    "Приоритет",
    "Номер закупки",
    "Закон",
    "Заказчик / предмет",
    "НМЦК / цена контракта",
    "ОКПД2 / КТРУ",
    "Нормативный статус",
    "Связь с контрактом",
    "Документы",
    "На что обратить внимание",
    "Статус использования",
    "Комментарий пользователя",
)
SHEET6_CLIENT_COLUMNS = EXPORT_DATA_COLUMNS

EXPORT_DATA_SERVICE_COLUMNS: tuple[str, ...] = (
    "Закон / приоритет",
    "НМЦК / цена",
    "Номер контракта",
    "Поставщик",
    "Протоколы",
    "КТРУ заказчика",
    "Подходящие КТРУ",
    "Более подходящая позиция",
    "Характеристики выбранного КТРУ",
    "Дополнительные характеристики в ТЗ",
    "Обоснование дополнительных характеристик",
    "Связь с ПП №1875",
    "Нормативное наблюдение",
    "Источники подтверждения",
    "Ссылка на извещение",
    "Ссылка на контракт",
    "Ссылка на документы",
    "Ссылка на КТРУ",
    "Служебный признак",
    "Уровень доказанности",
    "Номер закупки чистый",
    "Документы текст",
    "КТРУ текст",
    "Ссылка на протокол",
    "Ссылка на спецификацию",
    "Примечание",
    "Статус использования исходный",
)

EXPORT_DATA_ALL_COLUMNS: tuple[str, ...] = EXPORT_DATA_COLUMNS + EXPORT_DATA_SERVICE_COLUMNS

EXPORT_DATA_TABLE_HEADER_ROW = 27
EXPORT_DATA_TABLE_FIRST_ROW = 28
EXPORT_DATA_SELECTION_CELL = "B4"

EXPORT_DATA_USAGE_STATUSES: tuple[str, ...] = (
    "Использовать",
    "Использовать с ограничениями",
    "Не использовать без проверки",
    "Исключить",
    "Не применимо",
)

EXPORT_DATA_LABELS: tuple[str, ...] = (
    "Готово",
    "Наблюдение",
    "Нужна проверка",
    "Не проверено",
)

EXPORT_DATA_KTRU_STATUSES: tuple[str, ...] = (
    "КТРУ указан заказчиком",
    "КТРУ подтверждён",
    "КТРУ указан, есть нормативное наблюдение",
    "КТРУ указан, найден более подходящий вариант",
    "КТРУ указан, есть риск по характеристикам",
    "КТРУ не указан, возможна обязательность",
    "КТРУ не применимо",
    "КТРУ не проверен из-за недостатка данных",
)

EXPORT_DATA_EVIDENCE_LEVELS: tuple[str, ...] = (
    "Высокий",
    "Средний",
    "Низкий",
    "manual_review",
    "данных недостаточно",
)

EXPORT_DATA_CARD_GROUPS: tuple[tuple[str, tuple[str, ...]], ...] = (
    (
        "Сводка",
        (
            "Номер закупки",
            "Метка",
            "Закон / приоритет",
            "Заказчик / предмет",
            "НМЦК / цена",
            "Статус использования",
        ),
    ),
    (
        "Связь и документы",
        (
            "Связь с контрактом",
            "Номер контракта",
            "Поставщик",
            "Протоколы",
            "Документы (ссылка)",
            "Извещение (ссылка)",
        ),
    ),
    (
        "КТРУ и нормативка",
        (
            "КТРУ заказчика",
            "Подходящие КТРУ",
            "Более подходящая позиция",
            "Характеристики выбранного КТРУ",
            "Дополнительные характеристики в ТЗ",
            "Связь с ПП №1875",
        ),
    ),
    (
        "Наблюдение и действие",
        (
            "Нормативное наблюдение",
            "На что обратить внимание",
            "Уровень доказанности",
            "Источники подтверждения",
            "Карточка КТРУ (ссылка)",
            "Служебный признак",
        ),
    ),
)

KTRU_DETAIL_COLUMNS: tuple[str, ...] = (
    "№",
    "ОКПД2-группа",
    "Наименование позиции из листа 4 / ЕИС",
    "КТРУ примененный",
    "Наименование примененного КТРУ",
    "КТРУ найден в источнике",
    "Дата начала обязательного применения",
    "Обязателен на дату закупки",
    "Описание / характеристики есть",
    "Применимые КТРУ-кандидаты",
    "Похожие КТРУ",
    "Основание подбора",
    "Уровень близости",
    "Риск / наблюдение",
    "Можно использовать в аналитике",
    "Источник доказательства",
    "Служебный флаг",
)

SHEET6_TECHNICAL_FIELDS: tuple[str, ...] = (
    "no",
    "control_node",
    "checked_link",
    "sheet4_data",
    "sheet5_data",
    "updated_documents",
    "okpd2_ktru",
    "normative_status_measure",
    "determination_method",
    "demand_confirmation",
    "data_reliability_level",
    "observation",
    "transfer_to_recommendations",
    "source_evidence",
    "example_objects",
    "service_flag",
)

REFERENCE_VALUES: dict[str, tuple[str, ...]] = {
    "Статус использования": EXPORT_DATA_USAGE_STATUSES,
    "Метка": EXPORT_DATA_LABELS,
    "Статус КТРУ": EXPORT_DATA_KTRU_STATUSES,
    "Уровень доказанности": EXPORT_DATA_EVIDENCE_LEVELS,
    "Нормативный статус": (
        "1875: Запрет",
        "1875: Ограничение",
        "1875: Преимущество",
        "1875: Условия допуска",
        "1875: Ограничение / Запрет",
        "719/ГИСП: Требуется подтверждение",
        "719/ГИСП: Запись найдена",
        "719/ГИСП: Запись не найдена",
        "Не выявлено",
        "По документам",
        "Требует проверки",
        "Ошибка проверки",
        "Не проверено",
        "Не применимо",
    ),
    "Источник определения": (
        "Извещение ЕИС",
        "Текст извещения",
        "ТЗ / описание объекта закупки",
        "Проект контракта",
        "Протокол",
        "Контракт / спецификация",
        "Документы исполнения",
        "Лист 4",
        "Лист 5",
        "Предварительно по ОКПД2",
        "Предварительно по КТРУ",
        "Предварительно по наименованию",
        "ГИСП / реестр",
        "GPT-анализ",
        "Не проверено",
        "Ошибка",
    ),
    "Связь с контрактом": (
        "Связь подтверждена",
        "Контракт не найден в загруженных данных",
        "Связь требует проверки",
        "данных недостаточно",
        "не применимо",
    ),
    "Документы": (
        "Открыть документы",
        "документы найдены без ссылки",
        "документы не скачаны",
        "данных недостаточно",
        "не применимо",
    ),
    "Подтверждение спроса": (
        "Подтвержден контрактом",
        "Закупочный спрос без контракта",
        "Смежный спрос",
        "Не подтвержден",
        "Требует проверки",
        "Ошибка сопоставления",
    ),
    "Уровень надежности данных": (
        "Высокий",
        "Средний",
        "Низкий",
        "Требуется ручная проверка",
        "Недостаточно данных",
        "Ошибка проверки",
    ),
    "Итоговая пригодность": (
        "Использовать в расчетах",
        "Использовать с ограничениями",
        "Использовать только справочно",
        "Требуется ручная проверка",
        "Недостаточно данных",
        "Исключить из расчета",
    ),
    "Служебные флаги КТРУ": (
        "ktru_present",
        "ktru_missing",
        "ktru_applied_by_customer",
        "ktru_candidate_found",
        "ktru_mandatory_on_notice_date",
        "ktru_not_mandatory_on_notice_date",
        "similar_ktru_found",
        "alternative_ktru_with_characteristics_found",
        "ktru_no_characteristics",
        "empty_ktru_selected_possible_risk",
        "extra_characteristics_detected",
        "extra_characteristics_justification_missing",
        "ktru_1875_restricted_extra_characteristics",
        "manual_ktru_review_required",
    ),
}

EXPORT_READY = "export_data_ready"
EXPORT_READY_WITH_WARNINGS = "export_data_ready_with_warnings"
EXPORT_NOT_READY = "export_data_not_ready"

CHECK_RESULT_CONFIRMED = "Подтверждено"
CHECK_RESULT_PARTIAL = "Подтверждено частично"
CHECK_RESULT_REVIEW = "Требует ручной проверки"
CHECK_RESULT_NOT_CHECKED = "Не проверено"
CHECK_RESULT_ERROR = "Ошибка проверки"
CHECK_RESULT_NOT_APPLICABLE = "Не применимо"
CHECK_RESULT_INSUFFICIENT = "Недостаточно данных"

RELIABILITY_HIGH = "Высокий"
RELIABILITY_MEDIUM = "Средний"
RELIABILITY_LOW = "Низкий"
RELIABILITY_REVIEW = "Требуется ручная проверка"
RELIABILITY_INSUFFICIENT = "Недостаточно данных"

FORBIDDEN_EXPORT_WORDING = (
    "нарушение установлено",
    "заказчик нарушил",
    "точно неправомерно",
    "точно не соответствует",
    "закупка незаконна",
    "ктру выбран неправомерно",
    "товар клиента",
    "эталонный товар",
    "данные клиента",
    "сопоставление с товаром клиента",
    "клиентский товар",
    "client item",
)


class Sheet6CheckDataExportValidationError(Exception):
    """Raised when export-data inputs or outputs are unsafe."""


@dataclass(frozen=True)
class Sheet6CheckDataExportPaths:
    export_root: Path
    xlsx: Path
    csv: Path
    json: Path
    technical_readiness_json: Path
    gpt_handoff_md: Path
    export_report_md: Path

    def to_json_dict(self) -> dict[str, str]:
        return {
            "export_root": str(self.export_root),
            "xlsx": str(self.xlsx),
            "csv": str(self.csv),
            "json": str(self.json),
            "technical_readiness_json": str(self.technical_readiness_json),
            "gpt_handoff_md": str(self.gpt_handoff_md),
            "export_report_md": str(self.export_report_md),
        }


class Sheet6CheckDataExportService:
    """Build the user-facing export named "Экспорт данных"."""

    def __init__(
        self,
        project_root: Path | str,
        *,
        project_service: ProjectService | None = None,
        now_provider: Callable[[], datetime] | None = None,
    ) -> None:
        self.project_service = project_service or ProjectService()
        self.project: ProjectInfo = self.project_service.open_project(Path(project_root))
        self.project_root = self.project.project_root.resolve()
        self.working_sheets_root = self.project.folders["working_sheets"]
        self.clean_registers_root = self.project.folders["clean_registers"]
        self.app_working_docs_root = self.project.folders["app_working_docs"]
        self.output_root = self.working_sheets_root / EXPORT_DATA_FOLDER_NAME
        self.now_provider = now_provider or (lambda: datetime.now(timezone.utc))
        self._assert_inside_project(self.output_root)

    def build_sheet6_export(
        self,
        rows: Sequence[Mapping[str, Any]] | Sequence[Any] | None = None,
        *,
        row_ids: Sequence[str] | None = None,
        scope_label: str = "ui_all_rows",
    ) -> dict[str, Any]:
        """Aggregate available project artifacts and write export-data outputs."""

        created_at = _utc_iso(self.now_provider)
        paths = self._paths(created_at=created_at)
        for path in paths.to_json_dict().values():
            self._assert_output_path(Path(path))

        provided_rows = [_mapping(row) for row in (rows or [])]
        selected_ids = {_text(item) for item in (row_ids or []) if _text(item)}
        if selected_ids:
            provided_rows = [
                row
                for row in provided_rows
                if _text(row.get("row_id")) in selected_ids
                or _text(row.get("procurement_number")) in selected_ids
                or _text(row.get("purchase_number")) in selected_ids
                or _text(row.get("contract_registry_number")) in selected_ids
            ]

        inputs = self._collect_inputs(provided_rows=provided_rows)
        readiness = _export_data_readiness(inputs)
        if not readiness["can_export"]:
            return _blocked_payload(
                paths=paths,
                project_root=self.project_root,
                created_at=created_at,
                scope_label=scope_label,
                readiness=readiness,
            )

        records = self._build_table_records(inputs)
        export_rows = [record["export"] for record in records]
        technical_rows = [self._technical_row(record) for record in records]
        ktru_rows = self._build_ktru_detail_rows(inputs)
        safe_wording = _validate_safe_wording(
            {
                "export_rows": export_rows,
                "technical_rows": technical_rows,
                "ktru_rows": ktru_rows,
            }
        )
        status = (
            EXPORT_READY_WITH_WARNINGS
            if readiness["status"] == EXPORT_READY_WITH_WARNINGS or not safe_wording["is_valid"]
            else EXPORT_READY
        )
        summary = self._build_summary(
            created_at=created_at,
            scope_label=scope_label,
            inputs=inputs,
            export_rows=export_rows,
            technical_rows=technical_rows,
            ktru_rows=ktru_rows,
            safe_wording=safe_wording,
            readiness={**readiness, "status": status},
        )

        payload = {
            "schema_version": EXPORT_DATA_SCHEMA_VERSION,
            "stage": EXPORT_DATA_STAGE,
            "status": status,
            "export_data_status": status,
            "can_export_data": True,
            "project_root": str(self.project_root),
            "created_at": created_at,
            "scope_label": scope_label,
            "export_sheet_name": EXPORT_DATA_MAIN_SHEET_NAME,
            "export_columns": list(EXPORT_DATA_COLUMNS),
            "service_columns": list(EXPORT_DATA_SERVICE_COLUMNS),
            "all_columns": list(EXPORT_DATA_ALL_COLUMNS),
            "client_sheet_name": EXPORT_DATA_MAIN_SHEET_NAME,
            "client_columns": list(EXPORT_DATA_COLUMNS),
            "technical_fields": list(SHEET6_TECHNICAL_FIELDS),
            "items": export_rows,
            "summary": summary,
            "paths": paths.to_json_dict(),
            "live_network_used": False,
            "documents_downloaded": False,
            "source_excel_changed": False,
            "excel_sheets": [
                EXPORT_DATA_MAIN_SHEET_NAME,
                "КТРУ детализация",
                "Справочники",
                "README",
            ],
        }
        technical_payload = {
            "schema_version": EXPORT_DATA_SCHEMA_VERSION,
            "stage": EXPORT_DATA_STAGE,
            "status": status,
            "project_root": str(self.project_root),
            "created_at": created_at,
            "technical_fields": list(SHEET6_TECHNICAL_FIELDS),
            "items": technical_rows,
            "summary": summary,
        }

        paths.export_root.mkdir(parents=True, exist_ok=True)
        self._write_xlsx(paths.xlsx, export_rows, ktru_rows, created_at=created_at, inputs=inputs, summary=summary)
        self._write_csv(paths.csv, export_rows, EXPORT_DATA_COLUMNS)
        _write_json(paths.json, payload)
        _write_json(paths.technical_readiness_json, technical_payload)
        paths.gpt_handoff_md.write_text(
            _gpt_handoff_markdown(payload, technical_payload),
            encoding="utf-8",
        )
        paths.export_report_md.write_text(
            _export_report_markdown(payload),
            encoding="utf-8",
        )
        return payload

    def get_sheet6_export(self) -> dict[str, Any]:
        paths = self._paths()
        if not paths.json.is_file():
            return {
                "schema_version": EXPORT_DATA_SCHEMA_VERSION,
                "stage": EXPORT_DATA_STAGE,
                "status": "missing",
                "export_data_status": EXPORT_NOT_READY,
                "can_export_data": False,
                "project_root": str(self.project_root),
                "message": "Экспорт данных еще не создан.",
                "paths": paths.to_json_dict(),
                "export_sheet_name": EXPORT_DATA_MAIN_SHEET_NAME,
                "export_columns": list(EXPORT_DATA_COLUMNS),
                "client_sheet_name": EXPORT_DATA_MAIN_SHEET_NAME,
                "client_columns": list(EXPORT_DATA_COLUMNS),
                "technical_fields": list(SHEET6_TECHNICAL_FIELDS),
            }
        payload = _load_json(paths.json)
        payload.setdefault("paths", paths.to_json_dict())
        return payload

    def _collect_inputs(self, *, provided_rows: Sequence[Mapping[str, Any]]) -> dict[str, Any]:
        registry_payload, registry_path = self._load_latest_registry()
        registry_rows = provided_rows or _items(registry_payload, "rows")
        dataset_counts = _dataset_counts(registry_payload, registry_rows)
        procurement_rows = [row for row in registry_rows if _dataset_type(row) == "procurement"]
        contract_rows = [row for row in registry_rows if _dataset_type(row) == "contract"]

        enrichment_payload = _load_json_if_exists(
            self.app_working_docs_root / REGISTRY_ENRICHMENT_FOLDER_NAME / REGISTRY_ENRICHMENT_MANIFEST_NAME
        )
        enrichment_rows = _items(enrichment_payload, "rows")

        contract_link_payload = _load_json_if_exists(
            self.working_sheets_root / CONTRACT_LINK_EXPORT_FOLDER_NAME / CONTRACT_LINK_UI_EXPORT_JSON_NAME
        )
        contract_link_rows = _items_any(contract_link_payload)
        completed_procurement_payload = _load_json_if_exists(
            self.working_sheets_root / CONTRACT_LINK_EXPORT_FOLDER_NAME / COMPLETED_PROCUREMENT_REGISTRY_EXPORT_JSON_NAME
        )
        completed_contract_payload = _load_json_if_exists(
            self.working_sheets_root / CONTRACT_LINK_EXPORT_FOLDER_NAME / COMPLETED_CONTRACT_REGISTRY_EXPORT_JSON_NAME
        )
        contract_link_summary = _load_json_if_exists(
            self.working_sheets_root / CONTRACT_LINK_EXPORT_FOLDER_NAME / CONTRACT_LINK_COVERAGE_SUMMARY_JSON_NAME
        )

        protocol_payload = _load_json_if_exists(
            self.working_sheets_root / PROTOCOL_EXPORT_FOLDER_NAME / PROTOCOL_UI_EXPORT_JSON_NAME
        )
        protocol_rows = _items_any(protocol_payload)
        supplier_results = _load_json_if_exists(
            self.working_sheets_root / PROTOCOL_EXPORT_FOLDER_NAME / PROTOCOL_SUPPLIER_RESULTS_EXPORT_JSON_NAME
        )
        contract_evidence = _load_json_if_exists(
            self.working_sheets_root / PROTOCOL_EXPORT_FOLDER_NAME / PROTOCOL_CONTRACT_EVIDENCE_EXPORT_JSON_NAME
        )

        normative_payload = _load_json_if_exists(
            self.working_sheets_root / NORMATIVE_EXPORT_FOLDER_NAME / NORMATIVE_UI_EXPORT_JSON_NAME
        )
        normative_rows = _items(normative_payload, "ui_rows")
        normative_sheet6 = _load_json_if_exists(
            self.working_sheets_root / NORMATIVE_EXPORT_FOLDER_NAME / NORMATIVE_SHEET6_READINESS_JSON_NAME
        )
        normative_sheet6_rows = _items(normative_sheet6, "items")

        missing_inputs = []
        if not contract_link_rows:
            missing_inputs.append("contract_link_ui_export")
        if not protocol_rows:
            missing_inputs.append("protocol_ui_export")

        warnings = _collect_warning_texts(
            registry_rows,
            enrichment_rows,
            contract_link_rows,
            protocol_rows,
            normative_rows,
            normative_sheet6_rows,
        )
        conflict_rows = [
            row
            for row in contract_link_rows
            if "конфликт" in _text(row.get("user_contract_link_status")).casefold()
            or "conflict" in _text(row.get("technical_match_status")).casefold()
        ]
        manual_rows = [
            row
            for row in [*enrichment_rows, *contract_link_rows, *protocol_rows, *normative_rows]
            if bool(row.get("manual_review_required"))
            or _list(row.get("manual_review_flags"))
            or "провер" in _text(row.get("user_contract_link_status")).casefold()
            or "провер" in _text(row.get("status_label")).casefold()
        ]

        return {
            "registry_payload": registry_payload,
            "registry_path": registry_path,
            "registry_rows": registry_rows,
            "dataset_counts": dataset_counts,
            "procurement_rows": procurement_rows,
            "contract_rows": contract_rows,
            "enrichment_payload": enrichment_payload,
            "enrichment_rows": enrichment_rows,
            "contract_link_payload": contract_link_payload,
            "contract_link_rows": contract_link_rows,
            "completed_procurement_rows": _items_any(completed_procurement_payload),
            "completed_contract_rows": _items_any(completed_contract_payload),
            "contract_link_summary": contract_link_summary,
            "protocol_payload": protocol_payload,
            "protocol_rows": protocol_rows,
            "supplier_results": _items_any(supplier_results),
            "contract_evidence": _items_any(contract_evidence),
            "normative_payload": normative_payload,
            "normative_rows": normative_rows,
            "normative_sheet6": normative_sheet6,
            "normative_sheet6_rows": normative_sheet6_rows,
            "missing_inputs": missing_inputs,
            "warnings": warnings,
            "conflict_rows": conflict_rows,
            "manual_rows": manual_rows,
            "source_evidence": self._source_evidence(registry_path=registry_path),
        }

    def _build_table_records(self, inputs: Mapping[str, Any]) -> list[dict[str, Any]]:
        registry_rows = _list_of_mappings(inputs.get("registry_rows"))
        procurement_rows = _list_of_mappings(inputs.get("procurement_rows"))
        if not procurement_rows:
            procurement_rows = [row for row in registry_rows if _dataset_type(row) != "contract"] or registry_rows

        contract_rows = _list_of_mappings(inputs.get("contract_rows"))
        enrichment_rows = _list_of_mappings(inputs.get("enrichment_rows"))
        contract_link_rows = _list_of_mappings(inputs.get("contract_link_rows"))
        protocol_rows = _list_of_mappings(inputs.get("protocol_rows"))
        normative_rows = _list_of_mappings(inputs.get("normative_rows"))
        normative_sheet6_rows = _list_of_mappings(inputs.get("normative_sheet6_rows"))
        supplier_results = _list_of_mappings(inputs.get("supplier_results"))
        contract_evidence = _list_of_mappings(inputs.get("contract_evidence"))

        if not procurement_rows:
            return []

        records: list[dict[str, Any]] = []
        for index, row in enumerate(procurement_rows, start=1):
            contract_link = _find_related_row(row, contract_link_rows)
            contract_row = _find_related_contract_row(row, contract_rows, contract_link)
            enrichment_row = _find_related_row(row, enrichment_rows)
            protocol_row = _find_related_row(row, protocol_rows)
            supplier_row = _find_related_row(row, supplier_results)
            contract_evidence_row = _find_related_row(row, contract_evidence)
            normative_row = _find_related_row(row, normative_rows)
            readiness_row = _find_related_row(row, normative_sheet6_rows)
            export_row = _export_data_table_row(
                index=index,
                procurement_row=row,
                contract_row=contract_row,
                contract_link_row=contract_link,
                enrichment_row=enrichment_row,
                protocol_row=protocol_row,
                supplier_row=supplier_row,
                contract_evidence_row=contract_evidence_row,
                normative_row=normative_row,
                readiness_row=readiness_row,
            )
            records.append(
                {
                    "export": {column: _safe_cell_text(export_row.get(column, "")) for column in EXPORT_DATA_ALL_COLUMNS},
                    "source": export_row.get("Источники подтверждения", ""),
                    "service_flag": export_row.get("Служебный признак", ""),
                }
            )
        return records

    def _build_control_records(self, inputs: Mapping[str, Any]) -> list[dict[str, Any]]:
        registry_rows = _list_of_mappings(inputs.get("registry_rows"))
        procurement_rows = _list_of_mappings(inputs.get("procurement_rows"))
        contract_rows = _list_of_mappings(inputs.get("contract_rows"))
        enrichment_rows = _list_of_mappings(inputs.get("enrichment_rows"))
        contract_link_rows = _list_of_mappings(inputs.get("contract_link_rows"))
        protocol_rows = _list_of_mappings(inputs.get("protocol_rows"))
        normative_rows = _list_of_mappings(inputs.get("normative_rows"))
        normative_sheet6_rows = _list_of_mappings(inputs.get("normative_sheet6_rows"))
        supplier_results = _list_of_mappings(inputs.get("supplier_results"))
        contract_evidence = _list_of_mappings(inputs.get("contract_evidence"))
        missing_inputs = _list(inputs.get("missing_inputs"))
        warnings = _list(inputs.get("warnings"))
        conflict_count = len(_list_of_mappings(inputs.get("conflict_rows")))
        manual_count = len(_list_of_mappings(inputs.get("manual_rows")))

        procurement_count = len(procurement_rows)
        contract_count = len(contract_rows)
        linked_count = _linked_contracts_count(contract_link_rows)
        unlinked_count = max(procurement_count - linked_count, 0)
        documents_found = sum(_int(_mapping(row.get("documents")).get("document_count") or row.get("documents_found_count")) for row in enrichment_rows)
        protocol_count = len(protocol_rows)
        okpd2_count = sum(1 for row in registry_rows if _list(row.get("okpd2")) or _list(row.get("okpd2_codes")))
        ktru_count = sum(1 for row in registry_rows if _list(row.get("ktru")) or _list(row.get("ktru_codes")))
        ktru_candidate_count = sum(1 for row in normative_sheet6_rows if _text(row.get("ktru_candidate")) or _text(row.get("applicable_ktru_candidates")))
        normative_flags_count = sum(
            1
            for row in normative_rows
            if _text(row.get("normative_status")) and _text(row.get("normative_status")) != "Не проверено"
        )
        pp719_gisp_count = sum(
            1
            for row in normative_rows
            if _text(row.get("pp719_status")) not in {"", "pp719_not_checked"}
            or _text(row.get("gisp_status")) not in {"", "gisp_not_checked"}
        )
        price_count = sum(
            1
            for row in registry_rows
            if _text(row.get("nmck")) or _text(row.get("contract_price")) or _text(row.get("price"))
        )
        supplier_count = sum(1 for row in registry_rows if _text(row.get("supplier_name")) or _text(row.get("winner_display_name")))
        protocol_winner_count = sum(1 for row in protocol_rows if _text(row.get("winner_display_status")) or _text(row.get("winner_display_name")))
        noise_count = sum(1 for row in registry_rows if _text(row.get("row_quality_status")) in {"manual_review_required", "error"})
        error_count = sum(1 for item in warnings if "error" in _text(item).casefold() or "ошиб" in _text(item).casefold())

        records: list[dict[str, Any]] = []
        add = records.append
        add(_record(
            1,
            "Полнота идентификаторов листа 4",
            "Номер закупки, закон, источник, дата, заказчик и исходная строка.",
            _count_text("закупочных строк", procurement_count),
            _count_text("контрактных строк", contract_count),
            _count_text("обновленных строк", len(enrichment_rows)),
            _okpd_ktru_text(okpd2_count, ktru_count),
            "Не применимо",
            "Лист 4",
            _demand_confirmation_text(linked_count, procurement_count),
            RELIABILITY_MEDIUM if procurement_count else RELIABILITY_INSUFFICIENT,
            "Идентификаторы используются как техническая основа связки; пустые значения требуют ручной проверки.",
            "Переносить только строки с устойчивым номером закупки и source_row_number.",
            "normalized_registry",
            _examples(procurement_rows),
            "sheet4_identifiers_checked" if procurement_count else "missing_sheet4_rows",
        ))
        add(_record(
            2,
            "Связка лист 4 ↔ лист 5",
            "Сопоставление закупок с контрактами / договорами.",
            _count_text("закупочных строк", procurement_count),
            _count_text("контрактных строк", contract_count),
            _contract_link_limitation(contract_link_rows, conflict_count, manual_count),
            _okpd_ktru_text(okpd2_count, ktru_count),
            "Не применимо",
            "Лист 4 + Лист 5 + CONTRACT_LINK_UI_EXPORT",
            _demand_confirmation_text(linked_count, procurement_count),
            RELIABILITY_MEDIUM if linked_count else RELIABILITY_REVIEW,
            f"Связано: {linked_count}; не связано с листом 5: {unlinked_count}; конфликтов: {conflict_count}.",
            "Использовать подтвержденные связки; несвязанные строки переносить как ограничение.",
            "CONTRACT_LINK_UI_EXPORT",
            _examples(contract_link_rows or registry_rows),
            "contract_link_confirmed" if linked_count else "manual_contract_link_review_required",
        ))
        add(_record(
            3,
            "Предмет закупки и предмет контракта",
            "Сравнение предмета закупки с предметом контракта / договора.",
            _subject_sample(procurement_rows),
            _subject_sample(contract_rows),
            "Сопоставление строится по доступным нормализованным строкам и completed registries.",
            _okpd_ktru_text(okpd2_count, ktru_count),
            "Не выявлено" if linked_count else "Требует проверки",
            "Лист 4 / Лист 5",
            _demand_confirmation_text(linked_count, procurement_count),
            RELIABILITY_MEDIUM if linked_count else RELIABILITY_REVIEW,
            "Автоматическое совпадение предметов не является юридическим выводом.",
            "Переносить наблюдение о совпадении предметов только вместе с источником.",
            "normalized_registry / completed registries",
            _examples([*procurement_rows[:2], *contract_rows[:2]]),
            "subject_link_review_required",
        ))
        add(_record(
            4,
            "Источник / площадка / закон",
            "Источник строки, закон и маршрут получения сведений.",
            _law_source_text(procurement_rows),
            _law_source_text(contract_rows),
            _enrichment_status_text(enrichment_rows),
            "Предварительно по ОКПД2" if okpd2_count else "Не проверено",
            "Не выявлено",
            "Извещение ЕИС",
            _demand_confirmation_text(linked_count, procurement_count),
            RELIABILITY_MEDIUM if enrichment_rows else RELIABILITY_REVIEW,
            "Площадки вне поддержанного маршрута сохраняются как ручная проверка.",
            "Переносить источник и закон как контекст, а не как вывод.",
            "registry_enrichment_v1",
            _examples(enrichment_rows or registry_rows),
            "source_route_checked" if enrichment_rows else "registry_enrichment_not_started",
        ))
        add(_record(
            5,
            "Документы закупки",
            "Наличие извещения, ТЗ, проекта контракта, спецификации и приложений.",
            _count_text("закупочных строк", procurement_count),
            "Лист 5 не подтверждает документы закупки напрямую.",
            f"Документы найдены: {documents_found}; скачано: 0.",
            _okpd_ktru_text(okpd2_count, ktru_count),
            "По документам" if documents_found else "Не проверено",
            "Текст извещения",
            "Требует проверки" if not documents_found else _demand_confirmation_text(linked_count, procurement_count),
            RELIABILITY_MEDIUM if documents_found else RELIABILITY_INSUFFICIENT,
            "Если документы не скачаны или не прочитаны, выводы по документам считаются ограниченными.",
            "Переносить только факт наличия документов и ограничения проверки.",
            "registry_enrichment_v1",
            _examples(enrichment_rows or procurement_rows),
            "documents_found" if documents_found else "documents_missing_or_not_read",
        ))
        add(_record(
            6,
            "Протоколы и результат процедуры",
            "Участники, допуск / отклонение, победитель, цена победителя и итог процедуры.",
            _count_text("закупочных строк", procurement_count),
            _count_text("контрактных строк", contract_count),
            _protocol_limitation(protocol_rows, supplier_results, contract_evidence),
            "Не применимо",
            "По документам" if protocol_count else "Не проверено",
            "Протокол",
            _demand_confirmation_text(linked_count, procurement_count),
            RELIABILITY_MEDIUM if protocol_count else RELIABILITY_INSUFFICIENT,
            "Без протокола сведения о победителе и цене считаются ограниченными.",
            "Переносить только подтвержденные протоколом наблюдения.",
            "PROTOCOL_UI_EXPORT",
            _examples(protocol_rows),
            "protocol_found" if protocol_count else "protocol_missing",
        ))
        add(_record(
            7,
            "Контракт / договор",
            "Наличие реестрового номера, предмета, цены, поставщика и статуса.",
            _count_text("закупочных строк", procurement_count),
            _count_text("контрактов / договоров", contract_count),
            _count_text("completed contract rows", len(_list_of_mappings(inputs.get("completed_contract_rows")))),
            _okpd_ktru_text(okpd2_count, ktru_count),
            "Не применимо",
            "Контракт / спецификация",
            _demand_confirmation_text(linked_count, procurement_count),
            RELIABILITY_MEDIUM if contract_count else RELIABILITY_INSUFFICIENT,
            "Если лист 5 отсутствует или не сопоставлен, спрос по контрактам не подтверждается.",
            "Переносить контрактные сведения только при наличии номера и связи с закупкой.",
            "Лист 5 / completed contract registry",
            _examples(contract_rows),
            "contract_found" if contract_count else "contract_missing",
        ))
        add(_record(
            8,
            "Документы исполнения",
            "Документы исполнения, спецификации, страны происхождения и подтверждающие файлы.",
            "Лист 4 не содержит документы исполнения.",
            _count_text("контрактных строк", contract_count),
            _count_text("contract evidence rows", len(contract_evidence)),
            "Не применимо",
            "По документам" if contract_evidence else "Не проверено",
            "Документы исполнения",
            _demand_confirmation_text(linked_count, procurement_count),
            RELIABILITY_MEDIUM if contract_evidence else RELIABILITY_INSUFFICIENT,
            "Документы исполнения не считаются подтвержденными без отдельного evidence.",
            "Переносить как ограничение, если evidence отсутствует.",
            "PROTOCOL_CONTRACT_EVIDENCE_EXPORT / Лист 5",
            _examples(contract_evidence or contract_rows),
            "execution_docs_found" if contract_evidence else "execution_docs_missing",
        ))
        add(_record(
            9,
            "ОКПД2",
            "Наличие и пригодность ОКПД2 для классификации и нормативной проверки.",
            _count_text("строк с ОКПД2", okpd2_count),
            _count_text("контрактных строк", contract_count),
            _enrichment_status_text(enrichment_rows),
            _okpd_ktru_text(okpd2_count, ktru_count),
            "Предварительно по ОКПД2" if okpd2_count else "Не проверено",
            "Предварительно по ОКПД2",
            _demand_confirmation_text(linked_count, procurement_count),
            RELIABILITY_MEDIUM if okpd2_count else RELIABILITY_INSUFFICIENT,
            "ОКПД2 является классификационным признаком и требует проверки спорных строк.",
            "Переносить код и источник, не усиливая до финального вывода.",
            "normalized_registry / normative exports",
            _examples(procurement_rows),
            "okpd2_present" if okpd2_count else "okpd2_missing",
        ))
        add(_record(
            10,
            "КТРУ — что найдено",
            "Найденный, примененный или отсутствующий КТРУ.",
            _count_text("строк с КТРУ", ktru_count),
            _count_text("контрактных строк", contract_count),
            _enrichment_status_text(enrichment_rows),
            f"КТРУ найден: {ktru_count}; применимые кандидаты: {ktru_candidate_count}.",
            "Предварительно по КТРУ" if ktru_count else "Не проверено",
            "Предварительно по КТРУ",
            _demand_confirmation_text(linked_count, procurement_count),
            RELIABILITY_MEDIUM if ktru_count else RELIABILITY_REVIEW,
            "Найденный похожий или применимый КТРУ является основанием для ручной проверки.",
            "Переносить код, статус и основание подбора в КТРУ детализацию.",
            "Нормативная проверка / registry enrichment",
            _examples(procurement_rows),
            "ktru_present" if ktru_count else "ktru_missing",
        ))
        add(_record(
            11,
            "КТРУ — обязательность и применение",
            "Проверка обязательности КТРУ на дату закупки и признаков применения.",
            _count_text("закупочных строк", procurement_count),
            _count_text("контрактных строк", contract_count),
            _count_text("readiness rows", len(normative_sheet6_rows)),
            f"КТРУ найден: {ktru_count}; КТРУ обязателен: {_mandatory_ktru_count(normative_sheet6_rows)}.",
            "Требует проверки" if ktru_count else "Не проверено",
            "Предварительно по КТРУ",
            _demand_confirmation_text(linked_count, procurement_count),
            RELIABILITY_REVIEW,
            "Обязательность КТРУ не считается ошибкой без проверки даты и характеристик.",
            "Переносить как риск-флаг или ограничение для ручной проверки.",
            EXPORT_DATA_NORMATIVE_READINESS_EVIDENCE,
            _examples(normative_sheet6_rows or procurement_rows),
            "manual_ktru_review_required",
        ))
        add(_record(
            12,
            "Характеристики и описание КТРУ",
            "Наличие характеристик, описания объекта закупки и признаков дополнительных характеристик.",
            _count_text("закупочных строк", procurement_count),
            "Лист 5 может подтвердить только спецификацию при наличии evidence.",
            "Характеристики проверяются по доступным enrichment/normative полям.",
            f"КТРУ найден: {ktru_count}.",
            "Требует проверки",
            "ТЗ / описание объекта закупки",
            _demand_confirmation_text(linked_count, procurement_count),
            RELIABILITY_REVIEW,
            "Дополнительные характеристики требуют проверки основания и контекста.",
            "Переносить только подтвержденные признаки и ограничения.",
            "registry enrichment / normative exports",
            _examples(enrichment_rows or procurement_rows),
            "ktru_characteristics_review_required",
        ))
        add(_record(
            13,
            "Национальный режим / ПП №1875",
            "Наличие запрета, ограничения, преимущества или условий допуска.",
            _count_text("закупочных строк", procurement_count),
            "Лист 5 используется для подтверждения спроса, а не для самостоятельного нормативного вывода.",
            _count_text("normative rows", len(normative_rows)),
            _okpd_ktru_text(okpd2_count, ktru_count),
            "1875: Ограничение / Запрет" if normative_flags_count else "Не проверено",
            "Предварительно по ОКПД2",
            _demand_confirmation_text(linked_count, procurement_count),
            RELIABILITY_MEDIUM if normative_flags_count else RELIABILITY_REVIEW,
            "Выявленный признак по ПП №1875 требует проверки документов и дат.",
            "Переносить как риск/ограничение, не как юридический вывод.",
            "NORMATIVE_UI_EXPORT",
            _examples(normative_rows),
            "pp1875_risk_found" if normative_flags_count else "pp1875_not_checked",
        ))
        add(_record(
            14,
            "ПП №719 / ГИСП / происхождение",
            "Признаки страны происхождения, реестровой записи и подтверждения по ГИСП.",
            _count_text("закупочных строк", procurement_count),
            _count_text("контрактных строк", contract_count),
            _count_text("719/ГИСП rows", pp719_gisp_count or len(normative_sheet6_rows)),
            _okpd_ktru_text(okpd2_count, ktru_count),
            "719/ГИСП: Требуется подтверждение" if pp719_gisp_count or normative_sheet6_rows else "Не проверено",
            "ГИСП / реестр",
            _demand_confirmation_text(linked_count, procurement_count),
            RELIABILITY_REVIEW,
            "Отсутствие записи или несколько кандидатов не являются финальным выводом.",
            "Переносить как основание для ручной проверки происхождения.",
            EXPORT_DATA_NORMATIVE_READINESS_EVIDENCE,
            _examples(normative_sheet6_rows or normative_rows),
            "manual_origin_review_required",
        ))
        add(_record(
            15,
            "Цена и единица измерения",
            "НМЦК, цена контракта, цена победителя, единица измерения и сопоставимость.",
            _count_text("строк с ценой / НМЦК", price_count),
            _count_text("контрактных строк", contract_count),
            "Единицы измерения требуют отдельной нормализации, если нет evidence.",
            "Не применимо",
            "Не выявлено" if price_count else "Не проверено",
            "Лист 4 / Лист 5 / Протокол",
            _demand_confirmation_text(linked_count, procurement_count),
            RELIABILITY_MEDIUM if price_count else RELIABILITY_INSUFFICIENT,
            "Цены нельзя переносить в расчет без проверки единиц и состава поставки.",
            "Переносить только с ограничениями по единицам измерения.",
            "normalized_registry / protocol outputs",
            _examples(registry_rows),
            "price_present" if price_count else "price_missing",
        ))
        add(_record(
            16,
            "Поставщик / победитель",
            "Поставщик по контракту, победитель по протоколу, ИНН и confidence.",
            _count_text("закупочных строк", procurement_count),
            _count_text("контрактных строк с поставщиком", supplier_count),
            _count_text("protocol winner rows", protocol_winner_count),
            "Не применимо",
            "По документам" if supplier_count or protocol_winner_count else "Не проверено",
            "Протокол / Контракт / спецификация",
            _demand_confirmation_text(linked_count, procurement_count),
            RELIABILITY_MEDIUM if supplier_count or protocol_winner_count else RELIABILITY_INSUFFICIENT,
            "Участник не считается поставщиком без договора или надежного протокольного evidence.",
            "Переносить поставщика только с источником и статусом подтверждения.",
            "PROTOCOL_UI_EXPORT / Лист 5",
            _examples([*contract_rows[:2], *protocol_rows[:2]]),
            "supplier_or_winner_found" if supplier_count or protocol_winner_count else "supplier_winner_missing",
        ))
        add(_record(
            17,
            "Непригодные / шумовые данные",
            "Ошибки, неполные строки, manual review и шумовые признаки.",
            _count_text("закупочных строк", procurement_count),
            _count_text("контрактных строк", contract_count),
            _limitations_summary(missing_inputs, warnings, manual_count, conflict_count),
            _okpd_ktru_text(okpd2_count, ktru_count),
            "Ошибка проверки" if error_count else "Требует проверки",
            "Ошибка" if error_count else "Не проверено",
            "Требует проверки",
            RELIABILITY_REVIEW if manual_count or warnings else RELIABILITY_MEDIUM,
            f"Manual review: {manual_count}; warnings: {len(warnings)}; шумовые строки: {noise_count}.",
            "Исключать или использовать только справочно после ручной проверки.",
            "all_available_outputs",
            _examples(_list_of_mappings(inputs.get("manual_rows")) or registry_rows),
            "manual_review_required" if manual_count or warnings else "noise_not_detected",
        ))
        add(_record(
            18,
            "Итоговая пригодность данных",
            "Можно ли переносить данные в аналитические выводы и рекомендации.",
            _count_text("закупочных строк", procurement_count),
            _count_text("контрактных строк", contract_count),
            _overall_limitation(procurement_count, contract_count, missing_inputs, warnings),
            _okpd_ktru_text(okpd2_count, ktru_count),
            "Требует проверки" if warnings or missing_inputs else "Не выявлено",
            "Лист 4 / Лист 5 / нормативная проверка",
            _demand_confirmation_text(linked_count, procurement_count),
            RELIABILITY_MEDIUM if procurement_count and contract_count else RELIABILITY_REVIEW,
            "Экспорт фиксирует качество и доказательность данных, а не юридическую квалификацию действий заказчика.",
            "Использовать подтвержденные наблюдения, ограничения и риск-флаги.",
            "ETAP 32.1 export summary",
            _examples(registry_rows),
            "export_data_ready_with_warnings" if warnings or missing_inputs else "export_data_ready",
        ))
        return records

    def _build_ktru_detail_rows(self, inputs: Mapping[str, Any]) -> list[dict[str, Any]]:
        procurement_rows = _list_of_mappings(inputs.get("procurement_rows"))
        normative_by_row = _index_by_row_id(_list_of_mappings(inputs.get("normative_rows")))
        readiness_by_row = _index_by_row_id(_list_of_mappings(inputs.get("normative_sheet6_rows")))
        if not procurement_rows:
            return [
                _ktru_detail_row(
                    no=1,
                    okpd2="Недостаточно данных",
                    subject="Закупочные строки листа 4 не найдены",
                    applied_ktru="",
                    applied_name="",
                    found_status="КТРУ не проверено",
                    mandatory_date="",
                    mandatory_status="Требуется ручная проверка КТРУ",
                    has_characteristics="Не проверено",
                    candidates="",
                    similar="",
                    basis="Недостаточно данных",
                    closeness="",
                    risk="Недостаточно данных",
                    can_use="Требуется ручная проверка",
                    evidence="normalized_registry",
                    flag="manual_ktru_review_required",
                )
            ]

        rows: list[dict[str, Any]] = []
        for index, row in enumerate(procurement_rows, start=1):
            row_id = _text(row.get("row_id"))
            normative = normative_by_row.get(row_id, {})
            readiness = readiness_by_row.get(row_id, {})
            okpd2 = _join_values(_list(row.get("okpd2")) or _list(row.get("okpd2_codes")))
            ktru = _join_values(_list(row.get("ktru")) or _list(row.get("ktru_codes")))
            candidates = _join_values(
                _list(readiness.get("applicable_ktru_candidates"))
                or _list(normative.get("applicable_ktru_candidates"))
                or _list(readiness.get("ktru_candidate"))
            )
            similar = _join_values(_list(readiness.get("similar_ktru")) or _list(normative.get("similar_ktru")))
            mandatory = _mandatory_status(readiness, normative, bool(ktru))
            found_status = "КТРУ найден" if ktru else ("Похожие КТРУ найдены" if similar else "КТРУ не найден")
            flag = "ktru_present" if ktru else ("similar_ktru_found" if similar else "ktru_missing")
            if mandatory == "Требуется ручная проверка КТРУ":
                flag = "manual_ktru_review_required"
            rows.append(
                _ktru_detail_row(
                    no=index,
                    okpd2=okpd2 or "Не указано",
                    subject=_subject(row),
                    applied_ktru=ktru,
                    applied_name=_text(readiness.get("ktru_name") or normative.get("ktru_name")),
                    found_status=found_status,
                    mandatory_date=_text(readiness.get("mandatory_start_date") or normative.get("mandatory_start_date")),
                    mandatory_status=mandatory,
                    has_characteristics=_characteristics_status(readiness, normative),
                    candidates=candidates,
                    similar=similar,
                    basis=_text(readiness.get("match_basis") or normative.get("source_evidence") or "Предварительно по ОКПД2 / КТРУ"),
                    closeness=_text(readiness.get("closeness_level") or normative.get("confidence_label") or normative.get("normative_confidence")),
                    risk=_text(readiness.get("risk") or normative.get("risk_flags") or "Требуется ручная проверка КТРУ"),
                    can_use=_text(readiness.get("can_use_in_analytics") or normative.get("can_use_in_analytics") or "Использовать с ограничениями"),
                    evidence=_row_evidence(row),
                    flag=flag,
                )
            )
        return rows

    def _build_summary(
        self,
        *,
        created_at: str,
        scope_label: str,
        inputs: Mapping[str, Any],
        export_rows: Sequence[Mapping[str, Any]],
        technical_rows: Sequence[Mapping[str, Any]],
        ktru_rows: Sequence[Mapping[str, Any]],
        safe_wording: Mapping[str, Any],
        readiness: Mapping[str, Any],
    ) -> dict[str, Any]:
        status_counts = Counter(_text(row.get("Уровень надежности данных")) for row in export_rows)
        flag_counts = Counter(_text(row.get("Служебный флаг")) for row in export_rows)
        kpi = _kpi_rows(inputs)
        return {
            "schema_version": EXPORT_DATA_SCHEMA_VERSION,
            "stage": EXPORT_DATA_STAGE,
            "status": readiness.get("status", EXPORT_READY_WITH_WARNINGS),
            "created_at": created_at,
            "scope_label": scope_label,
            "control_rows_count": len(export_rows),
            "technical_rows_count": len(technical_rows),
            "export_columns_count": len(EXPORT_DATA_COLUMNS),
            "client_columns_count": len(EXPORT_DATA_COLUMNS),
            "technical_fields_count": len(SHEET6_TECHNICAL_FIELDS),
            "ktru_detail_rows_count": len(ktru_rows),
            "registry_rows_count": len(_list_of_mappings(inputs.get("registry_rows"))),
            "procurement_rows_count": len(_list_of_mappings(inputs.get("procurement_rows"))),
            "contract_rows_count": len(_list_of_mappings(inputs.get("contract_rows"))),
            "enrichment_rows_count": len(_list_of_mappings(inputs.get("enrichment_rows"))),
            "contract_link_rows_count": len(_list_of_mappings(inputs.get("contract_link_rows"))),
            "protocol_rows_count": len(_list_of_mappings(inputs.get("protocol_rows"))),
            "normative_rows_count": len(_list_of_mappings(inputs.get("normative_rows"))),
            "warnings_count": len(_list(inputs.get("warnings"))) + len(_list(readiness.get("warnings"))),
            "missing_inputs_count": len(_list(inputs.get("missing_inputs"))),
            "missing_inputs": _list(inputs.get("missing_inputs")),
            "controlled_warnings": [*_list(readiness.get("warnings")), *_list(inputs.get("warnings"))][:80],
            "readiness": dict(readiness),
            "kpi": kpi,
            "reliability_counts": dict(status_counts),
            "service_flag_counts": dict(flag_counts),
            "safe_wording": dict(safe_wording),
            "source_evidence": _list(inputs.get("source_evidence")),
            "live_network_used": False,
            "documents_downloaded": False,
            "source_excel_changed": False,
            "excel_sheets": [
                EXPORT_DATA_MAIN_SHEET_NAME,
                "КТРУ детализация",
                "Справочники",
                "README",
            ],
        }

    def _write_xlsx(
        self,
        path: Path,
        rows: Sequence[Mapping[str, Any]],
        ktru_rows: Sequence[Mapping[str, Any]],
        *,
        created_at: str,
        inputs: Mapping[str, Any],
        summary: Mapping[str, Any],
    ) -> None:
        self._assert_output_path(path)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = EXPORT_DATA_MAIN_SHEET_NAME

        title_fill = PatternFill("solid", fgColor="1F4E79")
        section_fill = PatternFill("solid", fgColor="EAF3F8")
        header_fill = PatternFill("solid", fgColor="D9EAF7")
        thin = Side(style="thin", color="B7C9D6")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(EXPORT_DATA_COLUMNS))
        title = sheet.cell(row=1, column=1)
        title.value = EXPORT_DATA_MAIN_SHEET_NAME
        title.font = Font(bold=True, color="FFFFFF", size=16)
        title.fill = title_fill
        title.alignment = Alignment(horizontal="center", vertical="center")
        sheet.row_dimensions[1].height = 26

        sheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(EXPORT_DATA_COLUMNS))
        note = sheet.cell(row=3, column=1)
        note.value = (
            "Универсальная проверочная выгрузка по результатам обработки листа 4 и листа 5: "
            "закупки, позиции, документы, протоколы, контракты, ОКПД2, КТРУ, национальный режим, "
            "подтверждение спроса и служебные риск-флаги."
        )
        note.alignment = Alignment(wrap_text=True, vertical="top")

        _write_key_value_block(
            sheet,
            start_row=5,
            start_col=1,
            title="Контекст экспорта",
            rows=_context_rows(inputs, created_at, summary),
            border=border,
            section_fill=section_fill,
        )
        _write_key_value_block(
            sheet,
            start_row=5,
            start_col=5,
            title="Контрольные показатели",
            rows=_kpi_rows(inputs),
            border=border,
            section_fill=section_fill,
        )

        rules_start = 24
        sheet.merge_cells(start_row=rules_start, start_column=1, end_row=rules_start, end_column=len(EXPORT_DATA_COLUMNS))
        cell = sheet.cell(row=rules_start, column=1)
        cell.value = "Правила использования"
        cell.font = Font(bold=True)
        cell.fill = section_fill
        for offset, rule in enumerate(_usage_rules(), start=1):
            sheet.merge_cells(
                start_row=rules_start + offset,
                start_column=1,
                end_row=rules_start + offset,
                end_column=len(EXPORT_DATA_COLUMNS),
            )
            rule_cell = sheet.cell(row=rules_start + offset, column=1)
            rule_cell.value = rule
            rule_cell.alignment = Alignment(wrap_text=True, vertical="top")

        header_row = rules_start + len(_usage_rules()) + 3
        for column_number, column_name in enumerate(EXPORT_DATA_COLUMNS, start=1):
            cell = sheet.cell(row=header_row, column=column_number)
            cell.value = column_name
            cell.font = Font(bold=True, color="1F2937")
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for row_offset, row_payload in enumerate(rows, start=1):
            excel_row = header_row + row_offset
            for column_number, column_name in enumerate(EXPORT_DATA_COLUMNS, start=1):
                cell = sheet.cell(row=excel_row, column=column_number)
                cell.value = row_payload.get(column_name, "")
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        widths = (8, 28, 34, 32, 32, 36, 28, 28, 26, 28, 24, 42, 42, 36, 32, 28)
        for column_number, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(column_number)].width = width
        last_column = get_column_letter(len(EXPORT_DATA_COLUMNS))
        last_row = header_row + len(rows)
        sheet.freeze_panes = f"A{header_row + 1}"
        sheet.auto_filter.ref = f"A{header_row}:{last_column}{last_row}"
        sheet.sheet_view.showGridLines = False

        self._write_ktru_sheet(workbook, ktru_rows, border=border, header_fill=header_fill)
        self._write_references_sheet(workbook, border=border, header_fill=header_fill, section_fill=section_fill)
        self._write_readme_sheet(workbook, inputs=inputs, summary=summary, border=border, section_fill=section_fill)

        path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(path)

    def _write_ktru_sheet(
        self,
        workbook: Workbook,
        rows: Sequence[Mapping[str, Any]],
        *,
        border: Border,
        header_fill: PatternFill,
    ) -> None:
        sheet = workbook.create_sheet("КТРУ детализация")
        for column_number, column_name in enumerate(KTRU_DETAIL_COLUMNS, start=1):
            cell = sheet.cell(row=1, column=column_number)
            cell.value = column_name
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row_offset, row_payload in enumerate(rows, start=2):
            for column_number, column_name in enumerate(KTRU_DETAIL_COLUMNS, start=1):
                cell = sheet.cell(row=row_offset, column=column_number)
                cell.value = row_payload.get(column_name, "")
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        widths = (8, 18, 42, 24, 34, 24, 24, 24, 26, 34, 34, 28, 20, 34, 28, 34, 28)
        for column_number, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(column_number)].width = width
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:{get_column_letter(len(KTRU_DETAIL_COLUMNS))}{max(2, len(rows) + 1)}"
        sheet.sheet_view.showGridLines = False

    def _write_references_sheet(
        self,
        workbook: Workbook,
        *,
        border: Border,
        header_fill: PatternFill,
        section_fill: PatternFill,
    ) -> None:
        sheet = workbook.create_sheet("Справочники")
        sheet.cell(row=1, column=1).value = "Справочник"
        sheet.cell(row=1, column=2).value = "Допустимое значение"
        for column in (1, 2):
            cell = sheet.cell(row=1, column=column)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.border = border
        row_number = 2
        for category, values in REFERENCE_VALUES.items():
            sheet.cell(row=row_number, column=1).value = category
            sheet.cell(row=row_number, column=1).font = Font(bold=True)
            sheet.cell(row=row_number, column=1).fill = section_fill
            sheet.cell(row=row_number, column=1).border = border
            row_number += 1
            for value in values:
                sheet.cell(row=row_number, column=1).value = category
                sheet.cell(row=row_number, column=2).value = value
                sheet.cell(row=row_number, column=1).border = border
                sheet.cell(row=row_number, column=2).border = border
                row_number += 1
        sheet.column_dimensions["A"].width = 34
        sheet.column_dimensions["B"].width = 52
        sheet.freeze_panes = "A2"
        sheet.sheet_view.showGridLines = False

    def _write_readme_sheet(
        self,
        workbook: Workbook,
        *,
        inputs: Mapping[str, Any],
        summary: Mapping[str, Any],
        border: Border,
        section_fill: PatternFill,
    ) -> None:
        sheet = workbook.create_sheet("README")
        lines = [
            ("Что это", "Экспорт данных — универсальная проверочная выгрузка по результатам обработки листа 4 и листа 5."),
            ("Источники", _truncate_text("; ".join(_list(inputs.get("source_evidence"))) or "Доступные данные проекта.")),
            ("Не юридическое заключение", "Выгрузка фиксирует качество и доказательность данных, риск-флаги и ограничения. Она не устанавливает нарушение."),
            ("Ограничения", _aggregate_warnings(_list(summary.get("controlled_warnings"))) or "Ограничения не зафиксированы."),
            ("Как читать статусы", "Статусы показывают готовность данных, уровень надежности и необходимость ручной проверки."),
            ("Где смотреть КТРУ", "Подробности вынесены на лист «КТРУ детализация»."),
        ]
        for row_number, (title, text) in enumerate(lines, start=1):
            title_cell = sheet.cell(row=row_number, column=1)
            value_cell = sheet.cell(row=row_number, column=2)
            title_cell.value = title
            value_cell.value = text
            title_cell.font = Font(bold=True)
            title_cell.fill = section_fill
            title_cell.border = border
            value_cell.border = border
            value_cell.alignment = Alignment(wrap_text=True, vertical="top")
        sheet.column_dimensions["A"].width = 30
        sheet.column_dimensions["B"].width = 100
        sheet.sheet_view.showGridLines = False

    def _write_csv(self, path: Path, rows: Sequence[Mapping[str, Any]], fieldnames: Sequence[str]) -> None:
        self._assert_output_path(path)
        path.parent.mkdir(parents=True, exist_ok=True)
        with path.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=list(fieldnames), delimiter=";", extrasaction="ignore")
            writer.writeheader()
            for row in rows:
                writer.writerow({key: _csv_value(row.get(key)) for key in fieldnames})

    def _load_latest_registry(self) -> tuple[dict[str, Any], Path | None]:
        direct = self.clean_registers_root / NORMALIZED_REGISTRY_JSON_NAME
        candidates = [path for path in self.clean_registers_root.glob("normalized_registry_v1*.json") if path.is_file()]
        if direct.is_file():
            path = direct
        elif candidates:
            path = max(candidates, key=lambda item: item.stat().st_mtime)
        else:
            return {}, None
        return _load_json(path), path

    def _source_evidence(self, *, registry_path: Path | None) -> list[str]:
        items = []
        if registry_path is not None:
            items.append(_relative_to_project(registry_path, self.project_root))
        for path in (
            self.app_working_docs_root / REGISTRY_ENRICHMENT_FOLDER_NAME / REGISTRY_ENRICHMENT_MANIFEST_NAME,
            self.working_sheets_root / CONTRACT_LINK_EXPORT_FOLDER_NAME / CONTRACT_LINK_UI_EXPORT_JSON_NAME,
            self.working_sheets_root / PROTOCOL_EXPORT_FOLDER_NAME / PROTOCOL_UI_EXPORT_JSON_NAME,
            self.working_sheets_root / NORMATIVE_EXPORT_FOLDER_NAME / NORMATIVE_UI_EXPORT_JSON_NAME,
            self.working_sheets_root / NORMATIVE_EXPORT_FOLDER_NAME / NORMATIVE_SHEET6_READINESS_JSON_NAME,
        ):
            if path.is_file():
                items.append(_export_data_evidence_label(_relative_to_project(path, self.project_root)))
        return items or ["available_project_state_missing"]

    def _paths(self, *, created_at: str | None = None) -> Sheet6CheckDataExportPaths:
        if created_at:
            date_label = created_at[:10]
        else:
            date_label = self._latest_export_date_label()
        return Sheet6CheckDataExportPaths(
            export_root=self.output_root,
            xlsx=self.output_root / _export_xlsx_name(date_label),
            csv=self.output_root / EXPORT_DATA_CSV_NAME,
            json=self.output_root / EXPORT_DATA_JSON_NAME,
            technical_readiness_json=self.output_root / EXPORT_DATA_TECHNICAL_JSON_NAME,
            gpt_handoff_md=self.output_root / EXPORT_DATA_HANDOFF_MD_NAME,
            export_report_md=self.output_root / EXPORT_DATA_REPORT_MD_NAME,
        )

    def _latest_export_date_label(self) -> str:
        if self.output_root.is_dir():
            candidates = sorted(
                self.output_root.glob("Экспорт_данных_проверка_листов_4_5_*.xlsx"),
                key=lambda item: item.stat().st_mtime,
                reverse=True,
            )
            if candidates:
                suffix = candidates[0].stem.rsplit("_", 1)[-1]
                if len(suffix) == 10:
                    return suffix
        return _utc_iso(self.now_provider)[:10]

    def _assert_output_path(self, path: Path) -> None:
        resolved = path.resolve()
        try:
            resolved.relative_to(self.project_root)
        except ValueError as exc:
            raise Sheet6CheckDataExportValidationError(
                f"Export-data output path outside user project is forbidden: {resolved}"
            ) from exc

    def _assert_inside_project(self, path: Path) -> None:
        resolved = path.resolve()
        try:
            resolved.relative_to(self.project_root)
        except ValueError as exc:
            raise Sheet6CheckDataExportValidationError(
                f"Path outside user project is forbidden for export-data outputs: {resolved}"
            ) from exc

    @staticmethod
    def _technical_row(record: Mapping[str, Any]) -> dict[str, Any]:
        row = _mapping(record.get("export"))
        technical = {
            "no": row.get("№", ""),
            "control_node": row.get("Контрольный узел", ""),
            "checked_link": row.get("Проверяемая связка", ""),
            "sheet4_data": row.get("Данные листа 4", ""),
            "sheet5_data": row.get("Данные листа 5", ""),
            "updated_documents": row.get("Обновленные сведения / документы", ""),
            "okpd2_ktru": row.get("ОКПД2 / КТРУ", ""),
            "normative_status_measure": row.get("Нормативный статус / мера", ""),
            "determination_method": row.get("Как определено", ""),
            "demand_confirmation": row.get("Подтверждение спроса", ""),
            "data_reliability_level": row.get("Уровень надежности данных", ""),
            "observation": row.get("Зафиксированное расхождение / наблюдение", ""),
            "transfer_to_recommendations": row.get("Что переносится в выводы / рекомендации", ""),
            "source_evidence": row.get("Источник / доказательство", ""),
            "example_objects": row.get("Объекты-примеры", ""),
            "service_flag": row.get("Служебный флаг", ""),
        }
        return {field: technical.get(field, "") for field in SHEET6_TECHNICAL_FIELDS}


def _record(
    no: int,
    control_node: str,
    checked_link: str,
    sheet4_data: str,
    sheet5_data: str,
    updated_documents: str,
    okpd2_ktru: str,
    normative_status: str,
    determination_method: str,
    demand_confirmation: str,
    reliability: str,
    observation: str,
    transfer: str,
    source_evidence: str,
    examples: str,
    service_flag: str,
) -> dict[str, Any]:
    row = {
        "№": no,
        "Контрольный узел": control_node,
        "Проверяемая связка": checked_link,
        "Данные листа 4": sheet4_data,
        "Данные листа 5": sheet5_data,
        "Обновленные сведения / документы": updated_documents,
        "ОКПД2 / КТРУ": okpd2_ktru,
        "Нормативный статус / мера": normative_status,
        "Как определено": determination_method,
        "Подтверждение спроса": demand_confirmation,
        "Уровень надежности данных": reliability,
        "Зафиксированное расхождение / наблюдение": observation,
        "Что переносится в выводы / рекомендации": transfer,
        "Источник / доказательство": source_evidence,
        "Объекты-примеры": examples,
        "Служебный флаг": service_flag,
    }
    return {"export": row, "source": source_evidence, "service_flag": service_flag}


def _export_data_table_row(
    *,
    index: int,
    procurement_row: Mapping[str, Any],
    contract_row: Mapping[str, Any],
    contract_link_row: Mapping[str, Any],
    enrichment_row: Mapping[str, Any],
    protocol_row: Mapping[str, Any],
    supplier_row: Mapping[str, Any],
    contract_evidence_row: Mapping[str, Any],
    normative_row: Mapping[str, Any],
    readiness_row: Mapping[str, Any],
) -> dict[str, Any]:
    procurement_number = _text(procurement_row.get("procurement_number") or procurement_row.get("purchase_number"))
    contract_number = _text(
        contract_link_row.get("contract_registry_number")
        or contract_row.get("contract_registry_number")
        or contract_row.get("contract_number")
        or contract_evidence_row.get("contract_registry_number")
    )
    customer_subject = _join_values(
        [
            procurement_row.get("customer_name"),
            _subject(procurement_row),
        ]
    )
    law_priority = _join_values([procurement_row.get("law") or procurement_row.get("law_type"), procurement_row.get("priority")])
    price_text = _join_values(
        [
            f"НМЦК: {_text(procurement_row.get('nmck') or procurement_row.get('price'))}"
            if _text(procurement_row.get("nmck") or procurement_row.get("price"))
            else "",
            f"контракт: {_text(contract_row.get('contract_price') or contract_row.get('price'))}"
            if _text(contract_row.get("contract_price") or contract_row.get("price"))
            else "",
        ]
    )
    ktru_text = _join_values(_list(procurement_row.get("ktru")) or _list(procurement_row.get("ktru_codes")))
    okpd2_text = _join_values(_list(procurement_row.get("okpd2")) or _list(procurement_row.get("okpd2_codes")))
    okpd_ktru = _join_values([f"ОКПД2: {okpd2_text}" if okpd2_text else "", f"КТРУ: {ktru_text}" if ktru_text else ""])

    normative_status = _text(
        readiness_row.get("normative_restriction")
        or readiness_row.get("mandatory_status")
        or normative_row.get("normative_status")
        or normative_row.get("pp1875_status")
        or normative_row.get("pp719_status")
        or normative_row.get("gisp_status")
    )
    contract_link_status = _text(contract_link_row.get("user_contract_link_status") or contract_link_row.get("technical_match_status"))
    if not contract_link_status:
        contract_link_status = "Связь подтверждена" if contract_number else "Контракт не найден в загруженных данных"

    documents = _mapping(enrichment_row.get("documents"))
    documents_count = _int(documents.get("document_count") or enrichment_row.get("documents_found_count"))
    documents_text = f"Документы найдены: {documents_count}" if documents_count else "документы не скачаны"
    protocol_text = _join_values(
        [
            protocol_row.get("protocol_status"),
            protocol_row.get("winner_display_status"),
            protocol_row.get("winner_display_name"),
        ]
    )
    supplier_text = _join_values(
        [
            contract_row.get("supplier_name"),
            supplier_row.get("supplier_name"),
            protocol_row.get("winner_display_name"),
            contract_evidence_row.get("contract_supplier_name"),
        ]
    )

    warnings = _collect_warning_texts(
        [procurement_row],
        [contract_link_row] if contract_link_row else [],
        [enrichment_row] if enrichment_row else [],
        [protocol_row] if protocol_row else [],
        [normative_row] if normative_row else [],
        [readiness_row] if readiness_row else [],
    )
    manual_review = any(
        bool(row.get("manual_review_required")) or _list(row.get("manual_review_flags"))
        for row in (
            procurement_row,
            contract_link_row,
            enrichment_row,
            protocol_row,
            normative_row,
            readiness_row,
        )
        if row
    )
    label = "Нужна проверка" if manual_review or warnings else "Готово"
    reliability = RELIABILITY_REVIEW if manual_review else (RELIABILITY_MEDIUM if documents_count or contract_number else RELIABILITY_INSUFFICIENT)
    usage_status = "Не использовать без проверки" if manual_review else ("Использовать с ограничениями" if warnings else "Использовать")
    attention = _aggregate_warnings(warnings) or _text(readiness_row.get("risk") or normative_row.get("risk") or "")
    if not attention and not contract_number:
        attention = "Контракт не найден в загруженных данных"

    source_evidence = _join_values(
        [
            _row_evidence(procurement_row),
            _row_evidence(contract_row) if contract_row else "",
            _row_evidence(contract_link_row) if contract_link_row else "",
            _row_evidence(normative_row) if normative_row else "",
        ]
    )
    service_flag = _text(
        readiness_row.get("service_flag")
        or normative_row.get("service_flag")
        or contract_link_row.get("technical_match_status")
        or ("manual_review_required" if manual_review else "export_data_row_ready")
    )

    row = {
        "№": index,
        "Метка": label,
        "Приоритет": _text(procurement_row.get("priority")),
        "Номер закупки": procurement_number,
        "Закон": _text(procurement_row.get("law") or procurement_row.get("law_type")),
        "Заказчик / предмет": customer_subject,
        "НМЦК / цена контракта": price_text,
        "ОКПД2 / КТРУ": okpd_ktru,
        "Нормативный статус": normative_status or "Не проверено",
        "Связь с контрактом": contract_link_status,
        "Документы": documents_text,
        "На что обратить внимание": attention,
        "Статус использования": usage_status,
        "Комментарий пользователя": "",
        "Закон / приоритет": law_priority,
        "НМЦК / цена": price_text,
        "Номер контракта": contract_number,
        "Поставщик": supplier_text,
        "Протоколы": protocol_text or "Не проверено",
        "КТРУ заказчика": ktru_text,
        "Подходящие КТРУ": _join_values(
            _list(readiness_row.get("applicable_ktru_candidates"))
            or _list(normative_row.get("applicable_ktru_candidates"))
            or _list(readiness_row.get("ktru_candidate"))
        ),
        "Более подходящая позиция": _join_values(_list(readiness_row.get("similar_ktru")) or _list(normative_row.get("similar_ktru"))),
        "Характеристики выбранного КТРУ": _characteristics_status(readiness_row, normative_row),
        "Дополнительные характеристики в ТЗ": _text(readiness_row.get("extra_characteristics") or normative_row.get("extra_characteristics")),
        "Обоснование дополнительных характеристик": _text(
            readiness_row.get("extra_characteristics_justification") or normative_row.get("extra_characteristics_justification")
        ),
        "Связь с ПП №1875": _text(normative_row.get("pp1875_status") or normative_row.get("normative_status")),
        "Нормативное наблюдение": attention or normative_status,
        "Источники подтверждения": source_evidence,
        "Ссылка на извещение": _text(procurement_row.get("notice_url") or procurement_row.get("url") or enrichment_row.get("notice_url")),
        "Ссылка на контракт": _text(contract_row.get("contract_url") or contract_link_row.get("contract_url")),
        "Ссылка на документы": _text(enrichment_row.get("documents_url") or enrichment_row.get("source_url")),
        "Ссылка на КТРУ": _text(readiness_row.get("ktru_url") or normative_row.get("ktru_url")),
        "Служебный признак": service_flag,
        "Уровень доказанности": reliability,
        "Номер закупки чистый": procurement_number,
        "Документы текст": documents_text,
        "КТРУ текст": ktru_text,
        "Ссылка на протокол": _text(protocol_row.get("protocol_url")),
        "Ссылка на спецификацию": _text(contract_row.get("specification_url") or contract_evidence_row.get("specification_url")),
        "Примечание": _aggregate_warnings(warnings, max_lines=10),
        "Статус использования исходный": usage_status,
    }
    return {column: row.get(column, "") for column in EXPORT_DATA_ALL_COLUMNS}


def _ktru_detail_row(
    *,
    no: int,
    okpd2: str,
    subject: str,
    applied_ktru: str,
    applied_name: str,
    found_status: str,
    mandatory_date: str,
    mandatory_status: str,
    has_characteristics: str,
    candidates: str,
    similar: str,
    basis: str,
    closeness: str,
    risk: str,
    can_use: str,
    evidence: str,
    flag: str,
) -> dict[str, Any]:
    return {
        "№": no,
        "ОКПД2-группа": okpd2,
        "Наименование позиции из листа 4 / ЕИС": subject,
        "КТРУ примененный": applied_ktru,
        "Наименование примененного КТРУ": applied_name,
        "КТРУ найден в источнике": found_status,
        "Дата начала обязательного применения": mandatory_date,
        "Обязателен на дату закупки": mandatory_status,
        "Описание / характеристики есть": has_characteristics,
        "Применимые КТРУ-кандидаты": candidates,
        "Похожие КТРУ": similar,
        "Основание подбора": basis,
        "Уровень близости": closeness,
        "Риск / наблюдение": risk,
        "Можно использовать в аналитике": can_use,
        "Источник доказательства": evidence,
        "Служебный флаг": flag,
    }


def _export_data_readiness(inputs: Mapping[str, Any]) -> dict[str, Any]:
    registry_rows = _list_of_mappings(inputs.get("registry_rows"))
    enrichment_payload = _mapping_or_none(inputs.get("enrichment_payload"))
    enrichment_rows = _list_of_mappings(inputs.get("enrichment_rows"))
    normative_payload = _mapping_or_none(inputs.get("normative_payload"))
    normative_sheet6 = _mapping_or_none(inputs.get("normative_sheet6"))
    normative_rows = _list_of_mappings(inputs.get("normative_rows"))
    normative_sheet6_rows = _list_of_mappings(inputs.get("normative_sheet6_rows"))
    missing_inputs = _list(inputs.get("missing_inputs"))
    warnings = []

    if not registry_rows:
        return _readiness(False, "registry_not_imported", "Сначала загрузите табличный реестр.", warnings)
    if not enrichment_payload:
        return _readiness(
            False,
            "registry_enrichment_not_started",
            "Экспорт данных доступен после обновления данных и нормативной проверки.",
            warnings,
        )

    enrichment_status = _text(enrichment_payload.get("status"))
    if enrichment_status in {"failed", "error", "blocked"}:
        return _readiness(False, "registry_enrichment_failed", "Обновление данных завершилось ошибкой.", warnings)
    if not enrichment_rows:
        return _readiness(False, "registry_enrichment_incomplete", "Обновление данных не содержит обработанных строк.", warnings)
    refreshed_count = _int(enrichment_payload.get("refreshed_rows_count") or len(enrichment_rows))
    if refreshed_count < len(registry_rows):
        warnings.append("registry_enrichment_incomplete")

    if not normative_payload and not normative_sheet6:
        return _readiness(
            False,
            "normative_check_not_started",
            "Экспорт данных доступен после обновления данных и нормативной проверки.",
            warnings,
        )
    normative_status = _text((normative_payload or normative_sheet6 or {}).get("status"))
    if normative_status in {"failed", "error", "blocked"}:
        return _readiness(False, "normative_check_failed", "Нормативная проверка завершилась ошибкой.", warnings)
    if not normative_rows and not normative_sheet6_rows:
        return _readiness(False, "normative_check_incomplete", "Нормативная проверка не содержит проверенных строк.", warnings)

    if enrichment_status == "ok_with_warnings" or normative_status == "ok_with_warnings":
        warnings.append("controlled_warnings_present")
    if missing_inputs:
        warnings.extend(f"missing_input:{item}" for item in missing_inputs)
    if _list(inputs.get("warnings")):
        warnings.append("row_level_warnings_present")

    status = EXPORT_READY_WITH_WARNINGS if warnings else EXPORT_READY
    message = (
        "Экспорт данных готов с предупреждениями."
        if warnings
        else "Экспорт данных готов."
    )
    return _readiness(True, status, message, warnings)


def _readiness(can_export: bool, status: str, message: str, warnings: Sequence[str]) -> dict[str, Any]:
    return {
        "can_export": can_export,
        "status": status,
        "message": message,
        "warnings": list(warnings),
    }


def _blocked_payload(
    *,
    paths: Sheet6CheckDataExportPaths,
    project_root: Path,
    created_at: str,
    scope_label: str,
    readiness: Mapping[str, Any],
) -> dict[str, Any]:
    return {
        "schema_version": EXPORT_DATA_SCHEMA_VERSION,
        "stage": EXPORT_DATA_STAGE,
        "status": readiness.get("status", EXPORT_NOT_READY),
        "export_data_status": readiness.get("status", EXPORT_NOT_READY),
        "can_export_data": False,
        "message": readiness.get("message") or "Экспорт данных недоступен.",
        "project_root": str(project_root),
        "created_at": created_at,
        "scope_label": scope_label,
        "export_sheet_name": EXPORT_DATA_MAIN_SHEET_NAME,
        "export_columns": list(EXPORT_DATA_COLUMNS),
        "client_sheet_name": EXPORT_DATA_MAIN_SHEET_NAME,
        "client_columns": list(EXPORT_DATA_COLUMNS),
        "technical_fields": list(SHEET6_TECHNICAL_FIELDS),
        "paths": paths.to_json_dict(),
        "summary": {
            "schema_version": EXPORT_DATA_SCHEMA_VERSION,
            "stage": EXPORT_DATA_STAGE,
            "status": readiness.get("status", EXPORT_NOT_READY),
            "readiness": dict(readiness),
            "control_rows_count": 0,
            "export_columns_count": len(EXPORT_DATA_COLUMNS),
            "warnings_count": len(_list(readiness.get("warnings"))),
            "missing_inputs_count": 0,
        },
        "live_network_used": False,
        "documents_downloaded": False,
        "source_excel_changed": False,
    }


def _write_key_value_block(
    sheet: Any,
    *,
    start_row: int,
    start_col: int,
    title: str,
    rows: Sequence[tuple[str, Any]],
    border: Border,
    section_fill: PatternFill,
) -> None:
    sheet.cell(row=start_row, column=start_col).value = title
    sheet.cell(row=start_row, column=start_col).font = Font(bold=True)
    sheet.cell(row=start_row, column=start_col).fill = section_fill
    sheet.cell(row=start_row, column=start_col).border = border
    sheet.cell(row=start_row, column=start_col + 1).fill = section_fill
    sheet.cell(row=start_row, column=start_col + 1).border = border
    for offset, (label, value) in enumerate(rows, start=1):
        label_cell = sheet.cell(row=start_row + offset, column=start_col)
        value_cell = sheet.cell(row=start_row + offset, column=start_col + 1)
        label_cell.value = label
        value_cell.value = _csv_value(value)
        label_cell.font = Font(bold=True)
        label_cell.border = border
        value_cell.border = border
        value_cell.alignment = Alignment(wrap_text=True, vertical="top")
        sheet.row_dimensions[start_row + offset].height = 18


def _context_rows(inputs: Mapping[str, Any], created_at: str, summary: Mapping[str, Any]) -> list[tuple[str, Any]]:
    registry_rows = _list_of_mappings(inputs.get("registry_rows"))
    procurement_rows = _list_of_mappings(inputs.get("procurement_rows"))
    contract_rows = _list_of_mappings(inputs.get("contract_rows"))
    laws = sorted({_text(row.get("law") or row.get("law_type")) for row in registry_rows if _text(row.get("law") or row.get("law_type"))})
    priorities = sorted({_text(row.get("priority")) for row in registry_rows if _text(row.get("priority"))})
    source4 = _source_files(procurement_rows)
    source5 = _source_files(contract_rows)
    readiness = _mapping(summary.get("readiness"))
    return [
        ("Проект", _text(inputs.get("project_name")) or "Активный пользовательский проект"),
        ("Дата экспорта", created_at[:10]),
        ("Режим экспорта", "safe/offline; документы и live-download не запускаются"),
        ("Период данных", _date_range(registry_rows)),
        ("Источник листа 4", source4 or "Не определено"),
        ("Источник листа 5", source5 or "Не определено"),
        ("Закон / контур", ", ".join(laws) if laws else "Не определено"),
        ("Приоритеты", ", ".join(priorities) if priorities else "Не указаны"),
        ("Статус обновления сведений", _text(readiness.get("status"))),
        ("Статус нормативной проверки", "выполнена" if readiness.get("can_export") else "не выполнена"),
        ("Ограничения", _aggregate_warnings(_list(summary.get("controlled_warnings"))) or "Нет"),
    ]


def _kpi_rows(inputs: Mapping[str, Any]) -> list[tuple[str, Any]]:
    registry_rows = _list_of_mappings(inputs.get("registry_rows"))
    procurement_rows = _list_of_mappings(inputs.get("procurement_rows"))
    contract_rows = _list_of_mappings(inputs.get("contract_rows"))
    enrichment_rows = _list_of_mappings(inputs.get("enrichment_rows"))
    contract_link_rows = _list_of_mappings(inputs.get("contract_link_rows"))
    protocol_rows = _list_of_mappings(inputs.get("protocol_rows"))
    normative_rows = _list_of_mappings(inputs.get("normative_rows"))
    normative_sheet6_rows = _list_of_mappings(inputs.get("normative_sheet6_rows"))
    documents_found = sum(_int(_mapping(row.get("documents")).get("document_count") or row.get("documents_found_count")) for row in enrichment_rows)
    linked = _linked_contracts_count(contract_link_rows)
    okpd2_count = sum(1 for row in registry_rows if _list(row.get("okpd2")) or _list(row.get("okpd2_codes")))
    ktru_count = sum(1 for row in registry_rows if _list(row.get("ktru")) or _list(row.get("ktru_codes")))
    normative_count = sum(
        1
        for row in normative_rows
        if _text(row.get("normative_status")) and _text(row.get("normative_status")) != "Не проверено"
    )
    manual_count = len(_list_of_mappings(inputs.get("manual_rows")))
    warnings = _list(inputs.get("warnings"))
    return [
        ("Строк листа 4 в экспорте", len(procurement_rows)),
        ("Строк листа 5 в экспорте", len(contract_rows)),
        ("Связано лист 4 ↔ лист 5", linked),
        ("Не связано с листом 5", max(len(procurement_rows) - linked, 0)),
        ("Документы найдены / скачаны", f"{documents_found} / 0"),
        ("Протоколы найдены", len(protocol_rows)),
        ("Контракты / договоры найдены", len(contract_rows)),
        ("Документы исполнения найдены", len(_list_of_mappings(inputs.get("contract_evidence")))),
        ("ОКПД2 найден", okpd2_count),
        ("КТРУ найден", ktru_count),
        ("КТРУ обязателен", _mandatory_ktru_count(normative_sheet6_rows)),
        ("Применимые КТРУ найдены", sum(1 for row in normative_sheet6_rows if _text(row.get("applicable_ktru_candidates")))),
        ("Похожие КТРУ найдены", sum(1 for row in normative_sheet6_rows if _text(row.get("similar_ktru")))),
        ("Нормативное ограничение выявлено", normative_count),
        ("Требует ручной проверки", manual_count),
        ("Ошибки проверки", sum(1 for item in warnings if "error" in _text(item).casefold() or "ошиб" in _text(item).casefold())),
    ]


def _usage_rules() -> tuple[str, ...]:
    return (
        "Экспорт фиксирует качество и доказательность данных, а не юридическую квалификацию действий заказчика.",
        "Если данные получены расчетно, это отражается в колонке «Как определено».",
        "Если лист 5 отсутствует или не сопоставлен, подтверждение спроса по контрактам не считается подтвержденным.",
        "Если документы не скачаны или не прочитаны, выводы по документам считаются ограниченными.",
        "Если найден похожий или применимый КТРУ, это не означает автоматическую ошибку заказчика; это основание для ручной проверки.",
        "В итоговые выводы переносятся только подтвержденные наблюдения, ограничения и риск-флаги.",
    )


def _contract_link_limitation(contract_link_rows: Sequence[Mapping[str, Any]], conflict_count: int, manual_count: int) -> str:
    if not contract_link_rows:
        return "Связь закупка-контракт не найдена; требуется ручная проверка."
    status_counts = Counter(_text(row.get("user_contract_link_status")) for row in contract_link_rows)
    parts = [f"Строк связи: {len(contract_link_rows)}"]
    if status_counts:
        parts.append("статусы: " + ", ".join(f"{status}: {count}" for status, count in status_counts.items() if status))
    if conflict_count:
        parts.append(f"конфликтные строки: {conflict_count}")
    if manual_count:
        parts.append(f"ручная проверка: {manual_count}")
    return "; ".join(parts) + "."


def _protocol_limitation(
    protocol_rows: Sequence[Mapping[str, Any]],
    supplier_results: Sequence[Mapping[str, Any]],
    contract_evidence: Sequence[Mapping[str, Any]],
) -> str:
    if not protocol_rows and not supplier_results and not contract_evidence:
        return "Протоколы и результаты процедуры не найдены в готовых outputs."
    return (
        f"Protocol UI rows: {len(protocol_rows)}; "
        f"supplier-results evidence: {len(supplier_results)}; "
        f"embedded contract evidence: {len(contract_evidence)}."
    )


def _limitations_summary(
    missing_inputs: Sequence[str],
    warnings: Sequence[str],
    manual_count: int,
    conflict_count: int,
) -> str:
    parts: list[str] = []
    if missing_inputs:
        parts.append("missing inputs: " + ", ".join(missing_inputs))
    if warnings:
        parts.append(f"warnings: {len(warnings)}")
    if manual_count:
        parts.append(f"manual review rows: {manual_count}")
    if conflict_count:
        parts.append(f"conflict rows: {conflict_count}")
    return "; ".join(parts) if parts else "Критичных controlled limitations по доступным outputs не найдено."


def _overall_limitation(procurement_count: int, contract_count: int, missing_inputs: Sequence[str], warnings: Sequence[str]) -> str:
    if not procurement_count and not contract_count:
        return "Нет достаточного registry input для вывода."
    parts = [f"закупки: {procurement_count}", f"контракты: {contract_count}"]
    if missing_inputs:
        parts.append("missing inputs: " + ", ".join(missing_inputs))
    if warnings:
        parts.append(f"warnings: {len(warnings)}")
    return "; ".join(parts)


_TECHNICAL_CODE_LABELS: dict[str, str] = {
    "controlled_warnings_present": "Предупреждения зафиксированы в данных",
    "row_level_warnings_present": "Предупреждения на уровне строк данных",
    "manual_review_required": "Требуется ручная проверка",
    "ktru_not_found": "КТРУ не найден в источниках",
    "extra_chars_risk": "Риск дополнительных характеристик КТРУ",
    "partial_data": "Данные получены частично",
    "normative_check_incomplete": "Нормативная проверка не завершена",
}


def _human_warning_text(value: Any) -> str:
    """Extract a human-readable string from a warning item (never returns raw JSON)."""
    if value is None:
        return ""
    if isinstance(value, str):
        stripped = value.strip()
        if stripped in _TECHNICAL_CODE_LABELS:
            return _TECHNICAL_CODE_LABELS[stripped]
        if stripped.startswith("missing_input:"):
            return f"Отсутствует: {stripped[14:]}"
        return stripped
    if isinstance(value, Mapping):
        for key in ("message", "label", "text", "description", "code"):
            candidate = value.get(key)
            if candidate and isinstance(candidate, str):
                return candidate.strip()
        parts = [f"{k}: {v}" for k, v in value.items() if v and not isinstance(v, (dict, list))]
        return "; ".join(parts)[:200]
    if isinstance(value, (list, tuple)):
        return "; ".join(_human_warning_text(item) for item in value if _human_warning_text(item))[:200]
    return str(value).strip()


def _truncate_text(text: str, max_len: int = 380) -> str:
    """Truncate text to max_len characters, appending ellipsis if cut."""
    return text if len(text) <= max_len else text[:max_len] + "…"


def _aggregate_warnings(warnings: Sequence[Any], max_lines: int = 6) -> str:
    """Produce a compact human-readable summary from a warnings list (max max_lines items)."""
    clean: list[str] = []
    for w in warnings:
        text = _human_warning_text(w)
        if text and text not in clean:
            clean.append(text)
    if not clean:
        return ""
    shown = clean[:max_lines]
    result = "; ".join(shown)
    if len(clean) > max_lines:
        result += f" (ещё {len(clean) - max_lines})"
    return result


def _linked_contracts_count(rows: Sequence[Mapping[str, Any]]) -> int:
    return sum(
        1
        for row in rows
        if _text(row.get("user_contract_link_status")) == "Подтвержден контрактом"
        or "confirmed" in _text(row.get("technical_match_status")).casefold()
    )


def _demand_confirmation_text(linked_count: int, procurement_count: int) -> str:
    if linked_count:
        return f"Подтвержден контрактом: {linked_count}"
    if procurement_count:
        return "Закупочный спрос без контракта"
    return "Не подтвержден"


def _okpd_ktru_text(okpd2_count: int, ktru_count: int) -> str:
    if okpd2_count or ktru_count:
        return f"ОКПД2 найден: {okpd2_count}; КТРУ найден: {ktru_count}"
    return "ОКПД2 / КТРУ не найдены"


def _count_text(label: str, count: int) -> str:
    return f"{label}: {count}"


def _examples(rows: Sequence[Mapping[str, Any]], limit: int = 3) -> str:
    values = []
    for row in rows[:limit]:
        value = (
            _text(row.get("procurement_number"))
            or _text(row.get("purchase_number"))
            or _text(row.get("contract_registry_number"))
            or _text(row.get("contract_number"))
            or _text(row.get("row_id"))
        )
        if value:
            values.append(value)
    return "; ".join(values) if values else "Нет примеров"


def _subject_sample(rows: Sequence[Mapping[str, Any]]) -> str:
    subjects = [_subject(row) for row in rows if _subject(row)]
    if not subjects:
        return "Недостаточно данных"
    return "; ".join(subjects[:2])


def _subject(row: Mapping[str, Any]) -> str:
    return _text(row.get("subject") or row.get("name") or row.get("purchase_object") or row.get("contract_subject"))


def _law_source_text(rows: Sequence[Mapping[str, Any]]) -> str:
    laws = sorted({_text(row.get("law") or row.get("law_type")) for row in rows if _text(row.get("law") or row.get("law_type"))})
    sources = sorted({_text(row.get("source") or row.get("platform") or row.get("source_file")) for row in rows if _text(row.get("source") or row.get("platform") or row.get("source_file"))})
    parts = []
    if laws:
        parts.append("законы: " + ", ".join(laws))
    if sources:
        parts.append("источники: " + ", ".join(sources[:3]))
    return "; ".join(parts) if parts else "Недостаточно данных"


def _enrichment_status_text(rows: Sequence[Mapping[str, Any]]) -> str:
    if not rows:
        return "Обновленные сведения не найдены"
    counts = Counter(_text(row.get("status_label") or row.get("status")) for row in rows)
    return "; ".join(f"{status}: {count}" for status, count in counts.items() if status)


def _mandatory_ktru_count(rows: Sequence[Mapping[str, Any]]) -> int:
    return sum(
        1
        for row in rows
        if "обяз" in _text(row.get("mandatory_status") or row.get("ktru_status") or row.get("service_flag")).casefold()
        or _text(row.get("service_flag")) == "ktru_mandatory_on_notice_date"
    )


def _mandatory_status(readiness: Mapping[str, Any], normative: Mapping[str, Any], has_ktru: bool) -> str:
    raw = _text(
        readiness.get("mandatory_status")
        or readiness.get("obligatory_status")
        or normative.get("mandatory_status")
        or normative.get("ktru_status")
    )
    if raw:
        return raw
    if has_ktru:
        return "Требуется ручная проверка КТРУ"
    return "КТРУ не проверено"


def _characteristics_status(readiness: Mapping[str, Any], normative: Mapping[str, Any]) -> str:
    raw = _text(readiness.get("characteristics_status") or normative.get("characteristics_status"))
    if raw:
        return raw
    if readiness.get("has_characteristics") is True or normative.get("has_characteristics") is True:
        return "Да"
    if readiness.get("has_characteristics") is False or normative.get("has_characteristics") is False:
        return "Нет"
    return "Не проверено"


def _row_evidence(row: Mapping[str, Any]) -> str:
    parts = []
    for key in ("source_file", "source_sheet", "source_row_number"):
        value = _text(row.get(key))
        if value:
            parts.append(value)
    return " / ".join(parts) if parts else _text(row.get("row_id")) or "normalized_registry"


def _index_by_row_id(rows: Sequence[Mapping[str, Any]]) -> dict[str, Mapping[str, Any]]:
    result: dict[str, Mapping[str, Any]] = {}
    for row in rows:
        row_id = _text(row.get("row_id"))
        if row_id and row_id not in result:
            result[row_id] = row
    return result


def _find_related_row(source_row: Mapping[str, Any], rows: Sequence[Mapping[str, Any]]) -> Mapping[str, Any]:
    if not rows:
        return {}
    source_row_id = _text(source_row.get("row_id"))
    if source_row_id:
        for row in rows:
            row_ids = {
                _text(row.get("row_id")),
                _text(row.get("procurement_row_id")),
                _text(row.get("source_row_id")),
                _text(row.get("registry_row_id")),
            }
            if source_row_id in row_ids:
                return row

    source_procurement_numbers = {
        _text(source_row.get("procurement_number")),
        _text(source_row.get("purchase_number")),
        _text(source_row.get("notice_number")),
        _text(source_row.get("registry_number")),
    }
    source_procurement_numbers.discard("")
    if source_procurement_numbers:
        for row in rows:
            row_numbers = {
                _text(row.get("procurement_number")),
                _text(row.get("purchase_number")),
                _text(row.get("notice_number")),
                _text(row.get("registry_number")),
            }
            row_numbers.discard("")
            if source_procurement_numbers & row_numbers:
                return row
    return {}


def _find_related_contract_row(
    procurement_row: Mapping[str, Any],
    contract_rows: Sequence[Mapping[str, Any]],
    contract_link_row: Mapping[str, Any],
) -> Mapping[str, Any]:
    if not contract_rows:
        return {}
    linked_contract_numbers = {
        _text(contract_link_row.get("contract_registry_number")),
        _text(contract_link_row.get("contract_number")),
    }
    linked_contract_numbers.discard("")
    if linked_contract_numbers:
        for row in contract_rows:
            row_contract_numbers = {
                _text(row.get("contract_registry_number")),
                _text(row.get("contract_number")),
                _text(row.get("registry_number")),
            }
            row_contract_numbers.discard("")
            if linked_contract_numbers & row_contract_numbers:
                return row

    procurement_numbers = {
        _text(procurement_row.get("procurement_number")),
        _text(procurement_row.get("purchase_number")),
        _text(contract_link_row.get("procurement_number")),
        _text(contract_link_row.get("purchase_number")),
    }
    procurement_numbers.discard("")
    if procurement_numbers:
        for row in contract_rows:
            row_procurement_numbers = {
                _text(row.get("procurement_number")),
                _text(row.get("purchase_number")),
                _text(row.get("notice_number")),
            }
            row_procurement_numbers.discard("")
            if procurement_numbers & row_procurement_numbers:
                return row
    return {}


def _source_files(rows: Sequence[Mapping[str, Any]]) -> str:
    values = sorted({_text(row.get("source_file")) for row in rows if _text(row.get("source_file"))})
    return "; ".join(values[:6])


def _date_range(rows: Sequence[Mapping[str, Any]]) -> str:
    values = sorted(
        {
            _text(row.get("date") or row.get("publication_date") or row.get("notice_date") or row.get("contract_date"))
            for row in rows
            if _text(row.get("date") or row.get("publication_date") or row.get("notice_date") or row.get("contract_date"))
        }
    )
    if not values:
        return "Не определен"
    if len(values) == 1:
        return values[0]
    return f"{values[0]} - {values[-1]}"


def _dataset_counts(payload: Mapping[str, Any], rows: Sequence[Mapping[str, Any]]) -> dict[str, int]:
    value = payload.get("dataset_counts") if isinstance(payload, Mapping) else None
    if isinstance(value, Mapping):
        return {str(key): int(count) for key, count in value.items() if isinstance(count, int)}
    counter = Counter(_dataset_type(row) or "unknown" for row in rows)
    return dict(counter)


def _dataset_type(row: Mapping[str, Any]) -> str:
    text = _text(row.get("source_dataset_type"))
    if text in {"procurement", "contract"}:
        return text
    if _text(row.get("contract_registry_number")) or _text(row.get("contract_number")):
        return "contract"
    return "procurement" if _text(row.get("procurement_number") or row.get("purchase_number")) else "unknown"


def _collect_warning_texts(*groups: Sequence[Mapping[str, Any]]) -> list[str]:
    result: list[str] = []
    for group in groups:
        for row in group:
            for value in _list(row.get("warnings")) + _list(row.get("limitations")) + _list(row.get("manual_review_flags")):
                text = _human_warning_text(value)
                if text and text not in result:
                    result.append(text)
            if bool(row.get("manual_review_required")) and "manual_review_required" not in result:
                result.append("manual_review_required")
    return result


def _validate_safe_wording(payload: Any) -> dict[str, Any]:
    serialized = json.dumps(payload, ensure_ascii=False).casefold()
    found = [word for word in FORBIDDEN_EXPORT_WORDING if word in serialized]
    return {
        "status": "ok" if not found else "error_check",
        "is_valid": not found,
        "forbidden_phrases_found": found,
    }


def _gpt_handoff_markdown(payload: Mapping[str, Any], technical_payload: Mapping[str, Any]) -> str:
    summary = _mapping(payload.get("summary"))
    paths = _mapping(payload.get("paths"))
    lines = [
        "# EXPORT_DATA_HANDOFF",
        "",
        "ETAP 32.1 реализован как пользовательское действие «Экспорт данных».",
        "Экспорт доступен только после обновления данных и нормативной проверки.",
        "Файл не является юридическим заключением.",
        "",
        f"- status: `{payload.get('status')}`",
        f"- created_at: `{payload.get('created_at')}`",
        f"- control rows: `{summary.get('control_rows_count')}`",
        f"- KTRU detail rows: `{summary.get('ktru_detail_rows_count')}`",
        f"- warnings: `{summary.get('warnings_count')}`",
        f"- Excel: `{paths.get('xlsx')}`",
        "",
        "## Excel sheets",
        "",
    ]
    for sheet in payload.get("excel_sheets", []):
        lines.append(f"- {sheet}")
    lines.extend(["", "## Основные колонки", ""])
    for index, column in enumerate(EXPORT_DATA_COLUMNS, start=1):
        lines.append(f"{index}. {column}")
    lines.append("")
    lines.append(f"Technical rows: `{len(_items(technical_payload, 'items'))}`.")
    return "\n".join(lines) + "\n"


def _export_report_markdown(payload: Mapping[str, Any]) -> str:
    summary = _mapping(payload.get("summary"))
    paths = _mapping(payload.get("paths"))
    lines = [
        "# EXPORT_DATA_REPORT",
        "",
        f"Status: `{payload.get('status')}`",
        "",
        "Экспорт данных формирует аналитическую выгрузку по результатам проверки листов 4 и 5.",
        "Экспорт не является юридическим заключением.",
        "",
        "## Outputs",
        "",
        f"- Excel: `{paths.get('xlsx')}`",
        f"- CSV: `{paths.get('csv')}`",
        f"- JSON: `{paths.get('json')}`",
        f"- Technical JSON: `{paths.get('technical_readiness_json')}`",
        "",
        "## Summary",
        "",
        f"- control_rows_count: `{summary.get('control_rows_count')}`",
        f"- export_columns_count: `{summary.get('export_columns_count')}`",
        f"- ktru_detail_rows_count: `{summary.get('ktru_detail_rows_count')}`",
        f"- registry_rows_count: `{summary.get('registry_rows_count')}`",
        f"- procurement_rows_count: `{summary.get('procurement_rows_count')}`",
        f"- contract_rows_count: `{summary.get('contract_rows_count')}`",
        f"- enrichment_rows_count: `{summary.get('enrichment_rows_count')}`",
        f"- normative_rows_count: `{summary.get('normative_rows_count')}`",
        f"- warnings_count: `{summary.get('warnings_count')}`",
        f"- missing_inputs_count: `{summary.get('missing_inputs_count')}`",
        "",
        "## Safety",
        "",
        "- Исходные Excel/CSV/ZIP/DOCX не изменялись сервисом.",
        "- Runtime/live-download/browser/cookies/tokens/passwords не используются сервисом.",
        "- Outputs пишутся только внутри активного пользовательского проекта.",
        "- Основная таблица агрегирует контрольные узлы и не является сырым полным реестром.",
    ]
    return "\n".join(lines) + "\n"


def _export_xlsx_name(date_label: str) -> str:
    return f"Экспорт_данных_проверка_листов_4_5_{date_label}.xlsx"


def _load_json(path: Path) -> dict[str, Any]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    return dict(payload) if isinstance(payload, Mapping) else {"items": payload}


def _load_json_if_exists(path: Path) -> dict[str, Any] | list[Any] | None:
    if not path.is_file():
        return None
    payload = json.loads(path.read_text(encoding="utf-8"))
    if isinstance(payload, (dict, list)):
        return payload
    return None


def _write_json(path: Path, payload: Mapping[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def _items(payload: Any, key: str) -> list[dict[str, Any]]:
    if isinstance(payload, Mapping):
        value = payload.get(key)
        if isinstance(value, list):
            return [_mapping(item) for item in value]
    if isinstance(payload, list):
        return [_mapping(item) for item in payload]
    return []


def _items_any(payload: Any) -> list[dict[str, Any]]:
    if isinstance(payload, list):
        return [_mapping(item) for item in payload]
    if isinstance(payload, Mapping):
        for key in ("items", "rows", "ui_rows", "data"):
            value = payload.get(key)
            if isinstance(value, list):
                return [_mapping(item) for item in value]
    return []


def _mapping(value: Any) -> dict[str, Any]:
    if isinstance(value, Mapping):
        return dict(value)
    if hasattr(value, "to_json_dict"):
        data = value.to_json_dict()
        return dict(data) if isinstance(data, Mapping) else {}
    return {}


def _mapping_or_none(value: Any) -> dict[str, Any] | None:
    if isinstance(value, Mapping):
        return dict(value)
    return None


def _list_of_mappings(value: Any) -> list[dict[str, Any]]:
    if isinstance(value, list):
        return [_mapping(item) for item in value]
    if isinstance(value, tuple):
        return [_mapping(item) for item in value]
    return []


def _list(value: Any) -> list[Any]:
    if value is None:
        return []
    if isinstance(value, list):
        return value
    if isinstance(value, tuple):
        return list(value)
    if isinstance(value, set):
        return list(value)
    return [value]


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, (list, tuple, set)):
        return "; ".join(_text(item) for item in value if _text(item))
    if isinstance(value, Mapping):
        return json.dumps(value, ensure_ascii=False, sort_keys=True)
    return str(value).strip()


def _safe_cell_text(value: Any) -> str:
    text = _text(value)
    if not text:
        return ""
    sanitized = "".join(char if char in "\t\n\r" or ord(char) >= 32 else " " for char in text)
    return sanitized[:32767]


def _join_values(values: Sequence[Any]) -> str:
    return "; ".join(_text(value) for value in values if _text(value))


def _int(value: Any) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return 0


def _csv_value(value: Any) -> str:
    if isinstance(value, (list, tuple, set)):
        return "; ".join(_text(item) for item in value)
    if isinstance(value, Mapping):
        return json.dumps(value, ensure_ascii=False, sort_keys=True)
    return _text(value)


def _utc_iso(now_provider: Callable[[], datetime]) -> str:
    value = now_provider()
    if value.tzinfo is None:
        value = value.replace(tzinfo=timezone.utc)
    return value.astimezone(timezone.utc).replace(microsecond=0).isoformat()


def _relative_to_project(path: Path, project_root: Path) -> str:
    try:
        return str(path.resolve().relative_to(project_root.resolve()))
    except ValueError:
        return str(path)


def _export_data_evidence_label(value: str) -> str:
    return str(value).replace(
        NORMATIVE_SHEET6_READINESS_JSON_NAME,
        f"{EXPORT_DATA_NORMATIVE_READINESS_EVIDENCE}.json",
    )
